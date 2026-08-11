from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook

from .db import now_iso


MAX_EMPLOYEE_WORKBOOK_BYTES = 10 * 1024 * 1024
EMPLOYEE_SHEET_NAME = "员工数据"
MAPPING_SOURCE_LEVEL2 = "employee_info.level2_department"
MAPPING_SOURCE_LEVEL3 = "employee_info.level3_department"
LEVEL3_SPLIT_DEPARTMENTS = {"凌翔/星铭供应链及职能中心"}


class EmployeeDepartmentError(ValueError):
    pass


@dataclass(frozen=True)
class EmployeeDepartmentRecord:
    user_id: str
    employee_name: str
    second_level_department: str
    third_level_department: str = ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_key(value: Any) -> str:
    return re.sub(r"[\s_\-（）()]+", "", _text(value)).lower()


HEADER_ALIASES = {
    "user_id": {"员工userid", "用户id", "userid", "员工id"},
    "employee_name": {"姓名", "员工姓名", "name"},
    "second_level_department": {"2级部门", "二级部门", "二级组织", "二级部门名称"},
    "third_level_department": {"3级部门", "三级部门", "三级组织", "三级部门名称"},
}


def _column_indexes(rows: list[list[Any]]) -> tuple[int, Dict[str, int]]:
    for row_index, row in enumerate(rows[:20]):
        normalized = {_header_key(value): index for index, value in enumerate(row) if _header_key(value)}
        indexes: Dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            match = next((normalized[alias] for alias in aliases if alias in normalized), None)
            if match is not None:
                indexes[field] = match
        if {"user_id", "employee_name", "second_level_department"}.issubset(indexes):
            return row_index, indexes
    raise EmployeeDepartmentError("员工表缺少“员工UserID、姓名、2级部门”列")


def _xlsx_rows(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise EmployeeDepartmentError("无法读取员工信息 Excel") from exc
    try:
        worksheet = workbook[EMPLOYEE_SHEET_NAME] if EMPLOYEE_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _xls_rows(content: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise EmployeeDepartmentError("服务器缺少旧版 XLS 读取组件") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        worksheet = workbook.sheet_by_name(EMPLOYEE_SHEET_NAME) if EMPLOYEE_SHEET_NAME in workbook.sheet_names() else workbook.sheet_by_index(0)
        return [worksheet.row_values(index) for index in range(worksheet.nrows)]
    except Exception as exc:
        raise EmployeeDepartmentError("无法读取员工信息 Excel") from exc


def parse_employee_department_workbook(content: bytes, filename: str) -> Dict[str, Any]:
    if not content:
        raise EmployeeDepartmentError("员工信息文件不能为空")
    if len(content) > MAX_EMPLOYEE_WORKBOOK_BYTES:
        raise EmployeeDepartmentError("员工信息文件不能超过 10MB")
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".xls", ".xlsx"}:
        raise EmployeeDepartmentError("员工信息文件只支持 XLS 或 XLSX")
    rows = _xls_rows(content) if suffix == ".xls" else _xlsx_rows(content)
    if not rows:
        raise EmployeeDepartmentError("员工信息文件没有数据")
    header_row, indexes = _column_indexes(rows)
    records: list[EmployeeDepartmentRecord] = []
    skipped_no_name = 0
    skipped_no_department = 0
    for row in rows[header_row + 1 :]:
        employee_name = _text(row[indexes["employee_name"]]) if indexes["employee_name"] < len(row) else ""
        user_id = _text(row[indexes["user_id"]]) if indexes["user_id"] < len(row) else ""
        department = _text(row[indexes["second_level_department"]]) if indexes["second_level_department"] < len(row) else ""
        if not employee_name:
            skipped_no_name += 1
            continue
        if not department:
            skipped_no_department += 1
            continue
        third_level_department = (
            _text(row[indexes["third_level_department"]])
            if indexes.get("third_level_department") is not None
            and indexes["third_level_department"] < len(row)
            else ""
        )
        records.append(EmployeeDepartmentRecord(user_id, employee_name, department, third_level_department))
    if not records:
        raise EmployeeDepartmentError("员工信息文件中没有可用的2级部门映射")
    unique_records: list[EmployeeDepartmentRecord] = []
    seen: set[tuple[str, str, str]] = set()
    for record in records:
        key = (
            record.user_id,
            record.employee_name,
            record.second_level_department,
            record.third_level_department,
        )
        if key not in seen:
            unique_records.append(record)
            seen.add(key)
    name_departments: Dict[str, set[str]] = {}
    for record in unique_records:
        name_departments.setdefault(record.employee_name, set()).add(assigned_department(record))
    return {
        "records": unique_records,
        "filename": Path(filename or "员工信息").name,
        "file_hash": hashlib.sha256(content).hexdigest(),
        "total_rows": max(0, len(rows) - header_row - 1),
        "imported_rows": len(unique_records),
        "skipped_no_name": skipped_no_name,
        "skipped_no_department": skipped_no_department,
        "ambiguous_names": sorted(name for name, departments in name_departments.items() if len(departments) > 1),
        "departments": sorted({assigned_department(record) for record in unique_records}),
    }


def assigned_department(record: EmployeeDepartmentRecord | Dict[str, Any]) -> str:
    if isinstance(record, EmployeeDepartmentRecord):
        second_level = _text(record.second_level_department)
        third_level = _text(record.third_level_department)
    else:
        second_level = _text(record.get("second_level_department"))
        third_level = _text(record.get("third_level_department"))
    if second_level in LEVEL3_SPLIT_DEPARTMENTS and third_level:
        return third_level
    return second_level


def replace_employee_department_mappings(
    conn,
    records: Iterable[EmployeeDepartmentRecord],
    *,
    filename: str,
    file_hash: str,
    actor_id: int,
) -> int:
    timestamp = now_iso()
    materialized = list(records)
    conn.execute("DELETE FROM employee_department_mappings")
    conn.executemany(
        """
        INSERT INTO employee_department_mappings (
            user_id, employee_name, second_level_department, third_level_department,
            source_file, source_file_hash, imported_by, imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                record.user_id or None,
                record.employee_name,
                record.second_level_department,
                record.third_level_department or None,
                filename,
                file_hash,
                actor_id,
                timestamp,
            )
            for record in materialized
        ],
    )
    return len(materialized)


def resolve_employee_department(
    conn,
    *,
    applicant_id: Any = None,
    applicant_name: Any = None,
) -> tuple[Optional[Dict[str, str]], str]:
    normalized_id = _text(applicant_id)
    normalized_name = _text(applicant_name)
    if normalized_id:
        rows = conn.execute(
            """
            SELECT user_id, employee_name, second_level_department, third_level_department
            FROM employee_department_mappings
            WHERE TRIM(COALESCE(user_id, '')) = ?
            ORDER BY id
            """,
            (normalized_id,),
        ).fetchall()
        departments = {assigned_department(dict(row)) for row in rows if assigned_department(dict(row))}
        if len(departments) == 1 and rows:
            row = rows[0]
            return {
                "user_id": str(row["user_id"] or "").strip(),
                "employee_name": str(row["employee_name"] or "").strip(),
                "second_level_department": str(row["second_level_department"] or "").strip(),
                "third_level_department": str(row["third_level_department"] or "").strip(),
                "assigned_department": next(iter(departments)),
            }, "user_id"
        if len(departments) > 1:
            return None, "ambiguous"
    if not normalized_name:
        return None, "missing_applicant"
    rows = conn.execute(
        """
        SELECT user_id, employee_name, second_level_department, third_level_department
        FROM employee_department_mappings
        WHERE TRIM(employee_name) = ?
        ORDER BY id
        """,
        (normalized_name,),
    ).fetchall()
    departments = {assigned_department(dict(row)) for row in rows if assigned_department(dict(row))}
    if len(departments) == 1 and rows:
        row = rows[0]
        return {
            "user_id": str(row["user_id"] or "").strip(),
            "employee_name": str(row["employee_name"] or "").strip(),
            "second_level_department": str(row["second_level_department"] or "").strip(),
            "third_level_department": str(row["third_level_department"] or "").strip(),
            "assigned_department": next(iter(departments)),
        }, "employee_name"
    if len(departments) > 1:
        return None, "ambiguous"
    return None, "unmatched"


def request_applicant_identity(data: Dict[str, Any]) -> tuple[str, str]:
    raw_extra = data.get("raw_extra")
    if not isinstance(raw_extra, dict) and data.get("raw_extra_json"):
        try:
            raw_extra = json.loads(data.get("raw_extra_json") or "{}")
        except (TypeError, json.JSONDecodeError):
            raw_extra = {}
    raw_extra = raw_extra if isinstance(raw_extra, dict) else {}
    external_source = raw_extra.get("external_source")
    external_source = external_source if isinstance(external_source, dict) else {}
    applicant_name = _text(data.get("applicant"))
    external_name = _text(external_source.get("applicant"))
    external_id = _text(external_source.get("applicant_id"))
    # A person edited in the cashier system takes precedence over the original
    # DingTalk identity.  Otherwise use the stable DingTalk user id first.
    if applicant_name and external_name and applicant_name != external_name:
        return "", applicant_name
    return external_id, applicant_name or external_name


def apply_employee_department_mapping(
    conn,
    data: Dict[str, Any],
) -> tuple[Dict[str, Any], Optional[Dict[str, str]], str]:
    applicant_id, applicant_name = request_applicant_identity(data)
    mapping, match_source = resolve_employee_department(
        conn,
        applicant_id=applicant_id,
        applicant_name=applicant_name,
    )
    if not mapping:
        return dict(data), None, match_source
    result = dict(data)
    current_sheet = _text(result.get("source_sheet"))
    assignment_source = (
        MAPPING_SOURCE_LEVEL3
        if mapping.get("third_level_department")
        and mapping.get("assigned_department") == mapping.get("third_level_department")
        else MAPPING_SOURCE_LEVEL2
    )
    result["source_sheet"] = mapping["assigned_department"]
    raw_extra = result.get("raw_extra")
    if not isinstance(raw_extra, dict) and result.get("raw_extra_json"):
        try:
            raw_extra = json.loads(result.get("raw_extra_json") or "{}")
        except (TypeError, json.JSONDecodeError):
            raw_extra = {}
    raw_extra = dict(raw_extra) if isinstance(raw_extra, dict) else {}
    external_source = raw_extra.get("external_source")
    if isinstance(external_source, dict):
        external_source = dict(external_source)
        external_source["applicant_department_level2"] = mapping["second_level_department"]
        external_source["applicant_department_level3"] = mapping.get("third_level_department") or None
        external_source["sheet_assignment_source"] = assignment_source
        raw_extra["external_source"] = external_source
    raw_extra["employee_department_mapping"] = {
        "user_id": mapping["user_id"] or None,
        "employee_name": mapping["employee_name"],
        "second_level_department": mapping["second_level_department"],
        "third_level_department": mapping.get("third_level_department") or None,
        "assigned_department": mapping["assigned_department"],
        "match_source": match_source,
        "source": assignment_source,
        "previous_source_sheet": current_sheet or None,
    }
    result["raw_extra"] = raw_extra
    result["raw_extra_json"] = json.dumps(raw_extra, ensure_ascii=False, default=str)
    return result, mapping, match_source
