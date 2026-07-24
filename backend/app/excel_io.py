from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CORE_FIELDS = {
    "dingding_id": "钉钉申请单号",
    "applicant": "申请人",
    "payment_account": "付款账户",
    "expense_type": "费用性质",
    "summary": "摘要",
    "style_name": "款式",
    "amount": "应付金额",
    "paid_amount": "已支付金额",
    "pending_amount": "待付款金额",
    "project": "项目归属",
    "bu": "BU归属",
    "payee_account": "收款账户",
    "payee_name": "账户名",
    "bank_name": "开户行",
    "invoice_status": "开票情况",
    "needed_payment_date": "需求付款日期",
    "owner_confirmation": "负责人确认",
    "finance_review": "财务审批",
    "finance_manager_approval": "财务主管审批",
    "general_manager_approval": "总经理审批",
    "general_manager_approval_date": "总经理审批时间",
    "general_manager_opinion": "总经理意见",
    "actual_payment_date": "财务付款时间",
    "remark": "备注",
    "overdue_status": "逾期情况",
    "payer": "付款人",
}

TARGET_FIELDS = list(CORE_FIELDS.keys())

FULL_DATE_RE = re.compile(r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})")
PAID_TEXT_RE = re.compile(r"(?:已经|已)\s*(?:付(?:款)?|支付|打款|转账?|转)|付讫|付款完成|打款完成")
DATED_PAID_RE = re.compile(r"(20\d{2}[./-]\d{1,2}[./-]\d{1,2})\s*(?:(?:已经|已))?\s*(?:付(?:款)?|支付|打款|转账?|转)")
LEGACY_PAYMENT_STATUS_OPTIONS = {"未支付", "部分付款", "已支付"}
FINANCE_REVIEW_OPTIONS = {"未付款", "部分付款", "已付款"}
UNPAID_REVIEW_VALUES = {"未选择", "待批付", "待付款", "未支付", "未付款"}
PAID_REVIEW_VALUES = {"已支付", "已付款", "已付", "付讫"}
PARTIAL_REVIEW_VALUES = {"部分付款", "部分支付", "部分已付", "已部分付款", "已部分支付"}
GENERAL_MANAGER_APPROVAL_OPTIONS = {"同意付款", "延缓批付", "存在争议", "无需审批"}
NON_FINANCE_PAYMENT_RE = re.compile(r"(垫付|Tiffany|T总|总经理|个人|老板|代付)")
IMAGE_FORMATS = {
    "png": (".png", "image/png"),
    "jpeg": (".jpeg", "image/jpeg"),
    "jpg": (".jpg", "image/jpeg"),
    "gif": (".gif", "image/gif"),
    "bmp": (".bmp", "image/bmp"),
    "webp": (".webp", "image/webp"),
}


KNOWN_HEADER_ALIASES = {
    "dingding_id": ["钉钉申请单号", "审批编号", "审批单号", "申请单号", "流程编号"],
    "applicant": ["申请人", "请款人", "提交人", "申请人姓名"],
    "payment_account": ["付款账户", "付款账号类型"],
    "expense_type": ["费用性质", "费用类型", "款项类型"],
    "summary": ["摘要", "支付节点明细", "事由", "付款事由", "申请内容", "明细"],
    "style_name": ["款式", "产品款式"],
    "amount": ["应付金额", "金额", "付款金额", "申请金额", "报销金额"],
    "paid_amount": ["已支付金额", "已付款金额", "已付金额", "累计支付金额"],
    "pending_amount": ["待付款金额", "未付款金额", "剩余应付金额", "剩余付款金额"],
    "project": ["项目归属", "项目", "所属项目"],
    "bu": ["BU归属", "BU"],
    "payee_account": ["收款信息", "收款账户", "账号", "银行账号", "收款账号"],
    "payee_name": ["账户名", "收款人", "收款方", "供应商名称", "户名"],
    "bank_name": ["开户行", "收款银行", "银行名称"],
    "invoice_status": ["开票情况", "发票情况", "票据情况"],
    "needed_payment_date": ["需求付款日期", "期望付款日期", "付款日期"],
    "owner_confirmation": ["负责人确认", "业务确认"],
    "finance_review": ["财务审核", "财务审批"],
    "finance_manager_approval": ["财务主管审批"],
    "general_manager_approval": ["总经理批复", "总经理确认", "总经理审批"],
    "general_manager_approval_date": ["总经理审批时间", "总经理批复时间", "总经理确认时间", "总经理审批日期", "总经理批复日期"],
    "general_manager_opinion": ["总经理意见", "总经理审批意见", "总经理批复意见"],
    "actual_payment_date": ["实际付款日期", "财务付款时间", "财务付款日期"],
    "remark": ["备注", "审核意见"],
    "overdue_status": ["逾期情况"],
    "payer": ["付款人"],
}


SHEET_HEADERS = {
    "all": [
        "钉钉申请单号",
        "申请人",
        "付款账户",
        "费用性质",
        "摘要",
        "应付金额",
        "已支付金额",
        "待付款金额",
        "项目归属",
        "BU归属",
        "收款信息",
        "账户名",
        "开户行",
        "开票情况",
        "需求付款日期",
        "负责人确认",
        "财务审批",
        "财务付款时间",
        "财务主管审批",
        "总经理审批",
        "总经理审批时间",
        "总经理意见",
        "备注",
        "逾期情况",
        "付款人",
        "来源 Sheet",
        "请款标识",
    ],
    "main": ["钉钉申请单号", "付款账户", "费用性质", "摘要", "应付金额", "已支付金额", "待付款金额", "项目归属", "收款信息", "开票情况", "需求付款日期", "财务审批", "财务付款时间", "总经理审批", "总经理审批时间", "总经理意见", "备注"],
    "mold": ["序号", "钉钉申请单号", "付款账户", "款式", "支付节点明细", "应付金额", "已支付金额", "待付款金额", "账户名", "需求付款日期", "财务审批", "财务付款时间", "备注", "财务主管审批", "总经理审批", "总经理审批时间", "总经理意见"],
    "yuewei": ["钉钉申请单号", "付款账户", "费用性质", "摘要", "应付金额", "已支付金额", "待付款金额", "项目归属", "账号", "账户名", "开户行", "开票情况", "需求付款日期", "负责人确认", "财务审批", "财务付款时间", "备注", "逾期情况", "总经理审批", "总经理审批时间", "总经理意见", "备注", "付款人"],
    "logistics": ["钉钉申请单号", "付款账户", "费用性质", "摘要", "应付金额", "已支付金额", "待付款金额", "BU归属", "项目归属", "账号", "账户名", "开户行", "开票情况", "需求付款日期", "负责人确认", "财务审批", "财务付款时间", "备注", "总经理意见"],
    "hr": ["钉钉申请单号", "付款账户", "费用性质", "摘要", "应付金额", "已支付金额", "待付款金额", "账户名", "项目", "开票情况", "需求付款日期", "财务主管审批", "总经理审批", "总经理审批时间", "总经理意见"],
    "reimburse": ["钉钉申请单号", "付款账户", "费用性质", "摘要", "应付金额", "已支付金额", "待付款金额", "开票情况", "收款账户", "项目归属", "备注", "财务审批", "财务付款时间", "财务主管审批", "总经理审批", "总经理审批时间", "总经理意见"],
    "growth": ["钉钉申请单号", "付款账户", "费用性质", "摘要", "应付金额", "已支付金额", "待付款金额", "收款账户", "备注", "财务审批", "财务付款时间", "财务主管审批", "总经理审批", "总经理审批时间", "总经理意见"],
}

for _headers in SHEET_HEADERS.values():
    if "申请人" not in _headers:
        _headers.insert(_headers.index("钉钉申请单号") + 1, "申请人")
    if "请款标识" not in _headers:
        _headers.append("请款标识")

PAYMENT_DETAIL_SHEET = "付款明细"
ALL_EXPORT_SHEET = "全部"
PAYMENT_DETAIL_HEADERS = [
    "付款标识",
    "请款标识",
    "钉钉单号",
    "来源 Sheet",
    "付款日期",
    "本次金额",
    "付款人",
    "付款账户",
    "流水号",
    "备注",
    "来源标记",
    "凭证信息",
]
SYSTEM_INFO_SHEET = "_系统信息"
EXPORT_FORMAT_VERSION = 2


def parse_batch_dates(filename: str) -> Tuple[Optional[str], Optional[str], str]:
    stem = Path(filename).stem
    match = re.search(r"(20\d{6})\s*[~\-至到]\s*(20\d{6})", stem)
    if not match:
        return None, None, stem or "新请款批次"
    start = datetime.strptime(match.group(1), "%Y%m%d").date()
    end = datetime.strptime(match.group(2), "%Y%m%d").date()
    return start.isoformat(), end.isoformat(), stem


def sheet_name_for_summary(end_date: Optional[str]) -> str:
    if not end_date:
        return "汇总可支付"
    dt = datetime.fromisoformat(end_date).date()
    return f"汇总{dt.month}.{dt.day}可支付"


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def stringify(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else None


def date_string(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def strict_date_string(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def strict_date_from_match(text: str) -> Optional[str]:
    match = FULL_DATE_RE.search(text)
    if not match:
        return None
    year, month, day = match.groups()
    try:
        return date(int(year), int(month), int(day)).isoformat()
    except ValueError:
        return None


def paid_signal(value: Any) -> bool:
    text = stringify(value) or ""
    if not text:
        return False
    return bool(PAID_TEXT_RE.search(text) or DATED_PAID_RE.search(text))


def paid_date_from_texts(*values: Any) -> Optional[str]:
    for value in values:
        text = stringify(value) or ""
        if not text:
            continue
        dated_paid_match = DATED_PAID_RE.search(text)
        if dated_paid_match:
            parsed = strict_date_from_match(dated_paid_match.group(1))
            if parsed:
                return parsed
        if paid_signal(text):
            parsed = strict_date_from_match(text)
            if parsed:
                return parsed
    return None


def append_texts(row: Dict[str, Any], field: str, *values: Any) -> None:
    existing = stringify(row.get(field)) or ""
    parts = [part.strip() for part in existing.split("\n") if part.strip()]
    for value in values:
        text = stringify(value)
        if not text:
            continue
        normalized = text.strip()
        if not normalized or normalized == "0":
            continue
        if normalized in existing or normalized in parts:
            continue
        parts.append(normalized)
        existing = "\n".join(parts)
    row[field] = "\n".join(parts) if parts else None


def append_remark_texts(row: Dict[str, Any], *values: Any) -> None:
    append_texts(row, "remark", *values)


def append_general_manager_opinions(row: Dict[str, Any], *values: Any) -> None:
    append_texts(row, "general_manager_opinion", *values)


def is_non_finance_payment_note(value: Any) -> bool:
    text = stringify(value) or ""
    return bool(text and NON_FINANCE_PAYMENT_RE.search(text))


def normalize_finance_review_value(value: Any) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    text = stringify(value)
    if not text:
        return None, None, None, None
    if strict_date_string(text):
        return None, None, None, strict_date_string(text)
    if text in FINANCE_REVIEW_OPTIONS:
        return text, None, text, None
    if text in UNPAID_REVIEW_VALUES:
        return "未付款", None, "未付款", None
    if is_non_finance_payment_note(text):
        return None, text, None, None
    if is_explicit_partial_payment_status(text):
        note = None if text in PARTIAL_REVIEW_VALUES else text
        return "部分付款", note, "部分付款", paid_date_from_texts(text)
    if text in PAID_REVIEW_VALUES or paid_signal(text):
        normalized = "部分付款" if "部分" in text else "已付款"
        note = None if text in PAID_REVIEW_VALUES else text
        return normalized, note, normalized, paid_date_from_texts(text)
    return None, text, None, None


def normalize_general_manager_approval_value(value: Any) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    text = stringify(value)
    if not text or text == "0":
        return None, None, None
    if strict_date_string(text):
        return None, None, strict_date_string(text)
    if text in GENERAL_MANAGER_APPROVAL_OPTIONS:
        return text, None, None
    if text == "同意支付":
        return "同意付款", None, None
    if "存在争议" in text:
        note = text if text != "存在争议" else None
        return "存在争议", note, None
    if "延缓批付" in text:
        note = text if text != "延缓批付" else None
        return "延缓批付", note, None
    if "同意付款" in text:
        note = text if text != "同意付款" else None
        return "同意付款", note, None
    return None, text, None


def is_explicit_paid_status(value: Any) -> bool:
    text = stringify(value) or ""
    if not text or text in {"同意支付", "同意付款"}:
        return False
    if is_explicit_partial_payment_status(text):
        return False
    return text in PAID_REVIEW_VALUES or paid_signal(text)


def is_explicit_partial_payment_status(value: Any) -> bool:
    text = stringify(value) or ""
    if not text:
        return False
    return text in PARTIAL_REVIEW_VALUES or (
        ("部分" in text or "一部分" in text) and paid_signal(text)
    )


def finance_review_from_legacy_payment_status(value: Any) -> Optional[str]:
    text = stringify(value) or ""
    if not text:
        return None
    if text in UNPAID_REVIEW_VALUES or text in {"同意支付", "同意付款"}:
        return "未付款"
    if is_non_finance_payment_note(text):
        return None
    if is_explicit_partial_payment_status(text):
        return "部分付款"
    if is_explicit_paid_status(text):
        return "已付款"
    return None


def normalize_request_business_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    finance_review_raw = row.get("finance_review")
    manager_approval_raw = row.get("general_manager_approval")
    payment_status_raw = row.get("payment_status")

    finance_review, finance_note, finance_status_hint, finance_date = normalize_finance_review_value(finance_review_raw)
    manager_approval, manager_note, manager_date = normalize_general_manager_approval_value(manager_approval_raw)
    legacy_status_hint = finance_review_from_legacy_payment_status(payment_status_raw)

    row["general_manager_approval"] = manager_approval

    if not strict_date_string(row.get("actual_payment_date")):
        row["actual_payment_date"] = finance_date or paid_date_from_texts(payment_status_raw)
    if manager_date and not strict_date_string(row.get("general_manager_approval_date")):
        row["general_manager_approval_date"] = manager_date

    append_remark_texts(row, finance_note)
    append_general_manager_opinions(row, manager_note)
    status_note = stringify(payment_status_raw)
    if status_note and status_note not in LEGACY_PAYMENT_STATUS_OPTIONS:
        append_remark_texts(row, f"原付款情况：{status_note}")

    has_payment_date = bool(strict_date_string(row.get("actual_payment_date")))
    final_finance_review = finance_status_hint or finance_review or legacy_status_hint
    if not final_finance_review:
        final_finance_review = "已付款" if has_payment_date else "未付款"
    if final_finance_review not in FINANCE_REVIEW_OPTIONS:
        final_finance_review = "未付款"

    amount = amount_number(row.get("amount"))
    paid_amount = amount_number(row.get("paid_amount"))
    if paid_amount is None:
        paid_amount = amount if final_finance_review == "已付款" and amount is not None else 0.0
    pending_amount = round(amount - paid_amount, 2) if amount is not None else None
    if amount is not None and paid_amount > 0:
        final_finance_review = "已付款" if pending_amount <= 0 else "部分付款"

    row["amount"] = amount
    row["paid_amount"] = paid_amount
    row["pending_amount"] = pending_amount
    row["finance_review"] = final_finance_review
    row["payment_status"] = None
    return row


def amount_number(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    text = text.replace(",", "").replace("¥", "").replace("￥", "").replace("$", "")
    try:
        return float(text)
    except ValueError:
        return None


def content_hash(row: Dict[str, Any]) -> str:
    parts = [
        stringify(row.get("dingding_id")) or "",
        str(row.get("amount") or ""),
        stringify(row.get("payee_account")) or stringify(row.get("payee_name")) or "",
        (stringify(row.get("summary")) or "")[:120],
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def detect_header_row(ws) -> Optional[int]:
    max_row = min(ws.max_row, 12)
    for row_idx in range(1, max_row + 1):
        headers = [normalize_header(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if "应付金额" in headers and ("钉钉申请单号" in headers or "序号" in headers):
            return row_idx
    return None


def build_header_columns(ws, header_row: int) -> List[Tuple[int, str]]:
    columns = []
    for col in range(1, ws.max_column + 1):
        value = stringify(ws.cell(header_row, col).value)
        if value:
            columns.append((col, value))
    return columns


def first_value(values_by_header: Dict[str, List[Any]], names: Iterable[str]) -> Any:
    for name in names:
        for key, values in values_by_header.items():
            if normalize_header(key) == normalize_header(name) and values:
                for value in values:
                    if value not in (None, ""):
                        return value
    return None


def collect_values(ws_values, ws_formulas, row_idx: int, header_columns: List[Tuple[int, str]]) -> Tuple[Dict[str, List[Any]], Dict[str, Any]]:
    values_by_header: Dict[str, List[Any]] = defaultdict(list)
    raw_extra: Dict[str, Any] = {}
    seen: Dict[str, int] = defaultdict(int)
    for col, header in header_columns:
        value = ws_values.cell(row_idx, col).value
        formula = ws_formulas.cell(row_idx, col).value
        values_by_header[header].append(value)
        seen[header] += 1
        key = header if seen[header] == 1 else f"{header}#{seen[header]}"
        raw_extra[key] = stringify(value)
        if isinstance(formula, str) and formula.startswith("="):
            raw_extra[f"{key}__formula"] = formula
    return values_by_header, raw_extra


def row_from_sheet(
    ws_values,
    ws_formulas,
    row_idx: int,
    header_columns: List[Tuple[int, str]],
    source_sheet: str,
) -> Optional[Dict[str, Any]]:
    values_by_header, raw_extra = collect_values(ws_values, ws_formulas, row_idx, header_columns)
    finance_review_raw = first_value(values_by_header, ["财务审核", "财务审批"])
    finance_payment_raw = first_value(values_by_header, ["财务付款时间", "财务付款日期", "实际付款日期"])
    finance_payment_date = strict_date_string(finance_payment_raw) or strict_date_string(finance_review_raw)
    finance_review = stringify(finance_review_raw)
    if strict_date_string(finance_review_raw):
        finance_review = None
    manager_approval_raw = first_value(values_by_header, ["总经理批复", "总经理确认", "总经理审批"])
    manager_opinion_raw = first_value(values_by_header, ["总经理意见", "总经理审批意见", "总经理批复意见"])
    manager_approval_date_raw = first_value(values_by_header, ["总经理审批时间", "总经理批复时间", "总经理确认时间", "总经理审批日期", "总经理批复日期"])
    manager_approval_date = strict_date_string(manager_approval_date_raw) or strict_date_string(manager_approval_raw)
    manager_approval = stringify(manager_approval_raw)
    if strict_date_string(manager_approval_raw):
        manager_approval = None
    row: Dict[str, Any] = {
        "_request_id": integer_value(first_value(values_by_header, ["请款标识"])),
        "dingding_id": stringify(first_value(values_by_header, ["钉钉申请单号"])),
        "applicant": stringify(first_value(values_by_header, ["申请人", "请款人", "提交人"])),
        "payment_account": stringify(first_value(values_by_header, ["付款账户"])),
        "expense_type": stringify(first_value(values_by_header, ["费用性质"])),
        "summary": stringify(first_value(values_by_header, ["摘要", "支付节点明细"])),
        "style_name": stringify(first_value(values_by_header, ["款式"])),
        "amount": amount_number(first_value(values_by_header, ["应付金额"])),
        "paid_amount": amount_number(first_value(values_by_header, ["已支付金额", "已付款金额", "已付金额"])),
        "pending_amount": amount_number(first_value(values_by_header, ["待付款金额", "未付款金额", "剩余应付金额"])),
        "project": stringify(first_value(values_by_header, ["项目归属", "项目"])),
        "bu": stringify(first_value(values_by_header, ["BU归属"])),
        "payee_account": stringify(first_value(values_by_header, ["收款信息", "收款账户", "账号"])),
        "payee_name": stringify(first_value(values_by_header, ["账户名"])),
        "bank_name": stringify(first_value(values_by_header, ["开户行"])),
        "invoice_status": stringify(first_value(values_by_header, ["开票情况"])),
        "needed_payment_date": date_string(first_value(values_by_header, ["需求付款日期"])),
        "owner_confirmation": stringify(first_value(values_by_header, ["负责人确认"])),
        "finance_review": finance_review,
        "finance_manager_approval": stringify(first_value(values_by_header, ["财务主管审批"])),
        "general_manager_approval": manager_approval,
        "general_manager_approval_date": manager_approval_date,
        "general_manager_opinion": stringify(manager_opinion_raw),
        "actual_payment_date": finance_payment_date,
        "remark": merge_texts(values_by_header.get("备注", [])),
        "payment_status": stringify(first_value(values_by_header, ["付款情况"])),
        "overdue_status": stringify(first_value(values_by_header, ["逾期情况"])),
        "payer": stringify(first_value(values_by_header, ["付款人"])),
        "source_sheet": source_sheet,
        "source_row": row_idx,
        "raw_extra_json": json.dumps(raw_extra, ensure_ascii=False, default=str),
        "_present_fields": present_request_fields(header_columns),
    }
    formula_amount = raw_extra.get("应付金额__formula")
    if row["amount"] is None and formula_amount and re.match(r"^=[0-9\s+\-*/().]+$", formula_amount):
        try:
            row["amount"] = float(eval(formula_amount[1:], {"__builtins__": {}}, {}))
        except Exception:
            pass
    normalize_request_business_fields(row)
    if not should_keep_row(row, formula_amount):
        return None
    row["content_hash"] = content_hash(row)
    return row


def merge_texts(values: Iterable[Any]) -> Optional[str]:
    texts = [stringify(value) for value in values]
    texts = [text for text in texts if text]
    return "\n".join(texts) if texts else None


def should_keep_row(row: Dict[str, Any], amount_formula: Optional[str]) -> bool:
    if row.get("amount") is None and not row.get("_request_id"):
        return False
    if not row.get("_request_id") and not row.get("dingding_id") and not row.get("payment_account"):
        return False
    has_business_text = any(
        row.get(key)
        for key in ("dingding_id", "payment_account", "expense_type", "summary", "payee_account", "payee_name")
    )
    if not has_business_text:
        return False
    if not row.get("dingding_id") and row.get("summary") in {"私户", "公户", "个人卡", "美金户"}:
        return False
    if not row.get("dingding_id") and amount_formula and re.match(r"^=[A-Z]+\$?\d+$", amount_formula):
        return False
    return True


def integer_value(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def present_request_fields(header_columns: List[Tuple[int, str]]) -> List[str]:
    normalized_headers = {normalize_header(header) for _, header in header_columns}
    present: List[str] = []
    for field, aliases in KNOWN_HEADER_ALIASES.items():
        if any(normalize_header(alias) in normalized_headers for alias in aliases):
            present.append(field)
    return present


def image_format_info(image: ExcelImage) -> Tuple[str, str]:
    image_format = (getattr(image, "format", "") or "").lower()
    if not image_format:
        image_format = Path(str(getattr(image, "path", ""))).suffix.lower().lstrip(".")
    return IMAGE_FORMATS.get(image_format, (".png", "image/png"))


def image_anchor_position(image: ExcelImage) -> Optional[Tuple[int, int]]:
    try:
        marker = image.anchor._from
        return marker.row + 1, marker.col + 1
    except Exception:
        return None


def collect_embedded_images(ws, rows_by_source_row: Dict[int, Dict[str, Any]]) -> Dict[str, int]:
    images = list(getattr(ws, "_images", []) or [])
    summary = {"found": len(images), "attached": 0, "skipped": 0}
    per_row_counts: Dict[int, int] = defaultdict(int)
    safe_sheet = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", ws.title).strip("_") or "sheet"
    for image in images:
        position = image_anchor_position(image)
        if not position:
            summary["skipped"] += 1
            continue
        row_idx, col_idx = position
        parsed_row = rows_by_source_row.get(row_idx)
        if not parsed_row:
            summary["skipped"] += 1
            continue
        try:
            data = image._data()
        except Exception:
            summary["skipped"] += 1
            continue
        if not data:
            summary["skipped"] += 1
            continue
        per_row_counts[row_idx] += 1
        image_index = per_row_counts[row_idx]
        extension, mime_type = image_format_info(image)
        parsed_row.setdefault("_embedded_images", []).append(
            {
                "data": data,
                "extension": extension,
                "mime_type": mime_type,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_col": col_idx,
                "label": f"Excel图片 {ws.title}!{row_idx}#{image_index}",
                "filename": f"excel_{safe_sheet}_{row_idx}_{image_index}{extension}",
            }
        )
        summary["attached"] += 1
    return summary


def parse_weekly_excel(path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    wb_values = load_workbook(path, data_only=True, read_only=False)
    wb_formulas = load_workbook(path, data_only=False, read_only=False)
    rows: List[Dict[str, Any]] = []
    sheet_summaries = []
    image_summary = {"found": 0, "attached": 0, "skipped": 0}
    payment_details: List[Dict[str, Any]] = []
    payment_detail_sheet_present = False
    system_info = parse_system_info(wb_values)
    sheet_name_map = system_info.get("sheet_name_map") or {}
    view_sheets = set(system_info.get("view_sheets") or [])
    workbook_sheet_order: List[str] = []
    for ws_values in wb_values.worksheets:
        ws_formulas = wb_formulas[ws_values.title]
        if normalize_header(ws_values.title) == normalize_header(SYSTEM_INFO_SHEET):
            continue
        if normalize_header(ws_values.title) == normalize_header(PAYMENT_DETAIL_SHEET):
            payment_detail_sheet_present = True
            parsed_details, detail_summary = parse_payment_detail_sheet(ws_values)
            payment_details.extend(parsed_details)
            sheet_summaries.append(detail_summary)
            continue
        if ws_values.title in view_sheets:
            sheet_summaries.append({"sheet": ws_values.title, "imported": 0, "skipped": "view_sheet"})
            continue
        full_source_sheet = str(sheet_name_map.get(ws_values.title) or ws_values.title)
        workbook_sheet_order.append(full_source_sheet)
        header_row = detect_header_row(ws_values)
        if not header_row:
            image_counts = collect_embedded_images(ws_values, {})
            for key in image_summary:
                image_summary[key] += image_counts[key]
            sheet_summaries.append({"sheet": ws_values.title, "imported": 0, "skipped": "no_header", "images": image_counts})
            continue
        header_columns = build_header_columns(ws_values, header_row)
        imported = 0
        rows_by_source_row: Dict[int, Dict[str, Any]] = {}
        for row_idx in range(header_row + 1, ws_values.max_row + 1):
            parsed = row_from_sheet(ws_values, ws_formulas, row_idx, header_columns, full_source_sheet)
            if parsed:
                imported += 1
                rows_by_source_row[row_idx] = parsed
                rows.append(parsed)
        image_counts = collect_embedded_images(ws_values, rows_by_source_row)
        for key in image_summary:
            image_summary[key] += image_counts[key]
        sheet_summaries.append({"sheet": full_source_sheet, "workbook_sheet": ws_values.title, "imported": imported, "header_row": header_row, "images": image_counts})
    return rows, {
        "sheets": sheet_summaries,
        "images": image_summary,
        "payment_detail_sheet_present": payment_detail_sheet_present,
        "payment_details": payment_details,
        "format_version": system_info.get("format_version"),
        "export_batch_id": system_info.get("batch_id"),
        "export_batch_updated_at": system_info.get("batch_updated_at"),
        "exported_at": system_info.get("exported_at"),
        "sheet_name_map": sheet_name_map,
        "workbook_sheet_order": workbook_sheet_order,
    }


def parse_system_info(workbook) -> Dict[str, Any]:
    matching = next(
        (
            worksheet
            for worksheet in workbook.worksheets
            if normalize_header(worksheet.title) == normalize_header(SYSTEM_INFO_SHEET)
        ),
        None,
    )
    if matching is None:
        return {}
    info: Dict[str, Any] = {"sheet_name_map": {}, "view_sheets": []}
    for row_idx in range(1, matching.max_row + 1):
        key = stringify(matching.cell(row_idx, 1).value)
        value = stringify(matching.cell(row_idx, 2).value)
        if key in {"format_version", "batch_id", "batch_updated_at", "exported_at"}:
            info[key] = integer_value(value) if key in {"format_version", "batch_id"} else value
        if key == "sheet_map":
            workbook_name = stringify(matching.cell(row_idx, 2).value)
            source_name = stringify(matching.cell(row_idx, 3).value)
            if workbook_name and source_name:
                info["sheet_name_map"][workbook_name] = source_name
        if key == "view_sheet" and value:
            info["view_sheets"].append(value)
    return info


def parse_payment_detail_sheet(ws) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    header_row = None
    header_lookup: Dict[str, int] = {}
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        candidate = {
            normalize_header(ws.cell(row_idx, col).value): col
            for col in range(1, ws.max_column + 1)
            if normalize_header(ws.cell(row_idx, col).value)
        }
        if "本次金额" in candidate and ("请款标识" in candidate or "钉钉单号" in candidate):
            header_row = row_idx
            header_lookup = candidate
            break
    if header_row is None:
        return [], {"sheet": ws.title, "imported": 0, "skipped": "invalid_payment_header", "images": {"found": 0, "attached": 0, "skipped": len(getattr(ws, "_images", []) or [])}}

    def cell_value(row_idx: int, header: str) -> Any:
        col = header_lookup.get(normalize_header(header))
        return ws.cell(row_idx, col).value if col else None

    details: List[Dict[str, Any]] = []
    rows_by_source_row: Dict[int, Dict[str, Any]] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_request_id = cell_value(row_idx, "请款标识")
        raw_payment_id = cell_value(row_idx, "付款标识")
        raw_amount = cell_value(row_idx, "本次金额")
        raw_date = cell_value(row_idx, "付款日期")
        if all(value in (None, "") for value in (raw_request_id, raw_amount, raw_date, cell_value(row_idx, "钉钉单号"))):
            continue
        try:
            request_id = int(raw_request_id) if raw_request_id not in (None, "") else None
        except (TypeError, ValueError):
            request_id = None
        detail = {
            "payment_id": integer_value(raw_payment_id),
            "request_id": request_id,
            "dingding_id": stringify(cell_value(row_idx, "钉钉单号")),
            "source_sheet": stringify(cell_value(row_idx, "来源 Sheet")),
            "payment_date": date_string(raw_date),
            "amount": amount_number(raw_amount),
            "payer": stringify(cell_value(row_idx, "付款人")),
            "payment_account": stringify(cell_value(row_idx, "付款账户")),
            "bank_reference": stringify(cell_value(row_idx, "流水号")),
            "remark": stringify(cell_value(row_idx, "备注")),
            "source_type": stringify(cell_value(row_idx, "来源标记")) or "excel_detail",
            "voucher_info": stringify(cell_value(row_idx, "凭证信息")),
            "source_row": row_idx,
            "_present_fields": [
                field
                for field, header in (
                    ("payment_date", "付款日期"),
                    ("amount", "本次金额"),
                    ("payer", "付款人"),
                    ("payment_account", "付款账户"),
                    ("bank_reference", "流水号"),
                    ("remark", "备注"),
                )
                if normalize_header(header) in header_lookup
            ],
        }
        details.append(detail)
        rows_by_source_row[row_idx] = detail
    image_counts = collect_embedded_images(ws, rows_by_source_row)
    return details, {"sheet": ws.title, "imported": len(details), "header_row": header_row, "images": image_counts}


def detect_table_headers(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        preview = []
        for row in reader:
            preview.append(dict(row))
            if len(preview) >= 5:
                break
        return headers, preview
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    header_row = None
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        values = [stringify(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if sum(1 for value in values if value) >= 3:
            header_row = row_idx
            break
    if header_row is None:
        return [], []
    headers = [stringify(ws.cell(header_row, col).value) or f"列{col}" for col in range(1, ws.max_column + 1)]
    preview = []
    for row_idx in range(header_row + 1, min(ws.max_row, header_row + 5) + 1):
        preview.append({headers[col - 1]: stringify(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)})
    return headers, preview


def suggest_mapping(headers: List[str]) -> Dict[str, str]:
    mapping = {}
    normalized = {normalize_header(header): header for header in headers}
    for field, aliases in KNOWN_HEADER_ALIASES.items():
        for alias in aliases:
            if normalize_header(alias) in normalized:
                mapping[field] = normalized[normalize_header(alias)]
                break
    return mapping


def parse_dingtalk_file(path: Path, mapping: Dict[str, str]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            source_rows = list(csv.DictReader(handle))
    else:
        wb = load_workbook(path, data_only=True, read_only=False)
        ws = wb.worksheets[0]
        headers, _ = detect_table_headers(path)
        header_row = 1
        for row_idx in range(1, min(ws.max_row, 8) + 1):
            probe = [stringify(ws.cell(row_idx, col).value) for col in range(1, len(headers) + 1)]
            if probe == headers:
                header_row = row_idx
                break
        source_rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            item = {headers[col - 1]: ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)}
            if any(value not in (None, "") for value in item.values()):
                source_rows.append(item)
    rows: List[Dict[str, Any]] = []
    for index, source in enumerate(source_rows, start=2):
        row: Dict[str, Any] = {
            "source_sheet": "钉钉导入",
            "source_row": index,
            "raw_extra_json": json.dumps({k: stringify(v) for k, v in source.items()}, ensure_ascii=False, default=str),
        }
        for field, header in mapping.items():
            value = source.get(header)
            if field in {"amount", "paid_amount", "pending_amount"}:
                row[field] = amount_number(value)
            elif field in {"needed_payment_date", "actual_payment_date", "general_manager_approval_date"}:
                row[field] = date_string(value)
            else:
                row[field] = stringify(value)
        if should_keep_row(row, None):
            row["content_hash"] = content_hash(row)
            rows.append(row)
    return rows, {"source_rows": len(source_rows)}


def sheet_kind(sheet_name: str) -> str:
    if "模具" in sheet_name:
        return "mold"
    if "YUEWEI" in sheet_name or "OEM" in sheet_name:
        return "yuewei"
    if "物流" in sheet_name:
        return "logistics"
    if "人事" in sheet_name:
        return "hr"
    if "员工报销" in sheet_name:
        return "reimburse"
    if "成长金" in sheet_name or "待定" in sheet_name:
        return "growth"
    return "main"


def export_workbook(
    batch: Dict[str, Any],
    records: List[Dict[str, Any]],
    attachments: Dict[int, List[Dict[str, Any]]],
    payments: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        source = record.get("source_sheet") or "汇总"
        groups[source].append(record)
    if not groups:
        groups[sheet_name_for_summary(batch.get("end_date"))] = []

    raw_sheet_order = batch.get("sheet_order")
    if not isinstance(raw_sheet_order, list) and batch.get("sheet_order_json"):
        try:
            raw_sheet_order = json.loads(batch["sheet_order_json"])
        except (TypeError, json.JSONDecodeError):
            raw_sheet_order = []
    raw_sheet_order = raw_sheet_order if isinstance(raw_sheet_order, list) else []
    export_order: list[str] = []
    for item in raw_sheet_order:
        source = str(item or "").strip()
        if source and source not in export_order:
            export_order.append(source)
    order_index = {name: index for index, name in enumerate(export_order)}
    ordered_groups = sorted(
        groups.items(),
        key=lambda item: (order_index.get(item[0], len(order_index)), list(groups).index(item[0])),
    )

    used_sheet_titles: set[str] = {PAYMENT_DETAIL_SHEET, SYSTEM_INFO_SHEET}
    all_sheet_title = safe_sheet_title(ALL_EXPORT_SHEET, used_sheet_titles)
    all_ws = wb.create_sheet(all_sheet_title)
    write_sheet(all_ws, "all", ALL_EXPORT_SHEET, records, {})

    sheet_name_map: Dict[str, str] = {}
    for sheet_name, sheet_records in ordered_groups:
        kind = sheet_kind(sheet_name)
        workbook_title = safe_sheet_title(sheet_name, used_sheet_titles)
        ws = wb.create_sheet(workbook_title)
        sheet_name_map[workbook_title] = sheet_name
        write_sheet(ws, kind, sheet_name, sheet_records, attachments)

    payment_ws = wb.create_sheet(PAYMENT_DETAIL_SHEET)
    write_payment_detail_sheet(payment_ws, payments or [])
    write_system_info_sheet(wb, batch, sheet_name_map, [all_sheet_title])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def write_payment_detail_sheet(ws, payments: List[Dict[str, Any]]) -> None:
    image_count = max(
        (
            len(
                [
                    voucher
                    for voucher in payment.get("vouchers", [])
                    if str(voucher.get("mime_type") or "").startswith("image/") and voucher.get("absolute_path")
                ]
            )
            for payment in payments
        ),
        default=0,
    )
    image_headers = [f"凭证图片{index + 1}" for index in range(image_count)]
    headers = PAYMENT_DETAIL_HEADERS + image_headers
    header_fill = PatternFill("solid", fgColor="315B7D")
    thin = Side(style="thin", color="D6DEE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_idx, payment in enumerate(payments, start=2):
        vouchers = payment.get("vouchers", []) or []
        voucher_text = "\n".join(
            f"{voucher.get('original_filename') or voucher.get('label') or '凭证'}: {voucher.get('file_url') or ''}".rstrip(": ")
            for voucher in vouchers
        )
        values = [
            payment.get("id"),
            payment.get("request_id"),
            payment.get("dingding_id"),
            payment.get("request_source_sheet") or payment.get("source_sheet"),
            payment.get("payment_date"),
            payment.get("amount"),
            payment.get("payer"),
            payment.get("payment_account"),
            payment.get("bank_reference"),
            payment.get("remark"),
            payment.get("source_type"),
            voucher_text or None,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row_idx, PAYMENT_DETAIL_HEADERS.index("本次金额") + 1).number_format = '#,##0.00'
        if payment.get("payment_date"):
            ws.cell(row_idx, PAYMENT_DETAIL_HEADERS.index("付款日期") + 1).number_format = "yyyy-mm-dd"
        image_vouchers = [
            voucher
            for voucher in vouchers
            if str(voucher.get("mime_type") or "").startswith("image/") and voucher.get("absolute_path")
        ]
        if image_vouchers:
            ws.row_dimensions[row_idx].height = 78
        for index, voucher in enumerate(image_vouchers[:image_count]):
            col = len(PAYMENT_DETAIL_HEADERS) + index + 1
            cell = ws.cell(row_idx, col)
            cell.border = border
            path = Path(str(voucher["absolute_path"]))
            if not path.exists():
                continue
            try:
                image = ExcelImage(str(path))
            except Exception:
                continue
            scale = min(120 / max(image.width, 1), 92 / max(image.height, 1), 1)
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
            ws.add_image(image, f"{get_column_letter(col)}{row_idx}")
    widths = [14, 14, 22, 22, 14, 14, 16, 18, 22, 30, 18, 42]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(len(PAYMENT_DETAIL_HEADERS) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True


def write_system_info_sheet(
    workbook,
    batch: Dict[str, Any],
    sheet_name_map: Dict[str, str],
    view_sheet_titles: Optional[List[str]] = None,
) -> None:
    ws = workbook.create_sheet(SYSTEM_INFO_SHEET)
    rows = [
        ("format_version", EXPORT_FORMAT_VERSION, None),
        ("batch_id", batch.get("id"), None),
        ("batch_updated_at", batch.get("updated_at"), None),
        ("exported_at", datetime.now().replace(microsecond=0).isoformat(), None),
    ]
    rows.extend(("view_sheet", sheet_title, "read_only_summary") for sheet_title in (view_sheet_titles or []))
    rows.extend(("sheet_map", workbook_name, source_name) for workbook_name, source_name in sheet_name_map.items())
    for row_idx, values in enumerate(rows, start=1):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    ws.sheet_state = "veryHidden"


def safe_sheet_title(title: str, used_titles: Optional[set[str]] = None) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", title).strip()[:31] or "Sheet"
    if used_titles is None:
        return base
    candidate = base
    index = 2
    normalized_used = {value.casefold() for value in used_titles}
    while candidate.casefold() in normalized_used:
        suffix = f"~{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate


def write_sheet(ws, kind: str, sheet_name: str, records: List[Dict[str, Any]], attachments: Dict[int, List[Dict[str, Any]]]) -> None:
    headers = SHEET_HEADERS[kind]
    image_column_count = max((len(image_attachments_for_record(record, attachments)) for record in records), default=0)
    image_headers: List[str] = []
    if image_column_count == 1:
        image_headers = ["图片附件"]
    elif image_column_count > 1:
        image_headers = [f"图片附件{i + 1}" for i in range(image_column_count)]
    export_headers = headers + image_headers
    header_fill = PatternFill("solid", fgColor="1F6F6D")
    title_fill = PatternFill("solid", fgColor="F2C94C")
    light_fill = PatternFill("solid", fgColor="EAF6F3")
    thin = Side(style="thin", color="D6DEE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if kind == "mold":
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.cell(1, 1, value=f"{sheet_name}应付款合计")
        ws.cell(1, 1).fill = title_fill
        ws.cell(1, 1).font = Font(bold=True)
        last_row = max(3, len(records) + 2)
        for header in ("应付金额", "已支付金额", "待付款金额"):
            amount_col = headers.index(header) + 1
            ws.cell(1, amount_col, value=f"=SUM({get_column_letter(amount_col)}3:{get_column_letter(amount_col)}{last_row})")
            ws.cell(1, amount_col).number_format = '#,##0.00'
        header_row = 2
        data_start = 3
    elif kind == "all":
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_headers))
        note = ws.cell(1, 1, value="全部记录汇总视图；如需重新导入并合并更新，请在对应部门 Sheet 中修改。")
        note.fill = title_fill
        note.font = Font(bold=True, color="5A4500")
        note.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24
        header_row = 2
        data_start = 3
    else:
        header_row = 1
        data_start = 2

    for col, header in enumerate(export_headers, start=1):
        cell = ws.cell(header_row, col, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for offset, record in enumerate(records):
        row_idx = data_start + offset
        row_values = values_for_headers(headers, record)
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row_idx, col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if headers[col - 1] in {"应付金额", "已支付金额", "待付款金额"}:
                cell.number_format = '#,##0.00'
            if ("日期" in headers[col - 1] or "时间" in headers[col - 1]) and value:
                cell.number_format = "yyyy-mm-dd"
        if kind == "mold":
            ws.cell(row_idx, 1, value=f"=ROW()-{data_start - 1}")
        if image_headers:
            for col in range(len(headers) + 1, len(export_headers) + 1):
                cell = ws.cell(row_idx, col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        add_attachment_comment(ws, row_idx, headers, record, attachments)
        if image_headers:
            add_attachment_images(ws, row_idx, len(headers) + 1, record, attachments, image_column_count)

    if kind != "mold" and records:
        total_row = data_start + len(records)
        amount_columns = [headers.index(header) + 1 for header in ("应付金额", "已支付金额", "待付款金额")]
        for amount_col in amount_columns:
            ws.cell(total_row, amount_col, value=f"=SUM({get_column_letter(amount_col)}{data_start}:{get_column_letter(amount_col)}{total_row - 1})")
            ws.cell(total_row, amount_col).number_format = '#,##0.00'
        ws.cell(total_row, amount_columns[0] - 1, value="合计")
        for col in range(1, len(export_headers) + 1):
            ws.cell(total_row, col).fill = light_fill
            ws.cell(total_row, col).border = border

    for col, header in enumerate(export_headers, start=1):
        width = 14
        if header in {"摘要", "支付节点明细", "备注", "总经理意见"}:
            width = 36
        elif header in {"钉钉申请单号", "收款账户", "收款信息", "账户名", "开户行"}:
            width = 22
        elif header in {"应付金额", "已支付金额", "待付款金额"}:
            width = 14
        elif header.startswith("图片附件"):
            width = 18
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws["A3"] if kind in {"mold", "all"} else ws["A2"]
    ws.sheet_view.showGridLines = False
    request_id_col = headers.index("请款标识") + 1
    ws.column_dimensions[get_column_letter(request_id_col)].hidden = True


def values_for_headers(headers: List[str], record: Dict[str, Any]) -> List[Any]:
    values = []
    remark_used = False
    for header in headers:
        if header == "序号":
            values.append(None)
        elif header == "请款标识":
            values.append(record.get("id"))
        elif header == "钉钉申请单号":
            values.append(record.get("dingding_id"))
        elif header == "申请人":
            manual_applicant = record.get("applicant")
            external_source = (record.get("raw_extra") or {}).get("external_source") or {}
            values.append(manual_applicant if manual_applicant is not None else external_source.get("applicant"))
        elif header == "付款账户":
            values.append(record.get("payment_account"))
        elif header == "费用性质":
            values.append(record.get("expense_type"))
        elif header in {"摘要", "支付节点明细"}:
            values.append(record.get("summary"))
        elif header == "款式":
            values.append(record.get("style_name"))
        elif header == "应付金额":
            values.append(record.get("amount"))
        elif header == "已支付金额":
            values.append(record.get("paid_amount"))
        elif header == "待付款金额":
            values.append(record.get("pending_amount"))
        elif header in {"项目归属", "项目"}:
            values.append(record.get("project"))
        elif header == "BU归属":
            values.append(record.get("bu"))
        elif header in {"收款账户", "收款信息", "账号"}:
            values.append(record.get("payee_account"))
        elif header == "账户名":
            values.append(record.get("payee_name"))
        elif header == "开户行":
            values.append(record.get("bank_name"))
        elif header == "开票情况":
            values.append(record.get("invoice_status"))
        elif header == "需求付款日期":
            values.append(record.get("needed_payment_date"))
        elif header == "负责人确认":
            values.append(record.get("owner_confirmation"))
        elif header in {"财务审核", "财务审批"}:
            values.append(record.get("finance_review"))
        elif header == "财务主管审批":
            values.append(record.get("finance_manager_approval"))
        elif header in {"总经理批复", "总经理确认", "总经理审批"}:
            values.append(record.get("general_manager_approval"))
        elif header == "总经理审批时间":
            values.append(record.get("general_manager_approval_date"))
        elif header == "总经理意见":
            values.append(record.get("general_manager_opinion"))
        elif header in {"实际付款日期", "财务付款时间"}:
            values.append(record.get("actual_payment_date"))
        elif header == "备注":
            values.append(None if remark_used else record.get("remark"))
            remark_used = True
        elif header == "逾期情况":
            values.append(record.get("overdue_status"))
        elif header == "付款人":
            values.append(record.get("payer"))
        elif header == "来源 Sheet":
            values.append(record.get("source_sheet"))
        else:
            values.append(None)
    return values


def add_attachment_comment(ws, row_idx: int, headers: List[str], record: Dict[str, Any], attachments: Dict[int, List[Dict[str, Any]]]) -> None:
    links = attachments.get(record["id"], [])
    if not links:
        return
    comment_text = "\n".join(attachment_label_text(link) for link in links)
    target_header = "备注" if "备注" in headers else headers[-1]
    col = headers.index(target_header) + 1
    cell = ws.cell(row_idx, col)
    cell.comment = Comment(comment_text, "System")


def image_attachments_for_record(record: Dict[str, Any], attachments: Dict[int, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    return [
        attachment
        for attachment in attachments.get(record.get("id"), [])
        if attachment.get("attachment_type") == "image" and attachment.get("absolute_path")
    ]


def add_attachment_images(
    ws,
    row_idx: int,
    start_col: int,
    record: Dict[str, Any],
    attachments: Dict[int, List[Dict[str, Any]]],
    max_count: int,
) -> None:
    image_links = image_attachments_for_record(record, attachments)[:max_count]
    if not image_links:
        return
    ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, 78)
    for index, link in enumerate(image_links):
        col = start_col + index
        cell = ws.cell(row_idx, col)
        cell.comment = Comment(attachment_label_text(link), "System")
        path = Path(str(link.get("absolute_path") or ""))
        if not path.exists():
            cell.value = attachment_label_text(link)
            continue
        try:
            image = ExcelImage(str(path))
        except Exception:
            cell.value = attachment_label_text(link)
            continue
        max_width = 120
        max_height = 92
        scale = min(max_width / max(image.width, 1), max_height / max(image.height, 1), 1)
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
        ws.add_image(image, f"{get_column_letter(col)}{row_idx}")


def attachment_label_text(link: Dict[str, Any]) -> str:
    label = link.get("label") or link.get("original_filename") or "附件"
    path = link.get("original_filename") or link.get("url_path") or link.get("file_path") or ""
    return f"{label}: {path}" if path else str(label)
