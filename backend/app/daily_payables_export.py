from __future__ import annotations

import os
import tempfile
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .daily_payables import DailyPayablesError


SUMMARY_SHEET = "每日汇总"
DETAIL_SHEET = "逐日明细"
SUPPORTED_CURRENCIES = ("CNY", "USD", "MXN")
EXCEL_MAX_ROWS = 1_048_576


def _positive_env_int(name: str, default: int) -> int:
    try:
        value = int(os.environ.get(name, str(default)))
    except (TypeError, ValueError):
        return default
    return value if value > 0 else default


MAX_EXPORT_DETAIL_ROWS = min(
    _positive_env_int("PAYMENT_DAILY_EXPORT_MAX_ROWS", 250_000),
    EXCEL_MAX_ROWS - 1,
)
MAX_EXPORT_FILE_BYTES = _positive_env_int(
    "PAYMENT_DAILY_EXPORT_MAX_BYTES",
    200 * 1024 * 1024,
)
MAX_CONCURRENT_EXPORTS = _positive_env_int("PAYMENT_DAILY_EXPORT_MAX_CONCURRENT", 2)

SUMMARY_HEADERS = [
    "统计日期",
    "当日到期数量",
    "日终未付数量",
    "逾期数量",
    "当天新增到期（折合人民币）",
    "当日支付（折合人民币）",
    "日终待付（折合人民币）",
    "逾期待付（折合人民币）",
] + [
    f"{currency} {label}"
    for currency in SUPPORTED_CURRENCIES
    for label in ("当天新增到期", "当日支付", "日终待付", "逾期待付")
]

DETAIL_HEADERS = [
    "统计日期",
    "状态",
    "请款标识",
    "钉钉申请单号",
    "应付款公司（来源 Sheet）",
    "申请人",
    "摘要",
    "需求付款日期",
    "应付金额",
    "累计已付",
    "当日支付",
    "日终待付",
    "币种",
    "折合人民币应付金额",
    "折合人民币累计已付",
    "折合人民币当日支付",
    "折合人民币日终待付",
    "审批状态",
    "审批结果",
]


def _header_cells(worksheet, headers: list[str]) -> list[WriteOnlyCell]:
    fill = PatternFill("solid", fgColor="1F6F6D")
    cells: list[WriteOnlyCell] = []
    for value in headers:
        cell = WriteOnlyCell(worksheet, value=value)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cells.append(cell)
    return cells


def _data_cells(
    worksheet,
    values: list[Any],
    *,
    date_columns: set[int],
    money_columns: set[int],
) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for column, value in enumerate(values, start=1):
        if isinstance(value, str):
            value = ILLEGAL_CHARACTERS_RE.sub("", value)
            if value.lstrip(" \t\r\n").startswith(("=", "+", "-", "@")):
                value = f"'{value}"
        cell = WriteOnlyCell(worksheet, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if column in date_columns and value is not None:
            cell.number_format = "yyyy-mm-dd"
        elif column in money_columns:
            cell.number_format = "#,##0.00"
        cells.append(cell)
    return cells


def _set_column_widths(worksheet, widths: list[int]) -> None:
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _discard_write_only_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        try:
            worksheet.close()
        except Exception:
            pass
    for worksheet in workbook.worksheets:
        writer = getattr(worksheet, "_writer", None)
        if writer is None:
            continue
        try:
            writer.cleanup()
        except (FileNotFoundError, ValueError):
            pass


def _date_value(value: Any) -> date | None:
    raw = str(value or "").strip()[:10]
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise DailyPayablesError(
            "INVALID_EXPORT_DATA",
            f"每日应付历史包含无效日期：{raw}",
        ) from exc


def _currency_totals(snapshot: dict[str, Any], currency: str) -> dict[str, float]:
    return next(
        (item for item in snapshot.get("currency_totals", []) if item.get("currency") == currency),
        {"due_today": 0, "paid_today": 0, "end_pending": 0, "overdue_pending": 0},
    )


def _detail_status(item: dict[str, Any]) -> str:
    if float(item.get("paid_today") or 0) > 0 and float(item.get("pending_amount") or 0) <= 0:
        return "当日付清"
    if item.get("is_due_today"):
        return "当日到期"
    if item.get("is_overdue"):
        return "逾期待付"
    return "待付"


def export_daily_payables_workbook(snapshots: Iterable[dict[str, Any]]) -> Path:
    descriptor, raw_path = tempfile.mkstemp(prefix="daily-payables-", suffix=".xlsx")
    os.close(descriptor)
    output_path = Path(raw_path)
    workbook: Workbook | None = None
    try:
        workbook = Workbook(write_only=True)
        summary = workbook.create_sheet(SUMMARY_SHEET)
        detail = workbook.create_sheet(DETAIL_SHEET)
        summary.freeze_panes = "A2"
        detail.freeze_panes = "A2"
        summary.sheet_view.showGridLines = False
        detail.sheet_view.showGridLines = False
        _set_column_widths(summary, [13, 13, 13, 13] + [19] * (len(SUMMARY_HEADERS) - 4))
        _set_column_widths(detail, [13, 12, 14, 24, 24, 18, 42, 15] + [15] * 9 + [18, 18])
        detail.column_dimensions["C"].hidden = True
        summary.append(_header_cells(summary, SUMMARY_HEADERS))
        detail.append(_header_cells(detail, DETAIL_HEADERS))

        summary_rows = 1
        detail_rows = 1
        for snapshot in snapshots:
            selected = _date_value(snapshot.get("date"))
            totals = snapshot["totals_cny"]
            counts = snapshot["counts"]
            summary_values = [
                selected,
                counts["due_today"],
                counts["end_pending"],
                counts["overdue_pending"],
                totals["due_today"],
                totals["paid_today"],
                totals["end_pending"],
                totals["overdue_pending"],
                *[
                    _currency_totals(snapshot, currency)[key]
                    for currency in SUPPORTED_CURRENCIES
                    for key in ("due_today", "paid_today", "end_pending", "overdue_pending")
                ],
            ]
            summary.append(
                _data_cells(
                    summary,
                    summary_values,
                    date_columns={1},
                    money_columns=set(range(5, len(SUMMARY_HEADERS) + 1)),
                )
            )
            summary_rows += 1

            for item in snapshot.get("items", []):
                if detail_rows >= min(EXCEL_MAX_ROWS, MAX_EXPORT_DETAIL_ROWS + 1):
                    raise DailyPayablesError(
                        "EXPORT_TOO_LARGE",
                        "导出明细超过单次容量上限，请缩短日期区间",
                    )
                detail_values = [
                    selected,
                    _detail_status(item),
                    item.get("logical_request_id"),
                    item.get("dingding_id"),
                    item.get("source_sheet"),
                    item.get("applicant"),
                    item.get("summary"),
                    _date_value(item.get("needed_payment_date")),
                    item.get("amount"),
                    item.get("paid_amount"),
                    item.get("paid_today"),
                    item.get("pending_amount"),
                    item.get("currency"),
                    item.get("base_amount_cny"),
                    item.get("base_paid_amount_cny"),
                    item.get("base_paid_today_cny"),
                    item.get("base_pending_amount_cny"),
                    item.get("approval_status"),
                    item.get("approval_result"),
                ]
                detail.append(
                    _data_cells(
                        detail,
                        detail_values,
                        date_columns={1, 8},
                        money_columns={9, 10, 11, 12, 14, 15, 16, 17},
                    )
                )
                detail_rows += 1
        summary.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{summary_rows}"
        detail.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_HEADERS))}{detail_rows}"
        workbook.save(output_path)
        if output_path.stat().st_size > MAX_EXPORT_FILE_BYTES:
            raise DailyPayablesError(
                "EXPORT_TOO_LARGE",
                "导出文件超过服务器容量上限，请缩短日期区间",
            )
        return output_path
    except Exception:
        if workbook is not None:
            _discard_write_only_workbook(workbook)
        output_path.unlink(missing_ok=True)
        raise
