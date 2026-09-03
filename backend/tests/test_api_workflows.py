import base64
import io
import json
import os
import sqlite3
import shutil
import sys
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from threading import Barrier, Event, Lock
from time import sleep
from urllib.parse import unquote

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

TEST_DIR = Path(os.environ["PAYMENT_APP_DB"]).parent

from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from backend.app.db import backup_database, connect, migrate_sheet_registry_and_names, now_iso
from backend.app import db as db_module
from backend.app import external_expenses as external_expenses_module
from backend.app.external_expenses import (
    CHINA_WORKBENCH_REGION_ERROR,
    ExternalExpenseError,
    _external_expense_metadata,
    _parse_workflow_events,
    _preview_conditions,
    approval_result_is_disallowed,
    china_workbench_external_expense_allowed,
    execution_region_is_allowed,
    general_manager_approval_from_workflow_events,
    applicant_name_from_title,
    classify_dingtalk_payment_event,
    mark_china_workbench_external_expense,
    map_monthly_payment,
    map_external_expense,
    _monthly_attachments,
    valid_applicant_name,
)
from backend.app.excel_io import export_workbook
from backend.app import main as main_module
from backend.app import mexico_tracking as mexico_tracking_module
from backend.app.main import app
from backend.app.normalize_payment_data import normalize_payment_data
from backend.app.rebuild_weekly_data import rebuild_weekly_data


SAMPLE = Path("/Users/smk/Downloads/20260626~20260707请款明细.xlsx")


def test_expected_payment_account_schema_is_added_compatibly():
    with TestClient(app) as client:
        login(client)
        with connect() as conn:
            columns = {
                row["name"]
                for row in conn.execute("PRAGMA table_info(payment_requests)").fetchall()
            }

    assert "expected_payment_account" in columns
    assert "expected_payment_account_source" in columns


def test_external_expense_metadata_exposes_mapped_needed_payment_date():
    mapped = {
        "approval_no": "META-DATE-1",
        "source_type": "purchase",
        "source_label": "采购支出",
        "source_id": "9778",
        "needed_payment_date": "2026-07-24",
        "payment_account": "公户",
        "project": "规范项目",
        "request_data": {
            "raw_extra": {
                "external_source": {
                    "approval_status": "RUNNING",
                    "needed_payment_date": "stale-value",
                    "payment_account": "私户",
                    "project": "过期项目",
                }
            }
        },
    }

    metadata = _external_expense_metadata(mapped)

    assert metadata["approval_no"] == "META-DATE-1"
    assert metadata["source_id"] == "9778"
    assert metadata["needed_payment_date"] == "2026-07-24"
    assert metadata["payment_account"] == "公户"
    assert metadata["project"] == "规范项目"


@pytest.mark.parametrize(("invoice_value", "expected_account"), [
    ("是", "公户"), ("有", "公户"), ("有发票", "公户"), ("Sí", "公户"), ("Si", "公户"), ("Yes", "公户"),
    ("  SÍ!  ", "公户"), ("yEs", "公户"),
    ("否", "私户"), ("无", "私户"), ("无发票", "私户"), ("No", "私户"), ("No.", "私户"), ("待定", None), ("", None),
])
def test_external_expense_maps_invoice_choice_and_project(invoice_value, expected_account):
    mapped = map_external_expense({
        "source_type": "operation", "source_id": "invoice-1", "effective_date": "2026-08-01",
        "approval_no": "INVOICE-1", "creator_name": "u1", "applicant_department": "财务部",
        "approval_status": "RUNNING", "approval_result": "agree", "execution_region": "中国China",
        "beneficiary": "收款人", "expense_type": "办公", "base_currency_amount": 10,
        "project": "表格项目", "raw_data": {"formComponentValues": [
            {"name": "是否有发票 Existe Factura", "value": invoice_value},
            {"name": "项目归属 Pertenencia del Proyecto", "value": "表单项目"},
        ]},
    })
    assert mapped["request_data"]["payment_account"] == expected_account
    assert mapped["payment_account"] == expected_account
    assert mapped["project"] == "表单项目"
    external = mapped["request_data"]["raw_extra"]["external_source"]
    assert external["payment_account"] == expected_account
    assert external["project"] == "表单项目"
    assert external["invoice_value"] == invoice_value.strip()


def test_external_expense_maps_case_insensitive_spanish_invoice_prefix():
    mapped = map_external_expense({
        "source_type": "operation", "source_id": "invoice-spanish", "effective_date": "2026-08-01",
        "approval_no": "INVOICE-SPANISH", "approval_status": "RUNNING", "approval_result": "agree",
        "execution_region": "中国China", "beneficiary": "收款人", "base_currency_amount": 10,
        "raw_data": {"formComponentValues": [
            {"name": "EXISTE FACTURA 是否有发票", "value": "No."},
        ]},
    })
    assert mapped["payment_account"] == "私户"
    assert mapped["request_data"]["raw_extra"]["external_source"]["invoice_value"] == "No."


def test_external_expense_skips_blank_first_bilingual_invoice_component():
    mapped = map_external_expense({
        "source_type": "operation", "source_id": "invoice-blank-first", "effective_date": "2026-08-01",
        "approval_no": "INVOICE-BLANK-FIRST", "approval_status": "RUNNING", "approval_result": "agree",
        "execution_region": "中国China", "beneficiary": "收款人", "base_currency_amount": 10,
        "raw_data": {"formComponentValues": [
            {"name": "是否有发票", "value": ""},
            {"name": "EXISTE FACTURA", "value": "Yes"},
        ]},
    })
    external = mapped["request_data"]["raw_extra"]["external_source"]
    assert mapped["payment_account"] == "公户"
    assert external["invoice_value"] == "Yes"


def test_external_expense_decodes_json_invoice_and_project_values():
    mapped = map_external_expense({
        "source_type": "operation", "source_id": "json-values", "effective_date": "2026-08-01",
        "approval_no": "JSON-VALUES", "approval_status": "RUNNING", "approval_result": "agree",
        "execution_region": "中国China", "beneficiary": "收款人", "base_currency_amount": 10,
        "raw_data": {"formComponentValues": [
            {"name": "是否有发票", "value": '["Sí"]'},
            {"name": "项目归属", "value": '["墨西哥新工厂"]'},
        ]},
    })
    external = mapped["request_data"]["raw_extra"]["external_source"]
    assert mapped["payment_account"] == "公户"
    assert mapped["project"] == "墨西哥新工厂"
    assert mapped["request_data"]["project"] == "墨西哥新工厂"
    assert external["project"] == "墨西哥新工厂"
    assert external["invoice_value"] == "Sí"


def test_external_expense_recognizes_label_only_invoice_and_project_components():
    mapped = map_external_expense({
        "source_type": "operation", "source_id": "label-values", "effective_date": "2026-08-01",
        "approval_no": "LABEL-VALUES", "approval_status": "RUNNING", "approval_result": "agree",
        "execution_region": "中国China", "beneficiary": "收款人", "base_currency_amount": 10,
        "raw_data": {"formComponentValues": [
            {"label": "EXISTE FACTURA", "value": "Yes"},
            {"label": "Pertenencia del Proyecto", "value": "标签项目"},
        ]},
    })
    assert mapped["payment_account"] == "公户"
    assert mapped["project"] == "标签项目"


def test_external_expense_project_falls_back_to_source_row_for_purchase_and_monthly():
    for source_type in ("purchase", "monthly"):
        mapped = map_external_expense({
            "source_type": source_type, "source_id": "project-fallback", "effective_date": "2026-08-01",
            "approval_no": "PROJECT-FALLBACK", "approval_status": "RUNNING", "approval_result": "agree",
            "execution_region": "中国China", "beneficiary": "收款人", "base_currency_amount": 10,
            "project": "表格项目", "raw_data": {"formComponentValues": []},
        })
        assert mapped["project"] == "表格项目"
        assert mapped["request_data"]["project"] == "表格项目"
        assert mapped["request_data"]["raw_extra"]["external_source"]["project"] == "表格项目"


def test_monthly_invoice_account_wins_and_legacy_account_falls_back():
    base = {
        "source_id": "monthly-account", "process_instance_id": "proc-account", "create_time": "2026-08-01",
        "status": "RUNNING", "result": "agree", "title": "张三提交的月结付款",
        "raw_payload": {"businessId": "MONTHLY-ACCOUNT", "originatorUserId": "u1",
            "originatorDeptName": "财务部", "formComponentValues": [
                {"name": "合计总额", "value": "10"}, {"name": "收款账户信息", "value": "收款人"},
                {"name": "币种", "value": "人民币"}, {"name": "付款账户类型", "value": "私户"},
            ]},
    }
    recognized = dict(base, raw_payload=dict(base["raw_payload"], formComponentValues=[
        *base["raw_payload"]["formComponentValues"], {"name": "是否有发票 Existe Factura", "value": "Yes"},
    ]))
    mapped = map_monthly_payment(recognized)
    assert mapped["payment_account"] == mapped["request_data"]["payment_account"] == "公户"
    assert mapped["request_data"]["raw_extra"]["external_source"]["payment_account"] == "公户"
    fallback = map_monthly_payment(base)
    assert fallback["payment_account"] == fallback["request_data"]["payment_account"] == "私户"
    assert fallback["request_data"]["raw_extra"]["external_source"]["payment_account"] == "私户"


def test_external_expense_maps_explicit_expected_payment_account_across_layers():
    mapped = map_external_expense({
        "source_type": "operation",
        "source_id": "expected-explicit",
        "effective_date": "2026-09-02",
        "approval_no": "EXPECTED-EXPLICIT",
        "approval_status": "RUNNING",
        "approval_result": "agree",
        "execution_region": "中国China",
        "beneficiary": "收款人",
        "base_currency_amount": 100,
        "raw_data": {"formComponentValues": [
            {
                "name": "预计支付账户Cuenta de pago prevista",
                "value": '[{"name":"悦为智能 6221 主账户"}]',
            },
            {"name": "服务主体Sujeto de servicio", "value": "悦为智能 YW Tech_Ai"},
        ]},
    })

    assert mapped["expected_payment_account"] == "悦为智能 6221 主账户"
    assert mapped["expected_payment_account_source"] == "dingtalk_explicit"
    assert mapped["request_data"]["expected_payment_account"] == "悦为智能 6221 主账户"
    assert mapped["request_data"]["expected_payment_account_source"] == "dingtalk_explicit"
    external = mapped["request_data"]["raw_extra"]["external_source"]
    assert external["expected_payment_account"] == "悦为智能 6221 主账户"
    assert external["expected_payment_account_source"] == "dingtalk_explicit"
    assert external["service_subject"] == "悦为智能 YW Tech_Ai"
    metadata = _external_expense_metadata(mapped)
    assert metadata["expected_payment_account"] == "悦为智能 6221 主账户"
    assert metadata["expected_payment_account_source"] == "dingtalk_explicit"


@pytest.mark.parametrize("source_type", ["operation", "purchase"])
def test_external_expense_defaults_expected_payment_account_from_service_subject(source_type):
    form_values = [
        {"label": "服务主体", "value": '[{"label":"凌翔产品&开发"}]'},
    ]
    if source_type == "purchase":
        form_values.append({"name": "收款人", "value": "采购收款人"})
    mapped = map_external_expense({
        "source_type": source_type,
        "source_id": f"expected-default-{source_type}",
        "effective_date": "2026-09-02",
        "approval_no": f"EXPECTED-DEFAULT-{source_type}",
        "approval_status": "RUNNING",
        "approval_result": "agree",
        "execution_region": "中国China",
        "beneficiary": "运营收款人",
        "base_currency_amount": 100,
        "raw_data": {"formComponentValues": form_values},
    })

    assert mapped["expected_payment_account"] == "凌翔公司账户"
    assert mapped["expected_payment_account_source"] == "service_subject_default"
    assert mapped["request_data"]["expected_payment_account"] == "凌翔公司账户"
    assert mapped["request_data"]["expected_payment_account_source"] == "service_subject_default"
    assert "服务主体无法匹配预计支付账户，请人工填写" not in mapped["warnings"]


def test_external_expense_unknown_service_subject_warns_without_blocking():
    mapped = map_external_expense({
        "source_type": "operation",
        "source_id": "expected-unknown",
        "effective_date": "2026-09-02",
        "approval_no": "EXPECTED-UNKNOWN",
        "approval_status": "RUNNING",
        "approval_result": "agree",
        "execution_region": "中国China",
        "beneficiary": "收款人",
        "base_currency_amount": 100,
        "raw_data": {"formComponentValues": [
            {"name": "服务主体", "value": "新公司"},
        ]},
    })

    assert mapped["expected_payment_account"] is None
    assert mapped["expected_payment_account_source"] is None
    assert "服务主体无法匹配预计支付账户，请人工填写" in mapped["warnings"]
    assert mapped["errors"] == []


def test_monthly_payment_uses_shared_expected_payment_account_mapping():
    mapped = map_monthly_payment({
        "source_id": "monthly-expected",
        "process_instance_id": "proc-monthly-expected",
        "create_time": "2026-09-02",
        "status": "RUNNING",
        "result": "agree",
        "title": "张三提交的月结付款",
        "raw_payload": {
            "businessId": "MONTHLY-EXPECTED",
            "originatorUserId": "u1",
            "originatorDeptName": "财务部",
            "formComponentValues": [
                {"name": "合计总额", "value": "10"},
                {"name": "收款账户信息", "value": "收款人"},
                {"name": "币种", "value": "人民币"},
                {"name": "服务主体", "value": "YW MOLDES MX模具"},
            ],
        },
    })

    assert mapped["expected_payment_account"] == "YW MOLDES公司账户"
    assert mapped["expected_payment_account_source"] == "service_subject_default"
    assert mapped["request_data"]["expected_payment_account"] == "YW MOLDES公司账户"
    assert mapped["request_data"]["expected_payment_account_source"] == "service_subject_default"


def login(client: TestClient, username: str = "admin", password: str = "admin123") -> None:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200


@pytest.fixture(autouse=True)
def legacy_workflow_test_sheets_are_china(request, monkeypatch):
    """Keep pre-region workflow fixtures explicit about their original scope.

    This module's generic workflow cases predate multi-region isolation; their
    invented Sheet names represent China data.  The dedicated isolation case
    keeps production classification unchanged so unknown Sheets remain review.
    """

    if request.node.name == "test_china_region_isolation_filters_workbench_totals_sheets_and_export":
        return
    production_sheet_region = mexico_tracking_module.sheet_region

    def workflow_fixture_region(source_sheet):
        return production_sheet_region(source_sheet) or "china"

    monkeypatch.setattr(mexico_tracking_module, "sheet_region", workflow_fixture_region)
    monkeypatch.setattr(main_module, "sheet_region", workflow_fixture_region)


def test_sqlite_concurrency_pragmas_and_version_columns_are_enabled():
    with TestClient(app):
        with connect() as conn:
            assert conn.execute("PRAGMA journal_mode").fetchone()[0].lower() == "wal"
            assert conn.execute("PRAGMA busy_timeout").fetchone()[0] == 15000
            assert conn.execute("PRAGMA synchronous").fetchone()[0] == 2
            assert conn.execute("PRAGMA foreign_keys").fetchone()[0] == 1
            for table in ("request_batches", "payment_requests", "payment_records"):
                columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
                assert "version" in columns


def test_mexico_user_access_schema_and_admin_validation():
    username = f"mexico-access-{uuid.uuid4().hex}"
    with TestClient(app) as admin_client:
        login(admin_client)
        with connect() as conn:
            columns = {
                row["name"]
                for row in conn.execute("PRAGMA table_info(users)").fetchall()
            }
        assert {"mexico_access_scope", "mexico_identity_name"} <= columns

        missing_identity = admin_client.post(
            "/api/admin/users",
            json={
                "username": f"{username}-missing",
                "password": "Yuewei123",
                "display_name": "缺少绑定",
                "role": "business",
                "mexico_access_scope": "participant",
            },
        )
        assert missing_identity.status_code == 400

        invalid_scope = admin_client.post(
            "/api/admin/users",
            json={
                "username": f"{username}-invalid",
                "password": "Yuewei123",
                "display_name": "非法权限",
                "role": "business",
                "mexico_access_scope": "everyone",
                "mexico_identity_name": "Nobody",
            },
        )
        assert invalid_scope.status_code == 400

        created = admin_client.post(
            "/api/admin/users",
            json={
                "username": username,
                "password": "Yuewei123",
                "display_name": "墨西哥审批参与者",
                "role": "business",
                "mexico_access_scope": "participant",
                "mexico_identity_name": "CHONG.MARTINEZ.DAUL",
            },
        )
        assert created.status_code == 200, created.text
        created_user = created.json()["user"]
        assert created_user["mexico_access_scope"] == "participant"
        assert created_user["mexico_identity_name"] == "CHONG.MARTINEZ.DAUL"

        with TestClient(app) as participant_client:
            login(participant_client, username, "Yuewei123")
            me = participant_client.get("/api/me")
            assert me.status_code == 200
            assert me.json()["user"]["mexico_access_scope"] == "participant"
            assert me.json()["user"]["mexico_identity_name"] == "CHONG.MARTINEZ.DAUL"

        cleared_identity = admin_client.patch(
            f"/api/admin/users/{created_user['id']}",
            json={"mexico_identity_name": ""},
        )
        assert cleared_identity.status_code == 400

        disabled = admin_client.patch(
            f"/api/admin/users/{created_user['id']}",
            json={"mexico_access_scope": "none", "mexico_identity_name": ""},
        )
        assert disabled.status_code == 200
        assert disabled.json()["user"]["mexico_access_scope"] == "none"
        assert disabled.json()["user"]["mexico_identity_name"] is None
        assert admin_client.delete(f"/api/admin/users/{created_user['id']}").status_code == 200


def test_daily_payable_schema_backfills_logical_request_chain_and_one_baseline():
    with TestClient(app):
        with connect() as conn:
            columns = {
                row["name"]
                for row in conn.execute("PRAGMA table_info(payment_requests)").fetchall()
            }
            assert "logical_request_id" in columns
            assert conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'payable_history_versions'"
            ).fetchone()
            assert conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'app_settings'"
            ).fetchone()

            timestamp = now_iso()
            batch_id = conn.execute(
                """
                INSERT INTO request_batches (name, status, created_at, updated_at)
                VALUES ('每日应付链路测试', 'draft', ?, ?)
                """,
                (timestamp, timestamp),
            ).lastrowid
            root_id = conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, amount, paid_amount, pending_amount, currency,
                    needed_payment_date, source_sheet, created_at, updated_at
                ) VALUES (?, 20000, 0, 20000, 'CNY', '2026-08-21', '测试公司', ?, ?)
                """,
                (batch_id, timestamp, timestamp),
            ).lastrowid
            copy_id = conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, copied_from_request_id, amount, paid_amount, pending_amount,
                    currency, needed_payment_date, source_sheet, created_at, updated_at
                ) VALUES (?, ?, 20000, 0, 20000, 'CNY', '2026-08-21', '测试公司', ?, ?)
                """,
                (batch_id, root_id, timestamp, timestamp),
            ).lastrowid

            db_module.ensure_daily_payable_history_schema(conn)

            rows = conn.execute(
                "SELECT id, logical_request_id FROM payment_requests WHERE id IN (?, ?) ORDER BY id",
                (root_id, copy_id),
            ).fetchall()
            assert [row["logical_request_id"] for row in rows] == [root_id, root_id]
            assert conn.execute(
                "SELECT COUNT(*) FROM payable_history_versions WHERE logical_request_id = ?",
                (root_id,),
            ).fetchone()[0] == 1
            assert db_module.get_daily_payables_history_start_date(conn)


def test_sqlite_backup_api_creates_verified_consistent_copy(tmp_path):
    with TestClient(app):
        with connect() as conn:
            expected_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        target = tmp_path / "app-backup.db"
        result = backup_database(target)
        assert result["path"] == str(target)
        assert result["sha256"]
        assert result["size"] > 0
        with sqlite3.connect(target) as backup:
            assert backup.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
            assert backup.execute("SELECT COUNT(*) FROM users").fetchone()[0] == expected_users


def test_stale_request_update_returns_structured_version_conflict():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "请款乐观锁", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        created = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "并发", "summary": "初始", "amount": 100},
        ).json()["request"]
        assert created["version"] == 1

        first = client.patch(
            f"/api/batches/{batch['id']}/requests/{created['id']}",
            json={"summary": "用户甲", "expected_version": created["version"]},
        )
        assert first.status_code == 200
        assert first.json()["request"]["version"] == 2

        stale = client.patch(
            f"/api/batches/{batch['id']}/requests/{created['id']}",
            json={"summary": "用户乙", "expected_version": created["version"]},
        )
        assert stale.status_code == 409
        detail = stale.json()["detail"]
        assert detail["code"] == "VERSION_CONFLICT"
        assert detail["entity_type"] == "payment_request"
        assert detail["entity_id"] == created["id"]
        assert detail["current_version"] == 2


def test_bulk_save_rolls_back_when_any_request_version_is_stale():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "批量乐观锁", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        first = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "并发", "summary": "第一条", "amount": 100},
        ).json()["request"]
        second = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "并发", "summary": "第二条", "amount": 200},
        ).json()["request"]
        updated_first = client.patch(
            f"/api/batches/{batch['id']}/requests/{first['id']}",
            json={"summary": "已被别人修改", "expected_version": first["version"]},
        ).json()["request"]

        response = client.patch(
            f"/api/batches/{batch['id']}/requests/bulk",
            json={
                "creates": [],
                "updates": [
                    {"id": first["id"], "summary": "过期覆盖", "expected_version": first["version"]},
                    {"id": second["id"], "summary": "本不应保存", "expected_version": second["version"]},
                ],
                "deletes": [],
            },
        )
        assert response.status_code == 409
        rows = {
            row["id"]: row
            for row in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
        }
        assert rows[first["id"]]["summary"] == updated_first["summary"]
        assert rows[second["id"]]["summary"] == "第二条"
        assert rows[second["id"]]["version"] == second["version"]


def test_payment_mutation_checks_payment_and_request_versions():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "付款乐观锁", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        request_row = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "并发", "summary": "付款", "amount": 100},
        ).json()["request"]
        created_payment = client.post(
            f"/api/batches/{batch['id']}/requests/{request_row['id']}/payments",
            json={
                "amount": 20,
                "payment_date": "2026-08-17",
                "expected_request_version": request_row["version"],
            },
        )
        assert created_payment.status_code == 200
        payment = created_payment.json()["payment"]
        request_after_create = created_payment.json()["request"]
        assert payment["version"] == 1
        assert request_after_create["version"] == request_row["version"] + 1

        changed = client.patch(
            f"/api/batches/{batch['id']}/requests/{request_row['id']}/payments/{payment['id']}",
            json={
                "amount": 30,
                "expected_request_version": request_after_create["version"],
                "expected_payment_version": payment["version"],
            },
        )
        assert changed.status_code == 200

        stale = client.patch(
            f"/api/batches/{batch['id']}/requests/{request_row['id']}/payments/{payment['id']}",
            json={
                "amount": 40,
                "expected_request_version": request_after_create["version"],
                "expected_payment_version": payment["version"],
            },
        )
        assert stale.status_code == 409
        assert stale.json()["detail"]["code"] == "VERSION_CONFLICT"


def test_stale_batch_structure_update_is_rejected():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "批次乐观锁", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        assert batch["version"] == 1
        first = client.put(
            f"/api/batches/{batch['id']}/sheet-order",
            json={"sheet_order": ["财务"], "expected_batch_version": batch["version"]},
        )
        assert first.status_code == 200
        assert first.json()["batch"]["version"] == 2
        stale = client.put(
            f"/api/batches/{batch['id']}/sheet-order",
            json={"sheet_order": ["采购"], "expected_batch_version": batch["version"]},
        )
        assert stale.status_code == 409
        assert stale.json()["detail"]["code"] == "VERSION_CONFLICT"


def test_batch_operation_lease_prevents_duplicate_long_tasks_and_recovers_expired_lease():
    from backend.app.batch_operations import (
        acquire_batch_operation,
        complete_batch_operation,
    )

    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "长任务互斥", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        with connect() as conn:
            actor = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()["id"]

        first = acquire_batch_operation(batch["id"], "dingtalk-sync", actor)
        with pytest.raises(HTTPException) as conflict:
            acquire_batch_operation(batch["id"], "weekly-excel", actor)
        assert conflict.value.status_code == 409
        assert conflict.value.detail["code"] == "BATCH_OPERATION_IN_PROGRESS"
        assert conflict.value.detail["operation_id"] == first["id"]

        with connect() as conn:
            conn.execute(
                "UPDATE batch_operations SET lease_expires_at = ? WHERE id = ?",
                ("2000-01-01T00:00:00.000000", first["id"]),
            )
        recovered = acquire_batch_operation(batch["id"], "weekly-excel", actor)
        with connect() as conn:
            interrupted = conn.execute(
                "SELECT status FROM batch_operations WHERE id = ?",
                (first["id"],),
            ).fetchone()
        assert interrupted["status"] == "interrupted"
        complete_batch_operation(recovered["id"], {"imported": 3})
        with connect() as conn:
            finished = conn.execute(
                "SELECT status, result_json FROM batch_operations WHERE id = ?",
                (recovered["id"],),
            ).fetchone()
        assert finished["status"] == "succeeded"
        assert json.loads(finished["result_json"])["imported"] == 3


def test_nonblocking_batch_operation_is_reused_and_exposes_progress():
    from backend.app.batch_operations import (
        acquire_or_reuse_batch_operation,
        complete_batch_operation,
        get_batch_operation,
        update_batch_operation_progress,
    )

    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "同步任务复用", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        request_row = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "并发", "summary": "初始", "amount": 100},
        ).json()["request"]
        with connect() as conn:
            actor = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()["id"]

        first, first_reused = acquire_or_reuse_batch_operation(
            batch["id"],
            "dingtalk-workflow-sync",
            actor,
            blocks_writes=False,
        )
        second, second_reused = acquire_or_reuse_batch_operation(
            batch["id"],
            "dingtalk-workflow-sync",
            actor,
            blocks_writes=False,
        )
        assert first_reused is False
        assert second_reused is True
        assert second["id"] == first["id"]

        update_batch_operation_progress(
            first["id"],
            stage="attachment_download",
            progress_current=8,
            progress_total=32,
            progress_message="正在同步 8/32 个附件",
            timings={"metadata_query_ms": 125.4, "workflow_query_ms": 248.1},
            partial_result={"status_committed": True, "updated_requests": 4},
        )
        current = get_batch_operation(first["id"])
        assert current["stage"] == "attachment_download"
        assert current["progress_current"] == 8
        assert current["progress_total"] == 32
        assert current["timings"]["metadata_query_ms"] == 125.4
        assert current["partial_result"]["status_committed"] is True

        edited = client.patch(
            f"/api/batches/{batch['id']}/requests/{request_row['id']}",
            json={"summary": "同步期间仍可编辑", "expected_version": request_row["version"]},
        )
        assert edited.status_code == 200
        complete_batch_operation(first["id"], {"status": "synced"})


def test_dingtalk_task_start_reuses_running_operation_and_can_be_polled(monkeypatch):
    worker_started = Event()
    release_worker = Event()
    worker_calls = []

    def blocked_worker(operation_id, batch_id, user):
        worker_calls.append((operation_id, batch_id, user["id"]))
        worker_started.set()
        assert release_worker.wait(timeout=10)

    monkeypatch.setattr(main_module, "run_dingtalk_sync_task", blocked_worker)

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "后台钉钉任务"}).json()["batch"]
        first = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata",
            params={"task_mode": "true"},
        )
        assert first.status_code == 202
        first_operation = first.json()["operation"]
        assert first_operation["status"] == "running"
        assert worker_started.wait(timeout=10)

        second = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata",
            params={"task_mode": "true"},
        )
        assert second.status_code == 202
        assert second.json()["reused"] is True
        assert second.json()["operation"]["id"] == first_operation["id"]
        assert len(worker_calls) == 1

        polled = client.get(f"/api/batch-operations/{first_operation['id']}")
        assert polled.status_code == 200
        assert polled.json()["operation"]["id"] == first_operation["id"]
        release_worker.set()


def test_dingtalk_task_commits_status_before_slow_attachment_inventory(monkeypatch):
    approval_no = "SYNC-STATUS-FIRST"
    attachment_query_started = Event()
    release_attachment_query = Event()

    monkeypatch.setattr(
        main_module,
        "fetch_external_expense_metadata",
        lambda approval_nos: [{
            "approval_no": approval_no,
            "source_type": "operation",
            "source_label": "运营支出",
            "source_id": "status-first-1",
            "approval_status": "COMPLETED",
            "approval_result": "agree",
            "applicant_id": "status-first-user",
            "applicant": "状态优先申请人",
            "applicant_department": "状态优先部门",
        }],
    )
    monkeypatch.setattr(
        main_module,
        "fetch_dingtalk_workflows",
        lambda approval_nos: [{
            "approval_no": approval_no,
            "process_instance_id": "PROC-STATUS-FIRST",
            "status": "COMPLETED",
            "result": "agree",
            "events": [],
        }],
    )

    def delayed_attachments(approval_nos):
        attachment_query_started.set()
        assert release_attachment_query.wait(timeout=10)
        return []

    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", delayed_attachments)

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "状态优先于附件"}).json()["batch"]
        request_row = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "dingding_id": approval_no,
                "summary": "同步前",
                "amount": 100,
                "source_sheet": "凌翔产品&开发",
            },
        ).json()["request"]
        started = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata",
            params={"task_mode": "true"},
        )
        assert started.status_code == 202
        operation_id = started.json()["operation"]["id"]
        assert attachment_query_started.wait(timeout=10)

        status_row = next(
            row for row in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
            if row["id"] == request_row["id"]
        )
        assert status_row["raw_extra"]["external_source"]["approval_status"] == "COMPLETED"
        operation = client.get(f"/api/batch-operations/{operation_id}").json()["operation"]
        assert operation["partial_result"]["status_committed"] is True
        assert operation["stage"] == "attachment_inventory"

        edited = client.patch(
            f"/api/batches/{batch['id']}/requests/{request_row['id']}",
            json={"summary": "附件查询期间仍可编辑", "expected_version": status_row["version"]},
        )
        assert edited.status_code == 200
        release_attachment_query.set()

        for _ in range(100):
            operation = client.get(f"/api/batch-operations/{operation_id}").json()["operation"]
            if operation["status"] != "running":
                break
            sleep(0.02)
        assert operation["status"] == "succeeded"
        current = next(
            row for row in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
            if row["id"] == request_row["id"]
        )
        assert current["summary"] == "附件查询期间仍可编辑"
        with connect() as conn:
            timing_audit = conn.execute(
                "SELECT new_value_json FROM audit_logs WHERE batch_id = ? AND action = 'external_expenses.sync_timing'",
                (batch["id"],),
            ).fetchone()
        assert timing_audit is not None
        timing_payload = json.loads(timing_audit["new_value_json"])
        assert timing_payload["operation_id"] == operation_id
        assert {
            "metadata_query_ms",
            "workflow_query_ms",
            "status_commit_ms",
            "attachment_query_ms",
            "attachment_commit_ms",
            "total_ms",
        }.issubset(timing_payload["timings"])


def test_dingtalk_attachment_downloads_use_bounded_parallelism(monkeypatch):
    state_lock = Lock()
    active = 0
    maximum_active = 0
    progress = []

    class FakeAttachmentClient:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback):
            return None

        def download(self, process_instance_id, file_id):
            nonlocal active, maximum_active
            with state_lock:
                active += 1
                maximum_active = max(maximum_active, active)
            sleep(0.04)
            with state_lock:
                active -= 1
            return file_id.encode(), "application/pdf"

    monkeypatch.setattr(main_module, "DingtalkAttachmentClient", FakeAttachmentClient)
    monkeypatch.setattr(
        main_module,
        "save_dingtalk_attachment_file",
        lambda attachment, content, content_type: {
            "relative_path": f"attachments/{attachment['file_id']}.pdf",
            "file_name": f"{attachment['file_id']}.pdf",
            "attachment_type": "pdf",
            "mime_type": content_type,
            "file_size": len(content),
        },
    )
    candidates = [
        {
            "approval_no": f"A-{index}",
            "attachment_id": f"ATT-{index}",
            "file_id": f"FILE-{index}",
            "file_name": f"file-{index}.pdf",
            "process_instance_id": f"PROC-{index}",
            "request_ids": [index],
        }
        for index in range(10)
    ]

    downloaded, failed, errors = main_module.download_dingtalk_attachment_candidates(
        candidates,
        max_workers=4,
        progress_callback=lambda current, total: progress.append((current, total)),
    )

    assert len(downloaded) == 10
    assert failed == 0
    assert errors == []
    assert 2 <= maximum_active <= 4
    assert progress[-1] == (10, 10)


def test_batch_structure_and_ordinary_writes_respect_versions_and_active_operation():
    from backend.app.batch_operations import acquire_batch_operation, complete_batch_operation

    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "批次写入互斥", "start_date": "2026-08-17", "end_date": "2026-08-23"},
        ).json()["batch"]
        request_row = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "并发", "summary": "初始", "amount": 100},
        ).json()["request"]
        latest_batch = client.get(f"/api/batches/{batch['id']}").json()["batch"]

        archived = client.post(
            f"/api/batches/{batch['id']}/archive",
            params={"expected_batch_version": latest_batch["version"]},
        )
        assert archived.status_code == 200
        assert archived.json()["batch"]["version"] == latest_batch["version"] + 1
        stale = client.post(
            f"/api/batches/{batch['id']}/unarchive",
            params={"expected_batch_version": latest_batch["version"]},
        )
        assert stale.status_code == 409
        assert stale.json()["detail"]["code"] == "VERSION_CONFLICT"

        current = archived.json()["batch"]
        restored = client.post(
            f"/api/batches/{batch['id']}/unarchive",
            params={"expected_batch_version": current["version"]},
        )
        assert restored.status_code == 200

        with connect() as conn:
            actor = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()["id"]
        operation = acquire_batch_operation(batch["id"], "dingtalk-sync", actor)
        blocked = client.patch(
            f"/api/batches/{batch['id']}/requests/{request_row['id']}",
            json={"summary": "不应写入", "expected_version": request_row["version"]},
        )
        assert blocked.status_code == 409
        assert blocked.json()["detail"]["code"] == "BATCH_OPERATION_IN_PROGRESS"
        complete_batch_operation(operation["id"], {"status": "ok"})


def test_dingtalk_sync_allows_edits_during_external_fetch_and_rejects_stale_commit(monkeypatch):
    approval_no = "SYNC-CONCURRENT-EDIT"
    fetch_started = Event()
    release_fetch = Event()

    def delayed_metadata(approval_nos):
        assert approval_nos == [approval_no]
        fetch_started.set()
        assert release_fetch.wait(timeout=10)
        return [{
            "approval_no": approval_no,
            "source_type": "operation",
            "source_label": "运营支出",
            "source_id": "concurrent-edit-1",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "applicant_id": "user-concurrent",
            "applicant": "并发申请人",
            "applicant_department": "并发测试部",
        }]

    monkeypatch.setattr(main_module, "fetch_external_expense_metadata", delayed_metadata)
    monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", lambda approval_nos: [])
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda approval_nos: [])

    with TestClient(app) as setup_client:
        login(setup_client)
        batch = setup_client.post("/api/batches", json={"name": "同步外部查询不锁编辑"}).json()["batch"]
        request_row = setup_client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": approval_no, "summary": "同步前", "amount": 100},
        ).json()["request"]

    def run_sync():
        with TestClient(app) as sync_client:
            login(sync_client)
            return sync_client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")

    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(run_sync)
        assert fetch_started.wait(timeout=10)
        with TestClient(app) as edit_client:
            login(edit_client)
            edited = edit_client.patch(
                f"/api/batches/{batch['id']}/requests/{request_row['id']}",
                json={"summary": "外部查询期间已编辑", "expected_version": request_row["version"]},
            )
        release_fetch.set()
        synced = future.result(timeout=15)

    assert edited.status_code == 200
    assert synced.status_code == 409
    assert synced.json()["detail"]["code"] == "VERSION_CONFLICT"
    assert synced.json()["detail"]["entity_type"] == "payment_request"
    with TestClient(app) as verify_client:
        login(verify_client)
        current = verify_client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
    assert current["summary"] == "外部查询期间已编辑"
    assert "external_source" not in current["raw_extra"]


def test_external_import_fetch_does_not_lock_batch_and_stale_plan_is_rejected(monkeypatch):
    fetch_started = Event()
    release_fetch = Event()

    def delayed_fetch(keys):
        fetch_started.set()
        assert release_fetch.wait(timeout=10)
        return [external_expense_test_row("IMPORT-CONCURRENT-1", "9001")]

    monkeypatch.setattr(main_module, "fetch_external_expenses", delayed_fetch)

    with TestClient(app) as setup_client:
        login(setup_client)
        batch = setup_client.post("/api/batches", json={"name": "外部导入查询不锁编辑"}).json()["batch"]
        request_row = setup_client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"summary": "原内容", "amount": 100},
        ).json()["request"]

    def run_import():
        with TestClient(app) as import_client:
            login(import_client)
            return import_client.post(
                f"/api/batches/{batch['id']}/imports/external-expenses",
                json={"items": [{"source_type": "operation", "source_id": "9001"}]},
            )

    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(run_import)
        assert fetch_started.wait(timeout=10)
        with TestClient(app) as edit_client:
            login(edit_client)
            edited = edit_client.patch(
                f"/api/batches/{batch['id']}/requests/{request_row['id']}",
                json={"summary": "导入查询期间已编辑", "expected_version": request_row["version"]},
            )
        release_fetch.set()
        imported = future.result(timeout=15)

    assert edited.status_code == 200
    assert imported.status_code == 409
    assert imported.json()["detail"]["code"] == "VERSION_CONFLICT"
    assert imported.json()["detail"]["entity_type"] == "request_batch"
    with TestClient(app) as verify_client:
        login(verify_client)
        rows = verify_client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
    assert len(rows) == 1
    assert rows[0]["summary"] == "导入查询期间已编辑"


def test_sheet_order_is_saved_and_inherited_by_rollover():
    with TestClient(app) as client:
        login(client)
        source = client.post(
            "/api/batches",
            json={"name": "Sheet 排序来源", "start_date": "2026-07-16", "end_date": "2026-07-21"},
        ).json()["batch"]
        for sheet_name in ("采购中心", "财务中心", "供应商"):
            created = client.post(
                f"/api/batches/{source['id']}/requests",
                json={"source_sheet": sheet_name, "payment_account": "私户", "amount": 100},
            )
            assert created.status_code == 200

        order = ["供应商", "财务中心", "采购中心"]
        saved = client.put(f"/api/batches/{source['id']}/sheet-order", json={"sheet_order": order})
        assert saved.status_code == 200
        assert saved.json()["batch"]["sheet_order"] == order
        assert client.get(f"/api/batches/{source['id']}").json()["batch"]["sheet_order"] == order

        exported = client.get(f"/api/batches/{source['id']}/export.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(io.BytesIO(exported.content))
        assert workbook.sheetnames[:4] == ["全部", *order]

        rollover = client.post(
            f"/api/batches/{source['id']}/rollover",
            json={"name": "Sheet 排序结转", "start_date": "2026-07-22", "end_date": "2026-07-28", "copy_mode": "all"},
        )
        assert rollover.status_code == 200
        assert rollover.json()["batch"]["sheet_order"] == order

        archived = client.post(f"/api/batches/{source['id']}/archive")
        assert archived.status_code == 200
        archived_order = list(reversed(order))
        reordered = client.put(f"/api/batches/{source['id']}/sheet-order", json={"sheet_order": archived_order})
        assert reordered.status_code == 200
        assert reordered.json()["batch"]["sheet_order"] == archived_order


def test_request_grid_preferences_are_saved_per_user():
    with TestClient(app) as admin_client, TestClient(app) as business_client:
        login(admin_client)
        default_preference = admin_client.get("/api/me/preferences/request-grid")
        assert default_preference.status_code == 200
        assert "summary" in default_preference.json()["preference"]["order"]
        assert "expected_payment_account" in default_preference.json()["preference"]["order"]
        assert "expected_payment_account" not in default_preference.json()["preference"]["hidden"]

        saved = admin_client.put(
            "/api/me/preferences/request-grid",
            json={"order": ["summary", "amount", "currency"], "hidden": ["amount"]},
        )
        assert saved.status_code == 200
        assert saved.json()["preference"]["order"][:3] == ["summary", "amount", "currency"]
        assert "expected_payment_account" in saved.json()["preference"]["order"]
        assert "amount" in saved.json()["preference"]["hidden"]
        assert "expected_payment_account" not in saved.json()["preference"]["hidden"]
        assert admin_client.get("/api/me/preferences/request-grid").json()["preference"] == saved.json()["preference"]

        hidden_expected_account = admin_client.put(
            "/api/me/preferences/request-grid",
            json={"order": ["summary", "expected_payment_account"], "hidden": ["expected_payment_account"]},
        )
        assert hidden_expected_account.status_code == 200
        assert "expected_payment_account" in hidden_expected_account.json()["preference"]["hidden"]

        created_user = admin_client.post(
            "/api/admin/users",
            json={
                "username": "grid-preference-user",
                "password": "Yuewei123",
                "display_name": "列偏好用户",
                "role": "business",
                "active": True,
                "sheet_permissions": [],
            },
        )
        assert created_user.status_code == 200
        login(business_client, "grid-preference-user", "Yuewei123")
        business_preference = business_client.get("/api/me/preferences/request-grid").json()["preference"]
        assert business_preference != saved.json()["preference"]
        assert "amount" not in business_preference["hidden"]

        invalid = business_client.put(
            "/api/me/preferences/request-grid",
            json={"order": ["not_a_column"], "hidden": []},
        )
        assert invalid.status_code == 400


def test_empty_sheet_registry_and_business_sheet_boundaries():
    with TestClient(app) as admin_client, TestClient(app) as business_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "空 Sheet 与业务边界", "start_date": "2026-07-28", "end_date": "2026-08-03"},
        ).json()["batch"]
        batch_id = batch["id"]
        registered_order = ["部门 A", "部门 B", "空部门"]
        registered = admin_client.put(
            f"/api/batches/{batch_id}/sheet-order",
            json={"sheet_order": registered_order},
        )
        assert registered.status_code == 200
        assert registered.json()["batch"]["sheet_order"] == registered_order

        existing = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "部门 A", "summary": "已有请款", "amount": 100},
        ).json()["request"]
        created_user = admin_client.post(
            "/api/admin/users",
            json={
                "username": "registered-sheet-user",
                "password": "Yuewei123",
                "display_name": "登记 Sheet 用户",
                "role": "business",
                "active": True,
                "sheet_permissions": ["部门 A", "部门 B", "未登记部门"],
            },
        )
        assert created_user.status_code == 200
        login(business_client, "registered-sheet-user", "Yuewei123")

        visible_batch = business_client.get(f"/api/batches/{batch_id}").json()["batch"]
        assert visible_batch["sheet_order"] == ["部门 A", "部门 B"]
        created_in_empty_sheet = business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "部门 B", "summary": "空 Sheet 中新增", "amount": 50},
        )
        assert created_in_empty_sheet.status_code == 200
        created_request_id = created_in_empty_sheet.json()["request"]["id"]
        assert business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "未登记部门", "summary": "越权创建 Sheet", "amount": 1},
        ).status_code == 403
        assert business_client.patch(
            f"/api/batches/{batch_id}/requests/{existing['id']}",
            json={"source_sheet": "部门 B"},
        ).status_code == 403
        assert business_client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={
                "creates": [],
                "updates": [{"id": existing["id"], "source_sheet": "部门 B", "summary": "不应保存"}],
                "deletes": [],
            },
        ).status_code == 403
        assert business_client.put(
            f"/api/batches/{batch_id}/sheet-order",
            json={"sheet_order": ["部门 B", "部门 A"]},
        ).status_code == 403

        deleted = business_client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={"creates": [], "updates": [], "deletes": [created_request_id]},
        )
        assert deleted.status_code == 200
        assert admin_client.get(f"/api/batches/{batch_id}").json()["batch"]["sheet_order"] == registered_order

        removed_empty_sheet = admin_client.put(
            f"/api/batches/{batch_id}/sheet-order",
            json={"sheet_order": ["部门 A", "空部门"]},
        )
        assert removed_empty_sheet.status_code == 200
        assert removed_empty_sheet.json()["batch"]["sheet_order"] == ["部门 A", "空部门"]


def test_employee_workbook_regroups_requests_by_level_two_department_and_keeps_unmatched_sheets():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "员工数据"
    worksheet.append(["员工UserID", "姓名", "1级部门", "2级部门"])
    worksheet.append(["LEVEL2-U1", "二级映射张甲", "一级", "二级制造"])
    worksheet.append(["LEVEL2-U2", "二级映射李乙", "一级", "二级供应链"])
    worksheet.append(["LEVEL2-U3", "二级映射王同名", "一级", "二级甲"])
    worksheet.append(["LEVEL2-U4", "二级映射王同名", "一级", "二级乙"])
    worksheet.append(["LEVEL2-U5", "二级映射无部门", "一级", ""])
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)

    with TestClient(app) as admin_client, TestClient(app) as business_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "员工2级部门归组", "start_date": "2026-08-01", "end_date": "2026-08-07"},
        ).json()["batch"]
        batch_id = batch["id"]
        old_order = ["旧 Sheet A", "旧 Sheet B", "保留原 Sheet", "保留空 Sheet"]
        assert admin_client.put(
            f"/api/batches/{batch_id}/sheet-order",
            json={"sheet_order": old_order},
        ).status_code == 200

        by_id_request = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={
                "applicant": "二级映射张甲",
                "source_sheet": "旧 Sheet A",
                "summary": "按钉钉用户 ID 匹配",
                "amount": 10,
                "raw_extra": {"external_source": {"applicant_id": "LEVEL2-U1", "applicant": "二级映射张甲"}},
            },
        ).json()["request"]
        by_name_request = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "二级映射李乙", "source_sheet": "旧 Sheet B", "summary": "按姓名匹配", "amount": 20},
        ).json()["request"]
        ambiguous_request = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "二级映射王同名", "source_sheet": "保留原 Sheet", "summary": "重名保留", "amount": 30},
        ).json()["request"]
        missing_request = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "保留原 Sheet", "summary": "无申请人保留", "amount": 40},
        ).json()["request"]
        unmatched_request = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "二级映射表外人员", "source_sheet": "保留原 Sheet", "summary": "表外人员保留", "amount": 50},
        ).json()["request"]

        imported = admin_client.post(
            f"/api/batches/{batch_id}/employee-departments/import",
            files={
                "file": (
                    "员工信息.xlsx",
                    workbook_buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert imported.status_code == 200, imported.text
        result = imported.json()
        assert result["mapping_rows"] == 4
        assert result["skipped_employee_no_department"] == 1
        assert result["matched_requests"] == 2
        assert result["moved_requests"] == 2
        assert result["missing_applicant"] == 1
        assert result["unmatched_applicant"] == 1
        assert result["ambiguous_applicant"] == 1
        assert set(result["removed_empty_sheets"]) == {"旧 Sheet A", "旧 Sheet B"}
        assert "保留空 Sheet" in result["sheet_order"]

        with connect() as conn:
            rows = {
                row["id"]: row
                for row in conn.execute(
                    "SELECT id, source_sheet, raw_extra_json FROM payment_requests WHERE batch_id = ?",
                    (batch_id,),
                ).fetchall()
            }
            assert rows[by_id_request["id"]]["source_sheet"] == "二级制造"
            assert rows[by_name_request["id"]]["source_sheet"] == "二级供应链"
            assert rows[ambiguous_request["id"]]["source_sheet"] == "保留原 Sheet"
            assert rows[missing_request["id"]]["source_sheet"] == "保留原 Sheet"
            assert rows[unmatched_request["id"]]["source_sheet"] == "保留原 Sheet"
            raw_extra = json.loads(rows[by_id_request["id"]]["raw_extra_json"])
            assert raw_extra["employee_department_mapping"]["second_level_department"] == "二级制造"

        future = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "二级映射张甲", "source_sheet": "人工临时 Sheet", "summary": "后续新增自动归组", "amount": 60},
        )
        assert future.status_code == 200
        assert future.json()["request"]["source_sheet"] == "二级制造"
        assert future.json()["request"]["raw_extra"]["employee_department_mapping"]["second_level_department"] == "二级制造"

        created_user = admin_client.post(
            "/api/admin/users",
            json={
                "username": "level-two-business-user",
                "password": "Yuewei123",
                "display_name": "二级部门业务用户",
                "role": "business",
                "active": True,
                "sheet_permissions": ["保留空 Sheet"],
            },
        )
        assert created_user.status_code == 200
        login(business_client, "level-two-business-user", "Yuewei123")
        business_created = business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "二级映射张甲", "source_sheet": "保留空 Sheet", "summary": "业务人员不被跨 Sheet 移动", "amount": 70},
        )
        assert business_created.status_code == 200
        assert business_created.json()["request"]["source_sheet"] == "保留空 Sheet"

        archived_batch = admin_client.post(
            "/api/batches",
            json={"name": "员工归组归档限制", "start_date": "2026-08-08", "end_date": "2026-08-14"},
        ).json()["batch"]
        assert admin_client.post(f"/api/batches/{archived_batch['id']}/archive").status_code == 200
        rejected = admin_client.post(
            f"/api/batches/{archived_batch['id']}/employee-departments/import",
            files={"file": ("员工信息.xlsx", workbook_buffer.getvalue())},
        )
        assert rejected.status_code == 400


def test_manual_sheet_move_wins_over_unchanged_employee_mapping_data():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "员工数据"
    worksheet.append(["员工UserID", "姓名", "1级部门", "2级部门"])
    worksheet.append(["MOVE-U1", "移动测试人员", "一级", "自动归组部门"])
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)

    with TestClient(app) as admin_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "手动移动优先", "start_date": "2026-08-15", "end_date": "2026-08-21"},
        ).json()["batch"]
        batch_id = batch["id"]
        imported = admin_client.post(
            f"/api/batches/{batch_id}/employee-departments/import",
            files={
                "file": (
                    "员工信息.xlsx",
                    workbook_buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert imported.status_code == 200, imported.text

        created = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={
                "applicant": "移动测试人员",
                "source_sheet": "导入临时 Sheet",
                "summary": "移动前",
                "amount": 100,
            },
        ).json()["request"]
        assert created["source_sheet"] == "自动归组部门"

        moved = admin_client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={
                "creates": [],
                "updates": [
                    {
                        "id": created["id"],
                        "expected_version": created["version"],
                        "source_sheet": "人工调整 Sheet",
                        # Simulate the legacy frontend sending the whole row.
                        "applicant": "移动测试人员",
                        "summary": "移动前",
                    }
                ],
                "deletes": [],
            },
        )
        assert moved.status_code == 200, moved.text
        refreshed = admin_client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        request = next(item for item in refreshed if item["id"] == created["id"])
        assert request["source_sheet"] == "人工调整 Sheet"


def test_unchanged_applicant_does_not_remap_sheet_during_unrelated_edit():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "员工数据"
    worksheet.append(["员工UserID", "姓名", "1级部门", "2级部门"])
    worksheet.append(["EDIT-U1", "普通编辑测试人员", "一级", "自动归组部门"])
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)

    with TestClient(app) as admin_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "普通编辑不归组", "start_date": "2026-08-15", "end_date": "2026-08-21"},
        ).json()["batch"]
        batch_id = batch["id"]
        assert admin_client.post(
            f"/api/batches/{batch_id}/employee-departments/import",
            files={"file": ("员工信息.xlsx", workbook_buffer.getvalue())},
        ).status_code == 200
        created = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "普通编辑测试人员", "summary": "修改前", "amount": 100},
        ).json()["request"]

        with connect() as conn:
            conn.execute(
                "UPDATE payment_requests SET source_sheet = '人工保留 Sheet', version = version + 1 WHERE id = ?",
                (created["id"],),
            )
            current_version = conn.execute(
                "SELECT version FROM payment_requests WHERE id = ?",
                (created["id"],),
            ).fetchone()["version"]

        updated = admin_client.patch(
            f"/api/batches/{batch_id}/requests/{created['id']}",
            json={
                "expected_version": current_version,
                "applicant": "普通编辑测试人员",
                "summary": "修改后",
            },
        )
        assert updated.status_code == 200, updated.text
        assert updated.json()["request"]["source_sheet"] == "人工保留 Sheet"
        assert updated.json()["request"]["summary"] == "修改后"


def test_changed_applicant_still_remaps_when_full_editor_payload_keeps_old_sheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "员工数据"
    worksheet.append(["员工UserID", "姓名", "1级部门", "2级部门"])
    worksheet.append(["EDITOR-U1", "抽屉原申请人", "一级", "原申请人部门"])
    worksheet.append(["EDITOR-U2", "抽屉新申请人", "一级", "新申请人部门"])
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)

    with TestClient(app) as admin_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "抽屉申请人换组", "start_date": "2026-08-15", "end_date": "2026-08-21"},
        ).json()["batch"]
        batch_id = batch["id"]
        assert admin_client.post(
            f"/api/batches/{batch_id}/employee-departments/import",
            files={"file": ("员工信息.xlsx", workbook_buffer.getvalue())},
        ).status_code == 200
        created = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"applicant": "抽屉原申请人", "summary": "修改申请人前", "amount": 100},
        ).json()["request"]
        assert created["source_sheet"] == "原申请人部门"

        updated = admin_client.patch(
            f"/api/batches/{batch_id}/requests/{created['id']}",
            json={
                "expected_version": created["version"],
                "applicant": "抽屉新申请人",
                # The drawer sends the full request, including the unchanged
                # old Sheet. Its mere presence must not suppress re-mapping.
                "source_sheet": "原申请人部门",
                "summary": "修改申请人后",
            },
        )
        assert updated.status_code == 200, updated.text
        assert updated.json()["request"]["source_sheet"] == "新申请人部门"


def test_sparse_grid_update_can_clear_optional_fields_without_overwriting_others():
    with TestClient(app) as admin_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "稀疏更新清空字段", "start_date": "2026-08-15", "end_date": "2026-08-21"},
        ).json()["batch"]
        batch_id = batch["id"]
        created = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={
                "source_sheet": "清空测试 Sheet",
                "summary": "保留摘要",
                "payee_name": "待清空收款人",
                "needed_payment_date": "2026-08-21",
                "remark": "待清空备注",
                "amount": 100,
            },
        ).json()["request"]

        cleared = admin_client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={
                "creates": [],
                "updates": [
                    {
                        "id": created["id"],
                        "expected_version": created["version"],
                        "payee_name": "",
                        "needed_payment_date": None,
                        "remark": "",
                    }
                ],
                "deletes": [],
            },
        )
        assert cleared.status_code == 200, cleared.text
        request = next(
            item
            for item in admin_client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
            if item["id"] == created["id"]
        )
        assert request["payee_name"] == ""
        assert request["needed_payment_date"] is None
        assert request["remark"] in (None, "")
        assert request["summary"] == "保留摘要"
        assert request["source_sheet"] == "清空测试 Sheet"


def test_employee_workbook_splits_supply_chain_center_by_level_three_department():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "员工数据"
    worksheet.append(["员工UserID", "姓名", "1级部门", "2级部门", "3级部门"])
    worksheet.append([
        "LEVEL3-U1",
        "三级映射李甲",
        "悦为集团",
        "凌翔/星铭供应链及职能中心",
        "凌翔产品&开发",
    ])
    worksheet.append([
        "LEVEL3-U2",
        "三级映射吴乙",
        "悦为集团",
        "凌翔/星铭供应链及职能中心",
        "星铭FC财务中心",
    ])
    worksheet.append([
        "LEVEL3-U3",
        "三级映射普通部门",
        "悦为集团",
        "普通二级部门",
        "普通三级部门",
    ])
    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)

    with TestClient(app) as client:
        login(client)
        batch_id = client.post(
            "/api/batches",
            json={"name": "员工3级部门例外归组", "start_date": "2026-08-08", "end_date": "2026-08-14"},
        ).json()["batch"]["id"]
        requests = []
        for applicant in ("三级映射李甲", "三级映射吴乙", "三级映射普通部门"):
            requests.append(
                client.post(
                    f"/api/batches/{batch_id}/requests",
                    json={"applicant": applicant, "source_sheet": "原 Sheet", "summary": applicant, "amount": 10},
                ).json()["request"]
            )

        response = client.post(
            f"/api/batches/{batch_id}/employee-departments/import",
            files={
                "file": (
                    "员工信息.xlsx",
                    workbook_buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200, response.text
        rows = client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        sheets = {row["applicant"]: row["source_sheet"] for row in rows}
        assert sheets == {
            "三级映射李甲": "凌翔产品&开发",
            "三级映射吴乙": "星铭FC财务中心",
            "三级映射普通部门": "普通二级部门",
        }
        mapped = next(row for row in rows if row["id"] == requests[0]["id"])
        assert mapped["raw_extra"]["employee_department_mapping"]["third_level_department"] == "凌翔产品&开发"
        assert mapped["raw_extra"]["employee_department_mapping"]["assigned_department"] == "凌翔产品&开发"


def test_legacy_mould_sheet_names_are_migrated_in_rows_order_and_permissions():
    assert main_module.canonical_sheet_name("赣瑞模具 7 月 后") == "赣瑞模具"
    assert main_module.canonical_sheet_name("志威模具 ( 7 月 前 )") == "志威模具"
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "模具 Sheet 归一", "start_date": "2026-07-28", "end_date": "2026-08-03"},
        ).json()["batch"]
        user = client.post(
            "/api/admin/users",
            json={
                "username": "legacy-mould-user",
                "password": "Yuewei123",
                "display_name": "模具负责人",
                "role": "business",
                "active": True,
                "sheet_permissions": ["其他部门"],
            },
        ).json()["user"]

        with connect() as conn:
            timestamp = now_iso()
            for index, sheet_name in enumerate(
                ("赣瑞模具（7月前）", "赣瑞模具（7月后）", "志威模具 (7月前)", "志威模具（7月后）"),
                start=1,
            ):
                conn.execute(
                    """
                    INSERT INTO payment_requests (
                        batch_id, source_sheet, summary, amount, paid_amount, pending_amount,
                        currency, raw_extra_json, created_by, updated_by, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, 0, ?, 'CNY', '{}', 1, 1, ?, ?)
                    """,
                    (batch["id"], sheet_name, f"迁移记录 {index}", index * 100, index * 100, timestamp, timestamp),
                )
            conn.execute(
                "UPDATE request_batches SET sheet_order_json = ? WHERE id = ?",
                (
                    json.dumps(
                        ["其他部门", "志威模具（7月后）", "赣瑞模具（7月前）", "志威模具 (7月前)", "赣瑞模具（7月后）"],
                        ensure_ascii=False,
                    ),
                    batch["id"],
                ),
            )
            conn.execute("DELETE FROM user_sheet_permissions WHERE user_id = ?", (user["id"],))
            for sheet_name in ("赣瑞模具（7月前）", "赣瑞模具（7月后）", "志威模具（7月前）", "志威模具（7月后）"):
                conn.execute(
                    """
                    INSERT INTO user_sheet_permissions (user_id, sheet_name, created_by, created_at)
                    VALUES (?, ?, 1, ?)
                    """,
                    (user["id"], sheet_name, timestamp),
                )
            conn.execute(
                "DELETE FROM schema_migrations WHERE key = 'sheet_registry_and_mould_names_v1'"
            )
            migrate_sheet_registry_and_names(conn)

            row_names = {
                row["source_sheet"]
                for row in conn.execute(
                    "SELECT source_sheet FROM payment_requests WHERE batch_id = ?",
                    (batch["id"],),
                ).fetchall()
            }
            assert row_names == {"赣瑞模具", "志威模具"}
            migrated_batch = conn.execute(
                "SELECT sheet_order_json FROM request_batches WHERE id = ?",
                (batch["id"],),
            ).fetchone()
            assert json.loads(migrated_batch["sheet_order_json"]) == ["其他部门", "志威模具", "赣瑞模具"]
            permissions = {
                row["sheet_name"]
                for row in conn.execute(
                    "SELECT sheet_name FROM user_sheet_permissions WHERE user_id = ?",
                    (user["id"],),
                ).fetchall()
            }
            assert permissions == {"赣瑞模具", "志威模具"}

        future = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"source_sheet": "赣瑞模具 (7月后)", "summary": "未来旧名称输入", "amount": 500},
        )
        assert future.status_code == 200
        assert future.json()["request"]["source_sheet"] == "赣瑞模具"


def test_filtered_export_matches_workspace_filters_and_keeps_all_sheet():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "筛选导出测试", "start_date": "2026-07-16", "end_date": "2026-07-23"},
        ).json()["batch"]

        def create(sheet: str, applicant: str, manager_approval: str, amount: float) -> dict:
            response = client.post(
                f"/api/batches/{batch['id']}/requests",
                json={
                    "source_sheet": sheet,
                    "applicant": applicant,
                    "summary": f"{applicant}的请款",
                    "amount": amount,
                    "general_manager_approval": manager_approval,
                },
            )
            assert response.status_code == 200
            return response.json()["request"]

        expected_a = create("采购中心", "张三", "同意付款", 100)
        create("采购中心", "李四", "存在争议", 200)
        expected_b = create("财务中心", "王五", "同意付款", 300)
        paid = create("财务中心", "赵六", "同意付款", 400)
        payment = client.post(
            f"/api/batches/{batch['id']}/requests/{paid['id']}/payments",
            json={"amount": 400, "payment_date": "2026-07-23"},
        )
        assert payment.status_code == 200
        order = client.put(
            f"/api/batches/{batch['id']}/sheet-order",
            json={"sheet_order": ["采购中心", "财务中心"]},
        )
        assert order.status_code == 200

        exported = client.get(
            f"/api/batches/{batch['id']}/export.xlsx",
            params={
                "filtered": "true",
                "finance_review": "未付款",
                "general_manager_approval": "同意付款",
            },
        )
        assert exported.status_code == 200
        assert "筛选结果" in unquote(exported.headers["content-disposition"])
        workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
        assert workbook.sheetnames[:3] == ["全部", "采购中心", "财务中心"]

        all_sheet = workbook["全部"]
        headers = [cell.value for cell in all_sheet[2]]
        request_id_column = headers.index("请款标识") + 1
        exported_request_ids = {
            all_sheet.cell(row, request_id_column).value
            for row in range(3, all_sheet.max_row + 1)
            if all_sheet.cell(row, request_id_column).value
        }
        assert exported_request_ids == {expected_a["id"], expected_b["id"]}
        assert workbook["付款明细"].max_row == 1

        amount_filtered = client.get(
            f"/api/batches/{batch['id']}/export.xlsx",
            params={
                "filtered": "true",
                "pending_amount_min": "250",
                "pending_amount_max": "350",
            },
        )
        assert amount_filtered.status_code == 200
        amount_workbook = load_workbook(io.BytesIO(amount_filtered.content), data_only=False)
        amount_all_sheet = amount_workbook["全部"]
        amount_headers = [cell.value for cell in amount_all_sheet[2]]
        amount_request_id_column = amount_headers.index("请款标识") + 1
        amount_request_ids = {
            amount_all_sheet.cell(row, amount_request_id_column).value
            for row in range(3, amount_all_sheet.max_row + 1)
            if amount_all_sheet.cell(row, amount_request_id_column).value
        }
        assert amount_request_ids == {expected_b["id"]}

        sheet_filtered = client.get(
            f"/api/batches/{batch['id']}/export.xlsx",
            params={
                "filtered": "true",
                "finance_review": "未付款",
                "general_manager_approval": "同意付款",
                "source_sheet": "采购中心",
            },
        )
        sheet_workbook = load_workbook(io.BytesIO(sheet_filtered.content), data_only=False)
        assert sheet_workbook.sheetnames[:2] == ["全部", "采购中心"]
        assert "财务中心" not in sheet_workbook.sheetnames


def external_expense_test_row(
    approval_no: str,
    source_id: str,
    *,
    source_type: str = "operation",
    status: str = "RUNNING",
    amount: float = 123.45,
    beneficiary: str = "测试收款信息",
    warnings=None,
    execution_region: str = "",
    source_sheet: str = "测试部门",
    expected_payment_account: str = "",
    expected_payment_account_source: str = "",
) -> dict:
    source_label = {
        "operation": "运营支出",
        "purchase": "采购支出",
        "monthly": "月结付款",
    }[source_type]
    request_data = {
        "dingding_id": approval_no,
        "expense_type": "测试支出",
        "summary": "中间表测试",
        "amount": amount,
        "currency": "CNY",
        "expected_payment_account": expected_payment_account or None,
        "expected_payment_account_source": expected_payment_account_source or None,
        "payee_account": beneficiary or None,
        "source_sheet": source_sheet,
        "raw_extra": {
            "external_source": {
                "system": "dingtalk_expense_database",
                "table": f"approval_expense_{source_type}",
                "record_id": source_id,
                "approval_no": approval_no,
                "approval_status": status,
                "applicant_id": "test-user-id",
                "applicant": "测试申请人",
                "applicant_department": source_sheet,
                "application_date": "2026-07-15",
                "execution_region": execution_region,
            }
        },
    }
    return {
        "source_type": source_type,
        "source_label": source_label,
        "source_id": source_id,
        "application_date": "2026-07-15",
        "approval_no": approval_no,
        "applicant_id": "test-user-id",
        "applicant": "测试申请人",
        "applicant_department": source_sheet,
        "approval_status": status,
        "approval_result": "agree",
        "summary": "中间表测试",
        "amount": amount,
        "beneficiary": beneficiary,
        "needed_payment_date": "2026-07-20",
        "expected_payment_account": expected_payment_account or None,
        "expected_payment_account_source": expected_payment_account_source or None,
        "warnings": warnings or [],
        "errors": [],
        "source_conflict": False,
        "request_data": request_data,
    }


def test_expected_payment_account_manual_create_patch_clear_and_bulk_source_spoof():
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "expected-account-manual"}).json()["batch"]

        created = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "source_sheet": "悦为智能",
                "amount": 100,
                "expected_payment_account": "  人工账户 A  ",
            },
        )
        assert created.status_code == 200
        created_row = created.json()["request"]
        assert created_row["expected_payment_account"] == "人工账户 A"
        assert created_row["expected_payment_account_source"] == "manual"

        patched = client.patch(
            f"/api/batches/{batch['id']}/requests/{created_row['id']}",
            json={
                "expected_payment_account": "人工账户 B",
                "expected_version": created_row["version"],
            },
        )
        assert patched.status_code == 200
        patched_row = patched.json()["request"]
        assert patched_row["expected_payment_account"] == "人工账户 B"
        assert patched_row["expected_payment_account_source"] == "manual"

        cleared = client.patch(
            f"/api/batches/{batch['id']}/requests/{created_row['id']}",
            json={
                "expected_payment_account": "",
                "expected_version": patched_row["version"],
            },
        )
        assert cleared.status_code == 200
        cleared_row = cleared.json()["request"]
        assert cleared_row["expected_payment_account"] is None
        assert cleared_row["expected_payment_account_source"] is None

        bulk = client.patch(
            f"/api/batches/{batch['id']}/requests/bulk",
            json={
                "creates": [{
                    "source_sheet": "悦为智能",
                    "amount": 50,
                    "expected_payment_account": "批量人工账户",
                    "expected_payment_account_source": "dingtalk_explicit",
                }],
            },
        )
        assert bulk.status_code == 200
        bulk_row = next(
            row
            for row in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
            if row["id"] in bulk.json()["created"]
        )
        assert bulk_row["expected_payment_account"] == "批量人工账户"
        assert bulk_row["expected_payment_account_source"] == "manual"


def test_expected_payment_account_external_import_preserves_trusted_source(monkeypatch):
    explicit = external_expense_test_row(
        "EXPECTED-IMPORT-EXPLICIT",
        "9301",
        execution_region="中国China",
        expected_payment_account="悦为智能 6221 主账户",
        expected_payment_account_source="dingtalk_explicit",
    )
    defaulted = external_expense_test_row(
        "EXPECTED-IMPORT-DEFAULT",
        "9302",
        execution_region="中国China",
        expected_payment_account="凌翔公司账户",
        expected_payment_account_source="service_subject_default",
    )
    monkeypatch.setattr(
        main_module,
        "fetch_external_expenses",
        lambda items: [explicit, defaulted],
    )

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "expected-account-import"}).json()["batch"]
        imported = client.post(
            f"/api/batches/{batch['id']}/imports/external-expenses",
            json={"items": [
                {"source_type": "operation", "source_id": "9301"},
                {"source_type": "operation", "source_id": "9302"},
            ]},
        )

        assert imported.status_code == 200
        assert imported.json()["imported_rows"] == 2
        rows = {
            row["dingding_id"]: row
            for row in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
        }
        assert rows["EXPECTED-IMPORT-EXPLICIT"]["expected_payment_account"] == "悦为智能 6221 主账户"
        assert rows["EXPECTED-IMPORT-EXPLICIT"]["expected_payment_account_source"] == "dingtalk_explicit"
        assert rows["EXPECTED-IMPORT-DEFAULT"]["expected_payment_account"] == "凌翔公司账户"
        assert rows["EXPECTED-IMPORT-DEFAULT"]["expected_payment_account_source"] == "service_subject_default"


def test_expected_payment_account_rollover_preserves_value_and_source():
    with TestClient(app) as client:
        login(client)
        source = client.post(
            "/api/batches",
            json={"name": "expected-account-rollover-source"},
        ).json()["batch"]
        manual = client.post(
            f"/api/batches/{source['id']}/requests",
            json={
                "dingding_id": "EXPECTED-ROLLOVER-MANUAL",
                "source_sheet": "悦为智能",
                "amount": 100,
                "expected_payment_account": "人工账户",
            },
        ).json()["request"]
        automatic = client.post(
            f"/api/batches/{source['id']}/requests",
            json={
                "dingding_id": "EXPECTED-ROLLOVER-AUTO",
                "source_sheet": "悦为智能",
                "amount": 100,
            },
        ).json()["request"]
        with connect() as conn:
            conn.execute(
                """
                UPDATE payment_requests
                SET expected_payment_account = '悦为智能公司账户',
                    expected_payment_account_source = 'service_subject_default'
                WHERE id = ?
                """,
                (automatic["id"],),
            )

        rollover = client.post(
            f"/api/batches/{source['id']}/rollover",
            json={"name": "expected-account-rollover-target", "copy_mode": "all"},
        )
        assert rollover.status_code == 200
        rows = {
            row["dingding_id"]: row
            for row in client.get(
                f"/api/batches/{rollover.json()['batch']['id']}/requests"
            ).json()["requests"]
        }
        assert rows["EXPECTED-ROLLOVER-MANUAL"]["expected_payment_account"] == "人工账户"
        assert rows["EXPECTED-ROLLOVER-MANUAL"]["expected_payment_account_source"] == "manual"
        assert rows["EXPECTED-ROLLOVER-AUTO"]["expected_payment_account"] == "悦为智能公司账户"
        assert rows["EXPECTED-ROLLOVER-AUTO"]["expected_payment_account_source"] == "service_subject_default"
        assert rows["EXPECTED-ROLLOVER-MANUAL"]["copied_from_request_id"] == manual["id"]


def test_rollover_copies_only_unfinished_rows():
    if not SAMPLE.exists():
        return
    with TestClient(app) as client:
        login(client)
        with SAMPLE.open("rb") as handle:
            imported = client.post(
                "/api/import/weekly-excel",
                files={"file": (SAMPLE.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert imported.status_code == 200
        imported_payload = imported.json()
        source_batch_id = imported_payload["batch_id"]
        assert imported_payload["meta"]["images"]["saved"] == 9
        listed_attachments = client.get(f"/api/batches/{source_batch_id}/attachments")
        assert listed_attachments.status_code == 200
        assert len(listed_attachments.json()["attachments"]) == 9
        exported_source = client.get(f"/api/batches/{source_batch_id}/export.xlsx")
        assert exported_source.status_code == 200
        exported_workbook = load_workbook(io.BytesIO(exported_source.content))
        exported_headers = [
            cell.value
            for worksheet in exported_workbook.worksheets
            for row in worksheet.iter_rows(max_row=2)
            for cell in row
        ]
        assert "图片附件" in exported_headers
        assert "付款情况" not in exported_headers
        assert sum(len(getattr(worksheet, "_images", [])) for worksheet in exported_workbook.worksheets) == 9
        source_requests = client.get(f"/api/batches/{source_batch_id}/requests").json()["requests"]
        partial_source = next(row for row in source_requests if row["finance_review"] == "未付款" and float(row.get("amount") or 0) > 1)
        partial_source_id = partial_source["id"]
        partial_update = client.post(
            f"/api/batches/{source_batch_id}/requests/{partial_source_id}/payments",
            json={"amount": round(float(partial_source["amount"]) / 2, 2), "payment_date": "2026-07-10"},
        )
        assert partial_update.status_code == 200
        assert partial_update.json()["request"]["finance_review"] == "部分付款"
        archived = client.post(f"/api/batches/{source_batch_id}/archive")
        assert archived.status_code == 200
        rollover = client.post(
            f"/api/batches/{source_batch_id}/rollover",
            json={"name": "20260708~20260714请款明细", "start_date": "2026-07-08", "end_date": "2026-07-14"},
        )
        assert rollover.status_code == 200
        copied_count = rollover.json()["copied_count"]
        assert copied_count > 0
        target_batch_id = rollover.json()["batch"]["id"]
        requests = client.get(f"/api/batches/{target_batch_id}/requests").json()["requests"]
        assert len(requests) == copied_count
        assert all(row.get("finance_review") != "已付款" for row in requests)
        assert any(row.get("copied_from_request_id") == partial_source_id and row.get("finance_review") == "部分付款" for row in requests)
        assert all(not row.get("actual_payment_date") or row.get("finance_review") == "部分付款" for row in requests)
        assert all(row.get("copied_from_request_id") for row in requests)

        rollover_all = client.post(
            f"/api/batches/{source_batch_id}/rollover",
            json={
                "name": "20260715~20260721请款明细",
                "start_date": "2026-07-15",
                "end_date": "2026-07-21",
                "copy_mode": "all",
            },
        )
        assert rollover_all.status_code == 200
        assert rollover_all.json()["copied_count"] == 161
        assert rollover_all.json()["copy_mode"] == "all"
        target_all_batch_id = rollover_all.json()["batch"]["id"]
        all_requests = client.get(f"/api/batches/{target_all_batch_id}/requests").json()["requests"]
        assert len(all_requests) == 161
        assert any(row.get("finance_review") == "已付款" for row in all_requests)
        assert all(row.get("copied_from_request_id") for row in all_requests)
        all_attachments = client.get(f"/api/batches/{target_all_batch_id}/attachments")
        assert all_attachments.status_code == 200
        assert len(all_attachments.json()["attachments"]) == 9


def test_rollover_skips_exact_dingtalk_duplicates_but_keeps_legitimate_split_rows():
    approval_no = "202607071243000115462"
    with TestClient(app) as client:
        login(client)
        source = client.post(
            "/api/batches",
            json={"name": "重复治理来源", "start_date": "2026-08-01", "end_date": "2026-08-07"},
        ).json()["batch"]
        exact_duplicate = {
            "dingding_id": approval_no,
            "source_sheet": "凌翔产品&开发",
            "summary": "2026年1份接收模具订单，超队2个型号",
            "amount": 21000,
            "currency": "CNY",
            "payee_account": "供应商收款账户",
        }
        first = client.post(f"/api/batches/{source['id']}/requests", json=exact_duplicate)
        second = client.post(
            f"/api/batches/{source['id']}/requests",
            json={**exact_duplicate, "source_sheet": "历史重组后 Sheet"},
        )
        split = client.post(
            f"/api/batches/{source['id']}/requests",
            json={**exact_duplicate, "summary": "同一审批拆分的第二笔请款", "amount": 5000},
        )
        assert first.status_code == second.status_code == split.status_code == 200

        latest_batch = client.get(f"/api/batches/{source['id']}").json()["batch"]
        rolled = client.post(
            f"/api/batches/{source['id']}/rollover",
            json={
                "name": "重复治理目标",
                "start_date": "2026-08-08",
                "end_date": "2026-08-14",
                "copy_mode": "all",
                "expected_batch_version": latest_batch["version"],
            },
        )

        assert rolled.status_code == 200
        assert rolled.json()["copied_count"] == 2
        assert rolled.json()["skipped_duplicate_rows"] == 1
        copied = client.get(
            f"/api/batches/{rolled.json()['batch']['id']}/requests",
            params={"dingtalk_lifecycle": "all"},
        ).json()["requests"]
        assert len(copied) == 2
        assert sorted(row["amount"] for row in copied) == [5000, 21000]
        assert {row["dingding_id"] for row in copied} == {approval_no}


def test_rollover_preserves_legacy_foreign_currency_rows_without_rate_anchor():
    with TestClient(app) as client:
        login(client)
        source = client.post(
            "/api/batches",
            json={"name": "legacy-fx-source", "start_date": "2026-08-01", "end_date": "2026-08-07"},
        ).json()["batch"]
        created = client.post(
            f"/api/batches/{source['id']}/requests",
            json={"source_sheet": "外币", "summary": "历史美元请款", "amount": 25000, "currency": "CNY"},
        )
        assert created.status_code == 200
        request_id = created.json()["request"]["id"]
        with connect() as conn:
            conn.execute(
                """
                UPDATE payment_requests
                SET currency = 'USD', base_amount_cny = NULL,
                    fx_rate_cny_per_unit = NULL, fx_rate_date = NULL,
                    fx_rate_actual_date = NULL
                WHERE id = ?
                """,
                (request_id,),
            )

        rollover = client.post(
            f"/api/batches/{source['id']}/rollover",
            json={
                "name": "legacy-fx-target",
                "start_date": "2026-08-08",
                "end_date": "2026-08-14",
                "copy_mode": "unfinished",
            },
        )
        assert rollover.status_code == 200
        target_id = rollover.json()["batch"]["id"]
        copied = client.get(f"/api/batches/{target_id}/requests").json()["requests"]
        assert len(copied) == 1
        assert copied[0]["currency"] == "USD"
        assert copied[0]["amount"] == 25000
        assert copied[0]["base_amount_cny"] == 25000
        assert copied[0]["fx_rate_cny_per_unit"] == 1


def test_rollback_latest_import_restores_previous_batch_state():
    if not SAMPLE.exists():
        return
    with TestClient(app) as client:
        login(client)
        with SAMPLE.open("rb") as handle:
            first = client.post(
                "/api/import/weekly-excel",
                files={"file": (SAMPLE.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert first.status_code == 200
        batch_id = first.json()["batch_id"]
        with SAMPLE.open("rb") as handle:
            second = client.post(
                "/api/import/weekly-excel",
                data={"batch_id": str(batch_id)},
                files={"file": (SAMPLE.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert second.status_code == 200
        assert client.get(f"/api/batches/{batch_id}/requests").json()["totals"]["count"] == 322
        assert len(client.get(f"/api/batches/{batch_id}/attachments").json()["attachments"]) == 18

        rolled_back = client.post(f"/api/batches/{batch_id}/imports/latest/rollback")
        assert rolled_back.status_code == 200
        payload = rolled_back.json()
        assert payload["deleted_requests"] == 161
        assert payload["deleted_attachments"] == 9
        assert client.get(f"/api/batches/{batch_id}/requests").json()["totals"]["count"] == 161
        assert len(client.get(f"/api/batches/{batch_id}/attachments").json()["attachments"]) == 9
        with connect() as conn:
            row = conn.execute("SELECT status, imported_rows FROM import_jobs WHERE id = ?", (payload["job_id"],)).fetchone()
            assert row["status"] == "rolled_back"
            assert row["imported_rows"] == 0


def test_bulk_save_create_update_delete_and_rollback():
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "bulk-test", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        batch_id = batch["id"]
        created = client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={
                "creates": [
                    {
                        "dingding_id": "D-1",
                        "payment_account": "私户",
                        "expense_type": "材料款",
                        "summary": "测试请款",
                        "amount": 123.45,
                        "general_manager_approval": "存在争议",
                        "general_manager_approval_date": "2026-07-10",
                    }
                ],
                "updates": [],
                "deletes": [],
            },
        )
        assert created.status_code == 200
        request_id = created.json()["created"][0]
        blocked_summary_update = client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={"creates": [], "updates": [{"id": request_id, "finance_review": "已付款"}], "deletes": []},
        )
        assert blocked_summary_update.status_code == 400
        paid = client.post(
            f"/api/batches/{batch_id}/requests/{request_id}/payments",
            json={"amount": 123.45, "payment_date": "2026-07-09"},
        )
        assert paid.status_code == 200
        row = client.get(f"/api/batches/{batch_id}/requests").json()["requests"][0]
        assert row["finance_review"] == "已付款"
        assert row["payment_status"] == "已付款"
        assert row["actual_payment_date"] == "2026-07-09"
        assert row["general_manager_approval"] == "存在争议"
        assert row["general_manager_approval_date"] == "2026-07-10"
        split_manager_opinion = client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={
                "creates": [],
                "updates": [{"id": request_id, "general_manager_approval": "存在争议，供应商还在谈解决方案。"}],
                "deletes": [],
            },
        )
        assert split_manager_opinion.status_code == 200
        row = client.get(f"/api/batches/{batch_id}/requests").json()["requests"][0]
        assert row["general_manager_approval"] == "存在争议"
        assert "供应商还在谈解决方案" in row["general_manager_opinion"]
        failed = client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={
                "creates": [{"dingding_id": "D-rollback", "summary": "不应保存", "amount": 1}],
                "updates": [{"id": 999999, "finance_review": "已付款"}],
                "deletes": [],
            },
        )
        assert failed.status_code == 404
        rows_after_failure = client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        assert all(row.get("dingding_id") != "D-rollback" for row in rows_after_failure)
        deleted = client.patch(
            f"/api/batches/{batch_id}/requests/bulk",
            json={"creates": [], "updates": [], "deletes": [request_id]},
        )
        assert deleted.status_code == 200
        assert client.get(f"/api/batches/{batch_id}/requests").json()["totals"]["count"] == 0


def test_partial_payment_amounts_are_calculated_and_summarized():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "partial-payment-test", "start_date": "2026-07-01", "end_date": "2026-07-07"},
        ).json()["batch"]
        batch_id = batch["id"]
        created = client.post(
            f"/api/batches/{batch_id}/requests",
            json={"summary": "分三次付款", "amount": 100},
        )
        assert created.status_code == 200
        row = created.json()["request"]
        assert row["paid_amount"] == 0
        assert row["pending_amount"] == 100
        assert row["finance_review"] == "未付款"

        direct_update = client.patch(
            f"/api/batches/{batch_id}/requests/{row['id']}",
            json={"paid_amount": 30},
        )
        assert direct_update.status_code == 400

        for index, amount in enumerate((30, 20, 50), start=1):
            payment = client.post(
                f"/api/batches/{batch_id}/requests/{row['id']}/payments",
                json={"amount": amount, "payment_date": f"2026-07-{index + 9:02d}", "payer": f"付款人{index}"},
            )
            assert payment.status_code == 200
            updated_row = payment.json()["request"]
            expected_paid = sum((30, 20, 50)[:index])
            assert updated_row["paid_amount"] == expected_paid
            assert updated_row["pending_amount"] == 100 - expected_paid
            assert updated_row["finance_review"] == ("已付款" if index == 3 else "部分付款")
            assert updated_row["general_manager_approval"] == ("同意付款" if index == 3 else None)

        detail = client.get(f"/api/batches/{batch_id}/requests/{row['id']}/payments").json()
        assert detail["summary"]["payment_count"] == 3
        assert len(detail["payments"]) == 3

        overpaid = client.post(
            f"/api/batches/{batch_id}/requests/{row['id']}/payments",
            json={"amount": 0.01, "payment_date": "2026-07-13"},
        )
        assert overpaid.status_code == 400

        totals = client.get(f"/api/batches/{batch_id}/requests").json()["totals"]
        assert totals == {"count": 1, "amount": 100.0, "paid_amount": 100.0, "pending_amount": 0.0}
        batch_totals = next(item for item in client.get("/api/batches").json()["batches"] if item["id"] == batch_id)
        assert batch_totals["total_amount"] == 100
        assert batch_totals["total_paid_amount"] == 100
        assert batch_totals["total_pending_amount"] == 0

        increased = client.patch(
            f"/api/batches/{batch_id}/requests/{row['id']}",
            json={"amount": 120},
        )
        assert increased.status_code == 200
        assert increased.json()["request"]["paid_amount"] == 100
        assert increased.json()["request"]["pending_amount"] == 20
        assert increased.json()["request"]["finance_review"] == "部分付款"


def test_concurrent_payments_cannot_exceed_request_amount():
    with TestClient(app) as setup_client:
        login(setup_client)
        batch = setup_client.post(
            "/api/batches",
            json={"name": "concurrent-payment-test", "start_date": "2026-07-01", "end_date": "2026-07-07"},
        ).json()["batch"]
        request = setup_client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"summary": "并发付款", "amount": 100},
        ).json()["request"]

    def submit_payment() -> int:
        with TestClient(app) as client:
            login(client)
            response = client.post(
                f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
                json={"amount": 80, "payment_date": "2026-07-10"},
            )
            return response.status_code

    with ThreadPoolExecutor(max_workers=2) as executor:
        statuses = sorted(executor.map(lambda _: submit_payment(), range(2)))
    assert statuses == [200, 400]
    with TestClient(app) as client:
        login(client)
        detail = client.get(f"/api/batches/{batch['id']}/requests/{request['id']}/payments").json()
        assert detail["summary"]["paid_amount"] == 80
        assert detail["summary"]["pending_amount"] == 20
        assert detail["summary"]["payment_count"] == 1


def test_admin_can_restore_archived_batch_to_draft():
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "status-test", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        batch_id = batch["id"]
        archived = client.post(f"/api/batches/{batch_id}/archive")
        assert archived.status_code == 200
        assert archived.json()["batch"]["status"] == "archived"
        restored = client.post(f"/api/batches/{batch_id}/unarchive")
        assert restored.status_code == 200
        assert restored.json()["batch"]["status"] == "draft"
        logs = client.get(f"/api/batches/{batch_id}/audit")
        assert logs.status_code == 200
        actions = [log["action"] for log in logs.json()["logs"]]
        assert "batch.archive" in actions
        assert "batch.unarchive" in actions


def test_all_roles_can_change_own_password_and_other_sessions_are_revoked():
    accounts = [
        ("self-password-business", "业务改密", "business", "business-old", "business-new"),
        ("self-password-finance", "财务改密", "finance", "finance-old", "finance-new"),
        ("self-password-manager", "总经理改密", "general_manager", "manager-old", "manager-new"),
    ]
    with TestClient(app) as admin_client:
        login(admin_client)
        created_user_ids = []
        for username, display_name, role, old_password, _ in accounts:
            created = admin_client.post(
                "/api/admin/users",
                json={"username": username, "password": old_password, "display_name": display_name, "role": role, "active": True},
            )
            assert created.status_code == 200
            created_user_ids.append(created.json()["user"]["id"])

        business = accounts[0]
        with TestClient(app) as current_client, TestClient(app) as other_client:
            login(current_client, business[0], business[3])
            login(other_client, business[0], business[3])

            assert current_client.post(
                "/api/auth/change-password",
                json={"current_password": "", "new_password": business[4], "confirm_password": business[4]},
            ).status_code == 400
            assert current_client.post(
                "/api/auth/change-password",
                json={"current_password": business[3], "new_password": "12345", "confirm_password": "12345"},
            ).status_code == 400
            assert current_client.post(
                "/api/auth/change-password",
                json={"current_password": business[3], "new_password": business[4], "confirm_password": "different"},
            ).status_code == 400
            assert current_client.post(
                "/api/auth/change-password",
                json={"current_password": "wrong-password", "new_password": business[4], "confirm_password": business[4]},
            ).status_code == 400
            assert current_client.post(
                "/api/auth/change-password",
                json={"current_password": business[3], "new_password": business[3], "confirm_password": business[3]},
            ).status_code == 400

            changed = current_client.post(
                "/api/auth/change-password",
                json={"current_password": business[3], "new_password": business[4], "confirm_password": business[4]},
            )
            assert changed.status_code == 200
            assert changed.json() == {"status": "ok", "signed_out_sessions": 1}
            assert current_client.get("/api/me").status_code == 200
            assert other_client.get("/api/me").status_code == 401

        old_login = admin_client.post("/api/auth/login", json={"username": business[0], "password": business[3]})
        assert old_login.status_code == 401
        with TestClient(app) as new_login_client:
            login(new_login_client, business[0], business[4])

        for username, _, _, old_password, new_password in accounts[1:]:
            with TestClient(app) as role_client:
                login(role_client, username, old_password)
                changed = role_client.post(
                    "/api/auth/change-password",
                    json={"current_password": old_password, "new_password": new_password, "confirm_password": new_password},
                )
                assert changed.status_code == 200
                assert role_client.get("/api/me").status_code == 200
            with TestClient(app) as new_login_client:
                login(new_login_client, username, new_password)

        admin_changed = admin_client.post(
            "/api/auth/change-password",
            json={"current_password": "admin123", "new_password": "admin456", "confirm_password": "admin456"},
        )
        assert admin_changed.status_code == 200
        assert admin_client.get("/api/me").status_code == 200
        admin_restored = admin_client.post(
            "/api/auth/change-password",
            json={"current_password": "admin456", "new_password": "admin123", "confirm_password": "admin123"},
        )
        assert admin_restored.status_code == 200

        with connect() as conn:
            audit_rows = conn.execute(
                "SELECT old_value_json, new_value_json FROM audit_logs WHERE action = 'user.change_password'"
            ).fetchall()
        assert len(audit_rows) >= 5
        audit_payload = json.dumps([dict(row) for row in audit_rows], ensure_ascii=False)
        assert "password_hash" not in audit_payload
        for password in [value for account in accounts for value in account[3:]] + ["admin123", "admin456"]:
            assert password not in audit_payload
        for user_id in created_user_ids:
            assert admin_client.delete(f"/api/admin/users/{user_id}").status_code == 200


def test_role_permissions_user_crud_and_audit_logs():
    with TestClient(app) as admin_client, TestClient(app) as business_client, TestClient(app) as finance_client, TestClient(app) as manager_client:
        login(admin_client)
        assert admin_client.get("/api/me").json()["user"]["role"] == "admin"
        with connect() as conn:
            active_roles = {
                row["value"]
                for row in conn.execute("SELECT value FROM dictionaries WHERE kind = 'role' AND active = 1").fetchall()
            }
        assert active_roles == {"business", "finance", "general_manager", "admin"}
        admin_users = admin_client.get("/api/admin/users")
        assert admin_users.status_code == 200

        business_user = admin_client.post(
            "/api/admin/users",
            json={
                "username": "biz-user",
                "password": "biz123",
                "display_name": "业务",
                "role": "business",
                "active": True,
                "sheet_permissions": ["未分 Sheet"],
            },
        )
        assert business_user.status_code == 200
        finance_user = admin_client.post(
            "/api/admin/users",
            json={"username": "finance-user", "password": "fin123", "display_name": "财务", "role": "finance", "active": True},
        )
        assert finance_user.status_code == 200
        finance_user_id = finance_user.json()["user"]["id"]
        manager_user = admin_client.post(
            "/api/admin/users",
            json={"username": "gm-user", "password": "gm123", "display_name": "总经理", "role": "general_manager", "active": True},
        )
        assert manager_user.status_code == 200
        manager_user_id = manager_user.json()["user"]["id"]

        batch = admin_client.post("/api/batches", json={"name": "permission-test", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        batch_id = batch["id"]
        created = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"summary": "权限测试", "amount": 10, "general_manager_approval": "同意付款"},
        )
        assert created.status_code == 200
        request_id = created.json()["request"]["id"]

        login(business_client, "biz-user", "biz123")
        blocked_business = business_client.patch(
            f"/api/batches/{batch_id}/requests/{request_id}",
            json={"summary": "业务改摘要", "finance_review": "已付款"},
        )
        assert blocked_business.status_code == 400
        allowed_business = business_client.patch(
            f"/api/batches/{batch_id}/requests/{request_id}",
            json={"summary": "业务改摘要"},
        )
        assert allowed_business.status_code == 200
        assert allowed_business.json()["request"]["summary"] == "业务改摘要"
        blocked_business_create = business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={
                "summary": "业务新增",
                "amount": 20,
                "finance_review": "已付款",
                "actual_payment_date": "2026-07-10",
                "general_manager_opinion": "不应保存",
            },
        )
        assert blocked_business_create.status_code == 400
        business_create = business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"summary": "业务新增", "amount": 20},
        )
        assert business_create.status_code == 200
        business_row = business_create.json()["request"]
        assert business_row["finance_review"] == "未付款"
        assert business_row["payment_status"] == "未付款"
        assert business_row["actual_payment_date"] is None
        assert business_row["general_manager_opinion"] is None

        login(finance_client, "finance-user", "fin123")
        blocked_payment_api = business_client.post(
            f"/api/batches/{batch_id}/requests/{request_id}/payments",
            json={"amount": 1, "payment_date": "2026-07-10"},
        )
        assert blocked_payment_api.status_code == 403
        partial_finance_update = finance_client.post(
            f"/api/batches/{batch_id}/requests/{request_id}/payments",
            json={"amount": 4, "payment_date": "2026-07-10"},
        )
        assert partial_finance_update.status_code == 200
        assert partial_finance_update.json()["request"]["finance_review"] == "部分付款"
        finance_update = finance_client.post(
            f"/api/batches/{batch_id}/requests/{request_id}/payments",
            json={"amount": 6, "payment_date": "2026-07-11"},
        )
        assert finance_update.status_code == 200
        assert finance_update.json()["request"]["finance_review"] == "已付款"
        blocked_finance = finance_client.patch(
            f"/api/batches/{batch_id}/requests/{request_id}",
            json={"general_manager_opinion": "财务不能改总经理意见"},
        )
        assert blocked_finance.status_code == 403

        login(manager_client, "gm-user", "gm123")
        manager_users = manager_client.get("/api/admin/users")
        assert manager_users.status_code == 200
        manager_update = manager_client.patch(
            f"/api/batches/{batch_id}/requests/{request_id}",
            json={"general_manager_opinion": "总经理可改"},
        )
        assert manager_update.status_code == 200
        assert manager_update.json()["request"]["general_manager_opinion"] == "总经理可改"

        reset = admin_client.patch(f"/api/admin/users/{finance_user_id}", json={"password": "fin456", "display_name": "财务改"})
        assert reset.status_code == 200
        assert reset.json()["user"]["display_name"] == "财务改"
        relogin = TestClient(app)
        with relogin as client:
            login(client, "finance-user", "fin456")
        default_reset = admin_client.post(f"/api/admin/users/{finance_user_id}/reset-password")
        assert default_reset.status_code == 200
        assert default_reset.json()["password"] == "123456"
        reset_login = TestClient(app)
        with reset_login as client:
            login(client, "finance-user", "123456")

        blocked_self_delete = admin_client.delete("/api/admin/users/1")
        assert blocked_self_delete.status_code == 400
        business_user_id = business_user.json()["user"]["id"]
        deactivated = admin_client.patch(f"/api/admin/users/{business_user_id}", json={"active": False})
        assert deactivated.status_code == 200
        listed_after_deactivate = admin_client.get("/api/admin/users").json()["users"]
        inactive_business = next(user for user in listed_after_deactivate if user["id"] == business_user_id)
        assert inactive_business["active"] is False
        inactive_login = business_client.post("/api/auth/login", json={"username": "biz-user", "password": "biz123"})
        assert inactive_login.status_code == 401
        deleted_business = admin_client.delete(f"/api/admin/users/{business_user_id}")
        assert deleted_business.status_code == 200
        listed_after_delete = admin_client.get("/api/admin/users").json()["users"]
        assert all(user["id"] != business_user_id for user in listed_after_delete)
        manager_deactivated = admin_client.delete(f"/api/admin/users/{manager_user_id}")
        assert manager_deactivated.status_code == 200
        blocked_last_privileged = admin_client.patch("/api/admin/users/1", json={"role": "finance"})
        assert blocked_last_privileged.status_code == 400

        logs = admin_client.get(f"/api/batches/{batch_id}/audit").json()["logs"]
        actions = [log["action"] for log in logs]
        assert "request.create" in actions
        assert "request.update" in actions
        assert "payment.create" in actions
        with connect() as conn:
            user_actions = {
                row["action"]
                for row in conn.execute("SELECT action FROM audit_logs WHERE entity_type = 'user'").fetchall()
            }
        assert {"user.create", "user.update", "user.deactivate", "user.delete", "user.reset_password"} <= user_actions


def test_business_users_are_strictly_isolated_by_sheet_permissions():
    png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wl2KpAAAAAASUVORK5CYII="
    )
    with TestClient(app) as admin_client, TestClient(app) as business_client:
        login(admin_client)
        batch = admin_client.post(
            "/api/batches",
            json={"name": "Sheet 权限隔离", "start_date": "2026-07-21", "end_date": "2026-07-28"},
        ).json()["batch"]
        batch_id = batch["id"]

        sheet_a = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "部门 A", "summary": "A 请款", "amount": 100},
        ).json()["request"]
        sheet_b = admin_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "部门 B", "summary": "B 请款", "amount": 200},
        ).json()["request"]
        assert admin_client.put(
            f"/api/batches/{batch_id}/sheet-order",
            json={"sheet_order": ["部门 B", "部门 A"]},
        ).status_code == 200

        attachment_a = admin_client.post(
            f"/api/batches/{batch_id}/requests/{sheet_a['id']}/attachments/image",
            files={"file": ("a.png", io.BytesIO(png_bytes), "image/png")},
        ).json()["attachment"]
        attachment_b = admin_client.post(
            f"/api/batches/{batch_id}/requests/{sheet_b['id']}/attachments/image",
            files={"file": ("b.png", io.BytesIO(png_bytes), "image/png")},
        ).json()["attachment"]
        payment_b = admin_client.post(
            f"/api/batches/{batch_id}/requests/{sheet_b['id']}/payments",
            json={"amount": 50, "payment_date": "2026-07-28"},
        ).json()["payment"]
        voucher_b = admin_client.post(
            f"/api/batches/{batch_id}/requests/{sheet_b['id']}/payments/{payment_b['id']}/vouchers",
            files={"file": ("voucher.png", io.BytesIO(png_bytes), "image/png")},
        ).json()["voucher"]

        with connect() as conn:
            attachment_objects = conn.execute(
                "SELECT file_object_id FROM attachment_links WHERE id IN (?, ?) ORDER BY id",
                (attachment_a["id"], attachment_b["id"]),
            ).fetchall()
            voucher_object = conn.execute(
                "SELECT file_object_id FROM payment_vouchers WHERE id = ?",
                (voucher_b["id"],),
            ).fetchone()
        referenced_object_ids = {row["file_object_id"] for row in attachment_objects}
        referenced_object_ids.add(voucher_object["file_object_id"])
        assert len(referenced_object_ids) == 1

        created_user = admin_client.post(
            "/api/admin/users",
            json={
                "username": "sheet-a-user",
                "password": "Yuewei123",
                "display_name": "部门 A 负责人",
                "role": "business",
                "active": True,
                "sheet_permissions": ["部门 A"],
            },
        )
        assert created_user.status_code == 200
        user_id = created_user.json()["user"]["id"]
        assert created_user.json()["user"]["sheet_permissions"] == ["部门 A"]
        login(business_client, "sheet-a-user", "Yuewei123")
        assert business_client.get("/api/me").json()["user"]["sheet_permissions"] == ["部门 A"]

        listed_batch = next(
            item for item in business_client.get("/api/batches").json()["batches"]
            if item["id"] == batch_id
        )
        assert listed_batch["request_count"] == 1
        assert listed_batch["total_amount"] == 100
        assert listed_batch["sheet_order"] == ["部门 A"]
        batch_detail = business_client.get(f"/api/batches/{batch_id}").json()
        assert batch_detail["batch"]["request_count"] == 1
        assert batch_detail["batch"]["sheet_order"] == ["部门 A"]

        requests = business_client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        assert [row["id"] for row in requests] == [sheet_a["id"]]
        assert business_client.get(
            f"/api/batches/{batch_id}/requests",
            params={"source_sheet": "部门 B"},
        ).json()["requests"] == []
        attachments = business_client.get(f"/api/batches/{batch_id}/attachments").json()["attachments"]
        assert [item["id"] for item in attachments] == [attachment_a["id"]]

        assert business_client.get(
            f"/api/batches/{batch_id}/requests/{sheet_b['id']}/payments"
        ).status_code == 403
        assert business_client.get(
            f"/api/batches/{batch_id}/requests/{sheet_b['id']}/attachments"
        ).status_code == 403
        assert business_client.get(f"/api/attachments/{attachment_b['id']}/file").status_code == 403
        assert business_client.get(f"/api/payment-vouchers/{voucher_b['id']}/file").status_code == 403
        assert business_client.get(f"/api/batches/{batch_id}/audit").status_code == 403

        exported = business_client.get(f"/api/batches/{batch_id}/export.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
        assert workbook.sheetnames[:2] == ["全部", "部门 A"]
        assert "部门 B" not in workbook.sheetnames

        assert business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "部门 B", "summary": "越权新增", "amount": 1},
        ).status_code == 403
        allowed_create = business_client.post(
            f"/api/batches/{batch_id}/requests",
            json={"source_sheet": "部门 A", "summary": "允许新增", "amount": 1},
        )
        assert allowed_create.status_code == 200
        assert business_client.patch(
            f"/api/batches/{batch_id}/requests/{sheet_a['id']}",
            json={"source_sheet": "部门 B"},
        ).status_code == 403

        changed = admin_client.patch(
            f"/api/admin/users/{user_id}",
            json={"sheet_permissions": ["部门 B"]},
        )
        assert changed.status_code == 200
        assert changed.json()["user"]["sheet_permissions"] == ["部门 B"]
        switched_rows = business_client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        assert [row["id"] for row in switched_rows] == [sheet_b["id"]]
        assert business_client.get(f"/api/attachments/{attachment_a['id']}/file").status_code == 403
        assert business_client.get(f"/api/attachments/{attachment_b['id']}/file").status_code == 200


def test_image_attachment_upload_and_export():
    png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "image-export-test", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        batch_id = batch["id"]
        created = client.post(
            f"/api/batches/{batch_id}/requests",
            json={
                "dingding_id": "IMG-1",
                "summary": "带图片附件",
                "amount": 100,
                "source_sheet": "凌翔产品&开发",
                "general_manager_approval_date": "2026-07-10",
            },
        )
        assert created.status_code == 200
        request_id = created.json()["request"]["id"]
        uploaded = client.post(
            f"/api/batches/{batch_id}/requests/{request_id}/attachments/image",
            data={"label": "付款截图"},
            files={"file": ("proof.png", io.BytesIO(png_bytes), "image/png")},
        )
        assert uploaded.status_code == 200
        attachment = uploaded.json()["attachment"]
        assert attachment["attachment_type"] == "image"
        assert attachment["file_url"].endswith(f"/api/attachments/{attachment['id']}/file")
        listed = client.get(f"/api/batches/{batch_id}/attachments")
        assert listed.status_code == 200
        assert len(listed.json()["attachments"]) == 1
        downloaded = client.get(attachment["file_url"])
        assert downloaded.status_code == 200
        assert "inline" in downloaded.headers["content-disposition"].lower()
        assert downloaded.content == png_bytes
        exported = client.get(f"/api/batches/{batch_id}/export.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(io.BytesIO(exported.content))
        worksheet = workbook["凌翔产品&开发"]
        headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
        assert "图片附件" in headers
        assert "财务付款时间" in headers
        assert "总经理审批时间" in headers
        assert "总经理意见" in headers
        assert worksheet._images


def test_restore_draft_baseline_restores_saved_rows_and_attachments():
    png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    second_png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAIAAAACCAYAAABytg0kAAAAFElEQVR42mP8z8BQz0AEYBxVSFIBADuVBAXnTKiNAAAAAElFTkSuQmCC"
    )
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "restore-test", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        batch_id = batch["id"]
        created = client.post(
            f"/api/batches/{batch_id}/requests",
            json={
                "dingding_id": "RESTORE-1",
                "summary": "基线记录",
                "amount": 100,
                "source_sheet": "凌翔产品&开发",
                "expected_payment_account": "基线人工账户",
            },
        )
        assert created.status_code == 200
        request_id = created.json()["request"]["id"]
        uploaded = client.post(
            f"/api/batches/{batch_id}/requests/{request_id}/attachments/image",
            files={"file": ("baseline.png", io.BytesIO(png_bytes), "image/png")},
        )
        assert uploaded.status_code == 200
        baseline_attachment = uploaded.json()["attachment"]
        baseline_file_path = TEST_DIR / "attachment-storage" / baseline_attachment["file_path"]
        assert baseline_file_path.exists()
        baseline = client.post(f"/api/batches/{batch_id}/snapshots/baseline")
        assert baseline.status_code == 200
        assert baseline.json()["snapshot"]["request_count"] == 1
        assert baseline.json()["snapshot"]["attachment_count"] == 1

        updated = client.patch(
            f"/api/batches/{batch_id}/requests/{request_id}",
            json={
                "summary": "已经保存的错误修改",
                "amount": 999,
                "expected_payment_account": "错误人工账户",
            },
        )
        assert updated.status_code == 200
        extra = client.post(
            f"/api/batches/{batch_id}/requests",
            json={"dingding_id": "RESTORE-EXTRA", "summary": "应被还原删除", "amount": 1},
        )
        assert extra.status_code == 200
        extra_id = extra.json()["request"]["id"]
        second_attachment = client.post(
            f"/api/batches/{batch_id}/requests/{extra_id}/attachments/image",
            files={"file": ("extra.png", io.BytesIO(second_png_bytes), "image/png")},
        )
        assert second_attachment.status_code == 200
        extra_file_path = TEST_DIR / "attachment-storage" / second_attachment.json()["attachment"]["file_path"]
        assert extra_file_path.exists()
        deleted_attachment = client.delete(f"/api/batches/{batch_id}/requests/{request_id}/attachments/{baseline_attachment['id']}")
        assert deleted_attachment.status_code == 200
        # Removing a logical relationship must not eagerly delete a shared hash object.
        assert baseline_file_path.exists()

        restored = client.post(f"/api/batches/{batch_id}/restore-baseline")
        assert restored.status_code == 200
        payload = restored.json()
        assert payload["before"]["requests"] == 2
        assert payload["after"]["requests"] == 1
        rows = client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        assert len(rows) == 1
        assert rows[0]["dingding_id"] == "RESTORE-1"
        assert rows[0]["summary"] == "基线记录"
        assert rows[0]["amount"] == 100
        assert rows[0]["expected_payment_account"] == "基线人工账户"
        assert rows[0]["expected_payment_account_source"] == "manual"
        attachments = client.get(f"/api/batches/{batch_id}/attachments").json()["attachments"]
        assert len(attachments) == 1
        assert attachments[0]["id"] == baseline_attachment["id"]
        assert baseline_file_path.exists()
        assert extra_file_path.exists()
        downloaded = client.get(attachments[0]["file_url"])
        assert downloaded.status_code == 200
        assert downloaded.content == png_bytes
        with connect() as conn:
            pre_restore_count = conn.execute(
                "SELECT COUNT(*) AS count FROM batch_snapshots WHERE batch_id = ? AND snapshot_type = 'pre_restore'",
                (batch_id,),
            ).fetchone()["count"]
            assert pre_restore_count == 1
        actions = [log["action"] for log in client.get(f"/api/batches/{batch_id}/audit").json()["logs"]]
        assert "batch.snapshot.baseline" in actions
        assert "batch.restore_baseline" in actions


def test_payment_vouchers_snapshot_rollover_and_archived_corrections():
    png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "payment-snapshot-test", "start_date": "2026-07-01", "end_date": "2026-07-07"},
        ).json()["batch"]
        batch_id = batch["id"]
        request = client.post(
            f"/api/batches/{batch_id}/requests",
            json={"dingding_id": "PAY-SNAPSHOT-1", "summary": "付款快照", "amount": 100, "source_sheet": "付款测试"},
        ).json()["request"]
        payment = client.post(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments",
            json={"amount": 40, "payment_date": "2026-07-10", "payer": "出纳A", "bank_reference": "FLOW-1"},
        ).json()["payment"]
        image_voucher = client.post(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}/vouchers",
            files={"file": ("proof.png", io.BytesIO(png_bytes), "image/png")},
        )
        assert image_voucher.status_code == 200
        assert image_voucher.json()["voucher"]["voucher_type"] == "image"
        pdf_voucher = client.post(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}/vouchers",
            files={"file": ("proof.pdf", io.BytesIO(b"%PDF-1.4\n%%EOF"), "application/pdf")},
        )
        assert pdf_voucher.status_code == 200
        assert pdf_voucher.json()["voucher"]["voucher_type"] == "pdf"
        baseline = client.post(f"/api/batches/{batch_id}/snapshots/baseline")
        assert baseline.status_code == 200
        assert baseline.json()["snapshot"]["payment_count"] == 1
        assert baseline.json()["snapshot"]["payment_voucher_count"] == 2

        changed = client.patch(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}",
            json={"amount": 60, "payment_date": "2026-07-11"},
        )
        assert changed.status_code == 200
        deleted_voucher = client.delete(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}/vouchers/{image_voucher.json()['voucher']['id']}"
        )
        assert deleted_voucher.status_code == 200
        restored = client.post(f"/api/batches/{batch_id}/restore-baseline")
        assert restored.status_code == 200
        restored_detail = client.get(f"/api/batches/{batch_id}/requests/{request['id']}/payments").json()
        assert restored_detail["summary"]["paid_amount"] == 40
        assert restored_detail["summary"]["pending_amount"] == 60
        assert len(restored_detail["payments"]) == 1
        assert len(restored_detail["payments"][0]["vouchers"]) == 2

        rollover = client.post(
            f"/api/batches/{batch_id}/rollover",
            json={"name": "payment-rollover", "start_date": "2026-07-08", "end_date": "2026-07-14", "copy_mode": "all"},
        )
        assert rollover.status_code == 200
        target_batch_id = rollover.json()["batch"]["id"]
        target_request = client.get(f"/api/batches/{target_batch_id}/requests").json()["requests"][0]
        inherited = client.get(f"/api/batches/{target_batch_id}/requests/{target_request['id']}/payments").json()["payments"][0]
        assert inherited["inherited"] is True
        assert inherited["amount"] == 40
        assert len(inherited["vouchers"]) == 2
        blocked_inherited = client.patch(
            f"/api/batches/{target_batch_id}/requests/{target_request['id']}/payments/{inherited['id']}",
            json={"amount": 30, "payment_date": "2026-07-12"},
        )
        assert blocked_inherited.status_code == 400

        source_changed = client.patch(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}",
            json={"amount": 50, "payment_date": "2026-07-12"},
        )
        assert source_changed.status_code == 200
        target_after_source_change = client.get(
            f"/api/batches/{target_batch_id}/requests/{target_request['id']}/payments"
        ).json()["payments"][0]
        assert target_after_source_change["amount"] == 40

        assert client.post(f"/api/batches/{batch_id}/archive").status_code == 200
        missing_reason = client.patch(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}",
            json={"amount": 45, "payment_date": "2026-07-12"},
        )
        assert missing_reason.status_code == 400
        corrected = client.patch(
            f"/api/batches/{batch_id}/requests/{request['id']}/payments/{payment['id']}",
            json={"amount": 45, "payment_date": "2026-07-12", "reason": "银行回单金额更正"},
        )
        assert corrected.status_code == 200
        logs = client.get(f"/api/batches/{batch_id}/audit").json()["logs"]
        assert any(log["action"] == "payment.correction" and log["reason"] == "银行回单金额更正" for log in logs)


def test_excel_payment_detail_import_export_duplicate_and_rollback(tmp_path):
    png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    voucher_path = tmp_path / "excel-proof.png"
    voucher_path.write_bytes(png_bytes)
    record = {
        "id": 999999,
        "dingding_id": "EXCEL-PAY-1",
        "summary": "Excel 分次付款",
        "amount": 100,
        "paid_amount": 30,
        "pending_amount": 70,
        "finance_review": "部分付款",
        "source_sheet": "Excel付款测试",
    }
    detail = {
        "request_id": 999999,
        "dingding_id": "EXCEL-PAY-1",
        "request_source_sheet": "Excel付款测试",
        "payment_date": "2026-07-10",
        "amount": 30,
        "payer": "Excel出纳",
        "payment_account": "公户",
        "bank_reference": "EXCEL-FLOW-1",
        "remark": "首笔",
        "source_type": "excel_detail",
        "vouchers": [
            {
                "id": 1,
                "payment_id": 1,
                "original_filename": "excel-proof.png",
                "mime_type": "image/png",
                "file_url": "/api/payment-vouchers/1/file",
                "absolute_path": str(voucher_path),
            }
        ],
    }
    workbook_bytes = export_workbook(
        {"name": "Excel付款导入", "end_date": "2026-07-07"},
        [record],
        {},
        [detail, {**detail, "vouchers": []}],
    )
    with TestClient(app) as client:
        login(client)
        imported = client.post(
            "/api/import/weekly-excel",
            files={
                "file": (
                    "excel-payment-detail.xlsx",
                    io.BytesIO(workbook_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert imported.status_code == 200
        payload = imported.json()
        batch_id = payload["batch_id"]
        assert payload["meta"]["payment_details"]["imported"] == 1
        assert payload["meta"]["payment_details"]["duplicates"] == 1
        assert payload["meta"]["payment_details"]["saved_vouchers"] == 1
        request = client.get(f"/api/batches/{batch_id}/requests").json()["requests"][0]
        assert request["paid_amount"] == 30
        assert request["pending_amount"] == 70
        assert request["payment_count"] == 1
        payments = client.get(f"/api/batches/{batch_id}/requests/{request['id']}/payments").json()["payments"]
        assert len(payments) == 1
        assert len(payments[0]["vouchers"]) == 1

        exported = client.get(f"/api/batches/{batch_id}/export.xlsx")
        assert exported.status_code == 200
        exported_workbook = load_workbook(io.BytesIO(exported.content))
        payment_sheet = exported_workbook["付款明细"]
        headers = [cell.value for cell in payment_sheet[1]]
        assert headers[:12] == ["付款标识", "请款标识", "钉钉单号", "来源 Sheet", "付款日期", "本次金额", "付款人", "付款账户", "流水号", "备注", "来源标记", "凭证信息"]
        assert payment_sheet.cell(2, 6).value == 30
        assert len(payment_sheet._images) == 1

        rolled_back = client.post(f"/api/batches/{batch_id}/imports/latest/rollback")
        assert rolled_back.status_code == 200
        assert rolled_back.json()["deleted_payments"] == 1
        assert rolled_back.json()["deleted_payment_vouchers"] == 1
        assert client.get(f"/api/batches/{batch_id}/requests").json()["totals"]["count"] == 0


def test_restore_and_delete_draft_require_privileged_role():
    with TestClient(app) as admin_client, TestClient(app) as business_client:
        login(admin_client)
        business_user = admin_client.post(
            "/api/admin/users",
            json={"username": "restore-biz", "password": "biz123", "display_name": "业务还原", "role": "business", "active": True},
        )
        assert business_user.status_code == 200
        batch = admin_client.post("/api/batches", json={"name": "restore-permission", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        login(business_client, "restore-biz", "biz123")
        blocked_restore = business_client.post(f"/api/batches/{batch['id']}/restore-baseline")
        assert blocked_restore.status_code == 403
        blocked_delete = business_client.delete(f"/api/batches/{batch['id']}")
        assert blocked_delete.status_code == 403


def test_delete_draft_batch_and_keep_archived_batches():
    png_bytes = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    with TestClient(app) as client:
        login(client)
        draft = client.post("/api/batches", json={"name": "delete-me", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        batch_id = draft["id"]
        request = client.post(
            f"/api/batches/{batch_id}/requests",
            json={"dingding_id": "DEL-1", "summary": "删除草稿", "amount": 12},
        ).json()["request"]
        uploaded = client.post(
            f"/api/batches/{batch_id}/requests/{request['id']}/attachments/image",
            files={"file": ("proof.png", io.BytesIO(png_bytes), "image/png")},
        )
        assert uploaded.status_code == 200
        file_path = TEST_DIR / "attachment-storage" / uploaded.json()["attachment"]["file_path"]
        assert file_path.exists()
        deleted = client.delete(f"/api/batches/{batch_id}")
        assert deleted.status_code == 200
        assert client.get(f"/api/batches/{batch_id}").status_code == 404
        # Content-addressed blobs are retained until an explicit verified cleanup.
        assert file_path.exists()

        archived = client.post("/api/batches", json={"name": "keep-me", "start_date": "2026-07-08", "end_date": "2026-07-14"}).json()["batch"]
        archived_id = archived["id"]
        assert client.post(f"/api/batches/{archived_id}/archive").status_code == 200
        blocked = client.delete(f"/api/batches/{archived_id}")
        assert blocked.status_code == 400
        assert client.get(f"/api/batches/{archived_id}").status_code == 200


def test_rebuild_weekly_data_resets_business_data():
    if not SAMPLE.exists():
        return
    with TestClient(app) as client:
        login(client)
        old_batch = client.post("/api/batches", json={"name": "old-data", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        created = client.post(
            f"/api/batches/{old_batch['id']}/requests",
            json={"dingding_id": "OLD-1", "summary": "旧数据", "amount": 1},
        )
        assert created.status_code == 200
        result = rebuild_weekly_data(SAMPLE)
        assert result["requests"] == 161
        assert result["attachments"] == 9
        assert result["batch"]["status"] == "draft"
        batches = client.get("/api/batches").json()["batches"]
        assert len(batches) == 1
        batch = batches[0]
        assert batch["id"] == 1
        assert batch["name"] == "20260626~20260707请款明细"
        assert batch["status"] == "draft"
        assert batch["request_count"] == 161
        attachments = client.get(f"/api/batches/{batch['id']}/attachments").json()["attachments"]
        assert len(attachments) == 9
        assert all(attachment["attachment_type"] == "image" for attachment in attachments)
        with connect() as conn:
            assert conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"] > 0
            assert conn.execute("SELECT COUNT(*) AS count FROM dictionaries").fetchone()["count"] > 0
            assert conn.execute("SELECT COUNT(*) AS count FROM import_jobs").fetchone()["count"] == 1
            assert conn.execute("SELECT COUNT(*) AS count FROM audit_logs").fetchone()["count"] == 1


def test_normalize_payment_data_repairs_saved_rows_and_dictionary():
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "normalize-test", "start_date": "2026-07-01", "end_date": "2026-07-07"}).json()["batch"]
        timestamp = now_iso()
        with connect() as conn:
            conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, dingding_id, amount, finance_review, general_manager_approval,
                    payment_status, remark, content_hash, raw_extra_json, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch["id"],
                    "FIX-1",
                    14500,
                    "已经支付14500元",
                    "我要审核一下\n审核意见反馈给小吴了。",
                    "",
                    "",
                    "bad-hash-1",
                    '{"财务审核": null, "财务审核#2": "已经支付14500元", "总经理批复": "我要审核一下\\n审核意见反馈给小吴了。"}',
                    timestamp,
                    timestamp,
                ),
            )
            conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, dingding_id, amount, finance_review, general_manager_approval,
                    payment_status, remark, content_hash, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch["id"],
                    "FIX-2",
                    50440,
                    "无票",
                    "2026.06.30 Tiffany垫付了50440元",
                    "同意支付",
                    "赵高雅",
                    "bad-hash-2",
                    timestamp,
                    timestamp,
                ),
            )
            conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, dingding_id, amount, payment_status, remark, content_hash,
                    raw_extra_json, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch["id"],
                    "FIX-3",
                    178750,
                    "已支付",
                    "已支付10万\nTiffany垫付的178750元",
                    "bad-hash-3",
                    '{"付款情况": null, "备注": "已支付10万", "备注#2": "Tiffany垫付的178750元"}',
                    timestamp,
                    timestamp,
                ),
            )
            conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, dingding_id, amount, general_manager_approval,
                    payment_status, remark, content_hash, raw_extra_json, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch["id"],
                    "FIX-4",
                    25000,
                    "存在争议，CMA船司在问YUEWEI索要巨款",
                    "未支付",
                    "广州思辉新\n存在争议，CMA船司在问YUEWEI索要巨款",
                    "bad-hash-4",
                    '{"备注": "广州思辉新", "总经理批复": "存在争议，CMA船司在问YUEWEI索要巨款"}',
                    timestamp,
                    timestamp,
                ),
            )
            conn.execute(
                """
                INSERT INTO payment_requests (
                    batch_id, dingding_id, amount, finance_review, actual_payment_date,
                    payment_status, remark, content_hash, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch["id"],
                    "FIX-5",
                    8800,
                    "部分付款",
                    "2026-07-10",
                    "已支付",
                    "",
                    "bad-hash-5",
                    timestamp,
                    timestamp,
                ),
            )
            conn.execute(
                "INSERT OR IGNORE INTO dictionaries (kind, value, active, created_at) VALUES ('payment_status', '待付款', 1, ?)",
                (timestamp,),
            )

        result = normalize_payment_data()
        assert result["changed_rows"] >= 3
        with connect() as conn:
            paid_row = conn.execute("SELECT * FROM payment_requests WHERE dingding_id = 'FIX-1'").fetchone()
            advance_row = conn.execute("SELECT * FROM payment_requests WHERE dingding_id = 'FIX-2'").fetchone()
            raw_extra_row = conn.execute("SELECT * FROM payment_requests WHERE dingding_id = 'FIX-3'").fetchone()
            dispute_row = conn.execute("SELECT * FROM payment_requests WHERE dingding_id = 'FIX-4'").fetchone()
            partial_row = conn.execute("SELECT * FROM payment_requests WHERE dingding_id = 'FIX-5'").fetchone()
            assert paid_row["finance_review"] == "已付款"
            assert paid_row["general_manager_approval"] == "同意付款"
            assert paid_row["payment_status"] == "已付款"
            assert conn.execute("SELECT COUNT(*) AS count FROM payment_records WHERE request_id = ?", (paid_row["id"],)).fetchone()["count"] == 1
            assert "已经支付14500元" in paid_row["remark"]
            assert "我要审核一下" in paid_row["general_manager_opinion"]
            assert not paid_row["remark"] or "我要审核一下" not in paid_row["remark"]
            assert advance_row["finance_review"] == "未付款"
            assert advance_row["general_manager_approval"] is None
            assert advance_row["payment_status"] == "未付款"
            assert "无票" in advance_row["remark"]
            assert "Tiffany垫付" in advance_row["general_manager_opinion"]
            assert "Tiffany垫付" not in advance_row["remark"]
            assert "原付款情况：同意支付" in advance_row["remark"]
            assert raw_extra_row["finance_review"] == "未付款"
            assert raw_extra_row["payment_status"] == "未付款"
            assert "Tiffany垫付" in raw_extra_row["remark"]
            assert dispute_row["general_manager_approval"] == "存在争议"
            assert "CMA船司" in dispute_row["general_manager_opinion"]
            assert dispute_row["remark"] == "广州思辉新"
            assert partial_row["finance_review"] == "未付款"
            assert partial_row["payment_status"] == "未付款"
            assert conn.execute("SELECT active FROM dictionaries WHERE kind = 'payment_status' AND value = '待付款'").fetchone()["active"] == 0
            active_payment_status_values = {
                row["value"]
                for row in conn.execute("SELECT value FROM dictionaries WHERE kind = 'payment_status' AND active = 1").fetchall()
            }
            assert active_payment_status_values == set()
            active_finance_values = {
                row["value"]
                for row in conn.execute("SELECT value FROM dictionaries WHERE kind = 'finance_review' AND active = 1").fetchall()
            }
            assert active_finance_values == {"未付款", "部分付款", "已付款"}


def test_external_expense_mapping_uses_base_currency_and_purchase_form_values():
    purchase = map_external_expense(
        {
            "source_type": "purchase",
            "source_id": "501",
            "effective_date": date(2026, 7, 15),
            "approval_no": "PURCHASE-501",
            "creator_name": "user-purchase-501",
            "applicant_department": "采购部",
            "approval_title": "备用姓名提交的采购支出",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "中国China",
            "beneficiary": None,
            "expense_type": "订单采购",
            "summary": None,
            "project": "项目A",
            "needed_payment_date": None,
            "source_currency": None,
            "source_amount": Decimal("90"),
            "base_currency_amount": Decimal("100.25"),
            "order_name": "订单A",
            "product_name": "产品A",
            "source_created_at": datetime(2026, 7, 15, 9, 10),
            "source_updated_at": datetime(2026, 7, 15, 10, 10),
            "raw_data": {
                "formComponentValues": [
                    {"name": "收款人beneficiario", "value": "供应商A 账号1"},
                    {"name": "收款人beneficiario", "value": "供应商B 账号2"},
                    {"name": "规格明细需求说明Descripción", "value": "采购测试摘要"},
                    {"name": "付款日期Fecha de pago", "value": "2026-07-20"},
                ]
            },
        },
        {"user-purchase-501": "快照表姓名"},
    )
    assert purchase["amount"] == 100.25
    assert purchase["beneficiary"] == "供应商A 账号1 / 供应商B 账号2"
    assert purchase["request_data"]["payee_name"] == purchase["beneficiary"]
    assert purchase["request_data"]["payee_account"] == purchase["beneficiary"]
    assert (
        purchase["request_data"]["raw_extra"]["external_source"]["beneficiary"]
        == purchase["beneficiary"]
    )
    assert purchase["summary"] == "采购测试摘要"
    assert purchase["needed_payment_date"] == "2026-07-20"
    assert "存在多个收款人，请确认" in purchase["warnings"]
    assert purchase["request_data"]["currency"] == "CNY"
    assert purchase["request_data"]["payee_account"] == purchase["beneficiary"]
    assert purchase["request_data"]["raw_extra"]["external_source"]["approval_status"] == "RUNNING"
    assert purchase["applicant"] == "快照表姓名"
    assert purchase["applicant_id"] == "user-purchase-501"
    assert purchase["request_data"]["source_sheet"] == "采购部"
    assert purchase["request_data"]["raw_extra"]["external_source"]["applicant_name_source"] == "ding_user_snapshot"

    refused = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "601",
            "effective_date": date(2026, 7, 15),
            "approval_no": "OP-601",
            "creator_name": "运营申请人",
            "applicant_department": "运营部",
            "approval_title": "Title enviado por Operador Uno",
            "approval_status": "COMPLETED",
            "approval_result": "refuse",
            "execution_region": "中国China",
            "beneficiary": "工资卡",
            "expense_type": "管理费用",
            "summary": "运营测试摘要",
            "project": None,
            "needed_payment_date": "2026-07-21",
            "source_currency": "美元",
            "source_amount": Decimal("10"),
            "base_currency_amount": Decimal("72.3"),
            "source_created_at": datetime(2026, 7, 15, 9, 10),
            "source_updated_at": datetime(2026, 7, 15, 10, 10),
            "raw_data": {},
        }
    )
    assert refused["amount"] == 10.0
    assert refused["request_data"]["currency"] == "USD"
    assert refused["request_data"]["base_amount_cny"] == 72.3
    assert "审批已拒绝、作废或终止，默认不导入" in refused["errors"]
    assert refused["applicant"] == "Operador Uno"

    mexico = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "602",
            "effective_date": date(2026, 8, 12),
            "approval_no": "OP-MX-602",
            "creator_name": "mx-user",
            "applicant_department": "原始墨西哥部门",
            "approval_title": "Solicitud enviado por Mario Gomez",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "墨西哥México",
            "beneficiary": "Proveedor MX",
            "expense_type": "Gastos administrativos",
            "summary": "Prueba México",
            "source_currency": "MXN",
            "source_amount": Decimal("1000"),
            "base_currency_amount": Decimal("400"),
            "raw_data": {},
        }
    )
    assert mexico["errors"] == []
    assert mexico["request_data"]["currency"] == "MXN"
    assert mexico["request_data"]["amount"] == 1000

    mexico_from_region = map_external_expense(
        {
            "source_type": "purchase",
            "source_id": "603",
            "effective_date": date(2026, 8, 12),
            "approval_no": "PURCHASE-MX-603",
            "creator_name": "mx-user",
            "applicant_department": "Compras",
            "approval_title": "Solicitud enviado por Mario Gomez",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "墨西哥 México",
            "beneficiary": None,
            "expense_type": "Compra",
            "source_currency": None,
            "source_amount": Decimal("2500"),
            "base_currency_amount": Decimal("1000"),
            "raw_data": {"formComponentValues": [{"name": "收款人", "value": "Proveedor MX"}]},
        }
    )
    assert mexico_from_region["errors"] == []
    assert mexico_from_region["request_data"]["currency"] == "MXN"
    assert mexico_from_region["request_data"]["amount"] == 2500
    assert mexico_from_region["request_data"]["raw_extra"]["external_source"]["currency_source"] == "execution_region"

    explicit_currency_wins = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "604",
            "effective_date": date(2026, 8, 12),
            "approval_no": "OP-MX-USD-604",
            "creator_name": "mx-user",
            "applicant_department": "México",
            "approval_title": "Solicitud enviado por Mario Gomez",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "墨西哥 México",
            "beneficiary": "Proveedor USD",
            "expense_type": "Compra",
            "source_currency": "美元 USD",
            "source_amount": Decimal("100"),
            "base_currency_amount": Decimal("680"),
            "raw_data": {},
        }
    )
    assert explicit_currency_wins["errors"] == []
    assert explicit_currency_wins["request_data"]["currency"] == "USD"
    assert explicit_currency_wins["request_data"]["amount"] == 100
    assert explicit_currency_wins["request_data"]["raw_extra"]["external_source"]["currency_source"] == "approval_currency"

    legacy_mexico_cny_default = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "605",
            "effective_date": date(2026, 8, 12),
            "approval_no": "OP-MX-PESO-605",
            "creator_name": "mx-user",
            "applicant_department": "Impresión UV",
            "approval_title": "Solicitud enviado por Mario Gomez",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "墨西哥 México",
            "beneficiary": "Proveedor MX",
            "expense_type": "Compra",
            "summary": "本地采购货架，合计76800比索（不含税）",
            "source_currency": "CNY",
            "source_amount": Decimal("76800"),
            "base_currency_amount": Decimal("29952"),
            "raw_data": {},
        }
    )
    assert legacy_mexico_cny_default["errors"] == []
    assert legacy_mexico_cny_default["request_data"]["currency"] == "MXN"
    assert legacy_mexico_cny_default["request_data"]["amount"] == 76800
    assert legacy_mexico_cny_default["request_data"]["raw_extra"]["external_source"]["currency_source"] == "summary_text"
    assert legacy_mexico_cny_default["request_data"]["raw_extra"]["external_source"]["source_currency_table"] == "CNY"

    form_currency_overrides_region = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "606",
            "effective_date": date(2026, 8, 12),
            "approval_no": "OP-MX-USD-606",
            "creator_name": "mx-user",
            "applicant_department": "México",
            "approval_title": "Solicitud enviado por Mario Gomez",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "墨西哥 México",
            "beneficiary": "Proveedor USD",
            "expense_type": "Compra",
            "summary": "Pago de servicio",
            "source_currency": "CNY",
            "source_amount": Decimal("100"),
            "base_currency_amount": Decimal("680"),
            "raw_data": {"formComponentValues": [{"name": "币种", "value": "USD"}]},
        }
    )
    assert form_currency_overrides_region["errors"] == []
    assert form_currency_overrides_region["request_data"]["currency"] == "USD"
    assert form_currency_overrides_region["request_data"]["amount"] == 100
    assert form_currency_overrides_region["request_data"]["raw_extra"]["external_source"]["currency_source"] == "approval_currency"

    assert execution_region_is_allowed("中国China")
    assert execution_region_is_allowed("墨西哥Mexico")
    assert execution_region_is_allowed("México")
    assert not execution_region_is_allowed("美国USA")

    for result in ("refuse", "REJECTED", "cancelled", "revoked", "voided", "已作废", "审批已拒绝", "已撤回"):
        assert approval_result_is_disallowed(result)
    for result in ("", "agree", "approved", "同意"):
        assert not approval_result_is_disallowed(result)


def test_monthly_payment_mapping_aggregates_details_and_extracts_attachments():
    table_value = json.dumps([
        {"rowValue": [
            {"label": "付款日期", "value": "2026-08-12"},
            {"label": "付款金额", "value": "1000"},
            {"label": "付款说明", "value": "第二笔"},
            {"label": "合同附件", "value": json.dumps([{"fileId": "detail-file", "fileName": "明细合同.pdf"}])},
        ]},
        {"rowValue": [
            {"label": "付款日期", "value": "2026-08-11"},
            {"label": "付款金额", "value": "1820"},
            {"label": "付款说明", "value": "第一笔"},
        ]},
    ], ensure_ascii=False)
    raw_payload = {
        "businessId": "202608101140000516846",
        "originatorUserId": "03413943004221481424",
        "originatorDeptName": "原部门",
        "formComponentValues": [
            {"name": "申请事由", "value": "电脑租赁月结"},
            {"name": "付款分类", "value": "办公用品采购付款"},
            {"name": "付款账户类型", "value": "公账付款"},
            {"name": "币种", "value": '["人民币"]'},
            {"name": "收款账户信息", "value": "测试公司 账号123"},
            {"name": "申请付款明细", "value": table_value},
            {"name": "合计总额（元）", "value": "2820"},
            {"name": "附件 Annex", "value": json.dumps([
                {"fileId": "231753805378", "fileName": "20260507电脑租赁合同.pdf", "fileSize": 516944}
            ], ensure_ascii=False)},
            {"name": "关联审批单", "value": '[{"businessId":"202608091234000000001"}]'},
        ],
    }
    instance = {
        "source_id": "35",
        "process_instance_id": "proc-monthly-35",
        "process_code": "PROC-EE85EDD4-5CF2-4C08-B948-1690A6ACC51C",
        "create_time": datetime.fromisoformat("2026-08-10T03:40:00+00:00"),
        "updated_at": datetime.fromisoformat("2026-08-10T04:00:00+00:00"),
        "status": "RUNNING",
        "result": "agree",
        "title": "施鸣坤提交的月结付款",
        "originator_user_id": "03413943004221481424",
        "raw_payload": raw_payload,
    }
    mapped = map_monthly_payment(instance, {"03413943004221481424": "施鸣坤"})
    assert mapped["source_type"] == "monthly"
    assert mapped["source_label"] == "月结付款"
    assert mapped["approval_no"] == "202608101140000516846"
    assert mapped["applicant"] == "施鸣坤"
    assert mapped["amount"] == 2820
    assert mapped["currency"] == "CNY"
    assert mapped["needed_payment_date"] == "2026-08-11"
    assert mapped["beneficiary"] == "测试公司 账号123"
    assert mapped["request_data"]["payment_account"] == "公账付款"
    assert mapped["request_data"]["raw_extra"]["external_source"]["monthly_payment_details"][0]["amount"] == 1000
    assert mapped["related_approval_nos"] == ["202608091234000000001"]
    assert "月结包含 2 行付款明细" in "；".join(mapped["warnings"])

    attachments = _monthly_attachments(instance)
    assert {attachment["file_id"] for attachment in attachments} == {"231753805378", "detail-file"}
    contract = next(attachment for attachment in attachments if attachment["file_id"] == "231753805378")
    assert contract["file_name"] == "20260507电脑租赁合同.pdf"
    assert contract["file_size"] == 516944


def test_monthly_payment_infers_mxn_from_execution_region(monkeypatch):
    def fake_fetch_rates(selected_date, currencies):
        assert selected_date == date(2026, 8, 10)
        assert currencies == ["MXN"]
        return {
            "MXN": {
                "currency": "MXN",
                "cny_per_unit": 0.4,
                "requested_date": "2026-08-10",
                "actual_date": "2026-08-08",
                "fallback": True,
            }
        }

    monkeypatch.setattr("backend.app.external_expenses.fetch_rates", fake_fetch_rates)
    instance = {
        "source_id": "36",
        "process_instance_id": "proc-monthly-36",
        "create_time": datetime.fromisoformat("2026-08-10T03:40:00+00:00"),
        "updated_at": datetime.fromisoformat("2026-08-10T04:00:00+00:00"),
        "status": "RUNNING",
        "result": "agree",
        "title": "Mario提交的月结付款",
        "raw_payload": {
            "businessId": "202608101140000516847",
            "originatorUserId": "mx-monthly-user",
            "originatorDeptName": "México",
            "formComponentValues": [
                {"name": "执行地区 Región de ejecución", "value": "墨西哥 México"},
                {"name": "合计总额", "value": "2500"},
                {"name": "收款账户信息", "value": "Proveedor MX"},
            ],
        },
    }

    mapped = map_monthly_payment(instance, {"mx-monthly-user": "Mario"})
    assert mapped["errors"] == []
    assert mapped["currency"] == "MXN"
    assert mapped["amount"] == 2500
    assert mapped["request_data"]["base_amount_cny"] == 1000
    assert mapped["request_data"]["fx_rate_actual_date"] == "2026-08-08"
    assert mapped["request_data"]["raw_extra"]["external_source"]["currency_source"] == "execution_region"


def test_external_expense_applicant_title_patterns():
    assert applicant_name_from_title("张三提交的运营支出") == "张三"
    assert applicant_name_from_title("Solicitud enviado por Zhang San") == "Zhang San"
    assert applicant_name_from_title("Expense submitted by John Smith") == "John Smith"
    assert applicant_name_from_title("Jane Doe's Purchase Expense") == "Jane Doe"
    assert applicant_name_from_title("无法识别的标题") == ""
    assert not valid_applicant_name("unknown")
    assert not valid_applicant_name("未识别人员")
    assert valid_applicant_name("王道伦")

    fallback = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "title-fallback",
            "effective_date": date(2026, 7, 15),
            "approval_no": "TITLE-FALLBACK",
            "creator_name": "user-with-placeholder",
            "applicant_department": "财务部",
            "approval_title": "王道伦提交的运营支出",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "中国China",
            "beneficiary": "测试收款人",
            "expense_type": "管理费用",
            "summary": "测试摘要",
            "base_currency_amount": Decimal("100"),
            "raw_data": {},
        },
        {"user-with-placeholder": "unknown"},
    )
    assert fallback["applicant"] == "王道伦"
    assert fallback["request_data"]["raw_extra"]["external_source"]["applicant_name_source"] == "approval_title"


def test_external_expense_exact_approval_number_ignores_dates():
    exact_sql, exact_params = _preview_conditions(None, None, ["operation", "purchase"], " 202607071704000140246 ", [])
    assert "effective_date" not in exact_sql
    assert "BTRIM(approval_no) = %s" in exact_sql
    assert "base_currency_amount <> 0" in exact_sql
    assert "execution_region" in exact_sql
    assert "BTRIM(COALESCE(execution_region, '')) = ''" in exact_sql
    assert exact_params[-3] == r"(中国|china)"
    assert "!~* %s" in exact_sql
    assert exact_params[-1] == "202607071704000140246"

    dated_sql, dated_params = _preview_conditions(date(2026, 7, 1), date(2026, 7, 15), ["operation"], "", [])
    assert "effective_date BETWEEN %s AND %s" in dated_sql
    assert dated_params[:2] == [date(2026, 7, 1), date(2026, 7, 15)]


def test_china_workbench_external_expense_region_boundary():
    explicit_china = external_expense_test_row(
        "REGION-CN",
        "9101",
        execution_region="中国China",
        source_sheet="YW MOLDES MX模具",
    )
    explicit_mexico = external_expense_test_row(
        "REGION-MX",
        "9102",
        execution_region="墨西哥México",
        source_sheet="悦为智能 YW Tech_Ai",
    )
    mexico_sheet_fallback = external_expense_test_row(
        "REGION-MX-SHEET",
        "9103",
        source_sheet="YW MOLDES MX模具",
    )
    unknown_legacy = external_expense_test_row("REGION-LEGACY", "9104")

    assert china_workbench_external_expense_allowed(explicit_china)
    assert not china_workbench_external_expense_allowed(explicit_mexico)
    assert not china_workbench_external_expense_allowed(mexico_sheet_fallback)
    assert china_workbench_external_expense_allowed(unknown_legacy)

    mark_china_workbench_external_expense(explicit_mexico)
    mark_china_workbench_external_expense(explicit_mexico)
    assert explicit_mexico["errors"].count(CHINA_WORKBENCH_REGION_ERROR) == 1


def test_external_expense_preview_excludes_mexico_monthly_rows_and_applicants(monkeypatch):
    def instance(source_id: str, approval_no: str, user_id: str, department: str, region: str) -> dict:
        return {
            "source_id": source_id,
            "status": "RUNNING",
            "result": "agree",
            "title": f"{user_id} submitted monthly payment",
            "raw_payload": {
                "businessId": approval_no,
                "originatorUserId": user_id,
                "originatorDeptName": department,
                "formComponentValues": [{"name": "执行地区", "value": region}],
            },
        }

    china_instance = instance(
        "9301",
        "MONTHLY-CN",
        "china-user",
        "YW MOLDES MX模具",
        "中国China",
    )
    mexico_instance = instance(
        "9302",
        "MONTHLY-MX",
        "mexico-user",
        "YW MOLDES MX模具",
        "墨西哥México",
    )

    monkeypatch.setattr(
        external_expenses_module,
        "_monthly_payment_query",
        lambda **kwargs: [china_instance, mexico_instance],
    )
    monkeypatch.setattr(
        external_expenses_module,
        "fetch_dingtalk_user_names",
        lambda user_ids: {
            "china-user": "中国申请人",
            "mexico-user": "Mexico Applicant",
        },
    )

    def fake_map_monthly(row, user_names):
        raw_payload = row["raw_payload"]
        region = raw_payload["formComponentValues"][0]["value"]
        return external_expense_test_row(
            raw_payload["businessId"],
            row["source_id"],
            source_type="monthly",
            execution_region=region,
            source_sheet=raw_payload["originatorDeptName"],
        ) | {
            "applicant_id": raw_payload["originatorUserId"],
            "applicant": user_names[raw_payload["originatorUserId"]],
        }

    monkeypatch.setattr(
        external_expenses_module,
        "map_monthly_payment",
        fake_map_monthly,
    )

    result = external_expenses_module.preview_external_expenses(
        date_from=date(2026, 8, 25),
        date_to=date(2026, 8, 25),
        source_types=["monthly"],
    )

    assert [row["approval_no"] for row in result["rows"]] == ["MONTHLY-CN"]
    assert result["applicant_options"] == [
        {
            "id": "china-user",
            "name": "中国申请人",
            "department": "YW MOLDES MX模具",
            "count": 1,
        }
    ]


def test_external_expense_zero_amount_is_not_importable():
    zero_amount = map_external_expense(
        {
            "source_type": "operation",
            "source_id": "zero-amount",
            "effective_date": date(2026, 7, 15),
            "approval_no": "ZERO-AMOUNT",
            "creator_name": "zero-user",
            "approval_title": "零金额申请人提交的运营支出",
            "applicant_department": "测试部门",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "中国China",
            "beneficiary": "测试收款人",
            "base_currency_amount": Decimal("0"),
            "raw_data": {},
        }
    )
    assert "应付金额为 0，暂不导入" in zero_amount["errors"]
    assert "应付金额为 0" not in zero_amount["warnings"]


def test_external_expense_preview_import_global_dedupe_and_rollback(monkeypatch):
    duplicate_row = external_expense_test_row("EXT-DUPLICATE", "701")
    new_row = external_expense_test_row("EXT-NEW", "702", status="COMPLETED", warnings=["缺少收款信息"], beneficiary="")
    new_row["payment_account"] = "公户"
    new_row["project"] = "导入项目"
    new_row["request_data"]["payment_account"] = "公户"
    new_row["request_data"]["project"] = "导入项目"
    new_row["request_data"]["raw_extra"]["external_source"].update({
        "payment_account": "公户", "project": "导入项目",
    })
    invalid_row = external_expense_test_row("EXT-INVALID", "703")
    invalid_row["errors"] = ["金额缺失"]

    def fake_preview(**kwargs):
        if kwargs["approval_no"] == "EXT-NEW":
            assert kwargs["date_from"] is None
            assert kwargs["date_to"] is None
        else:
            assert kwargs["date_from"] == date(2026, 7, 1)
            assert kwargs["date_to"] == date(2026, 7, 15)
        assert kwargs["source_types"] == ["operation", "purchase"]
        assert kwargs["applicant_ids"] == []
        return {
            "rows": [duplicate_row, new_row, invalid_row],
            "applicant_options": [{"id": "test-user-id", "name": "测试申请人", "department": "测试部门", "count": 3}],
        }

    def fake_fetch(items):
        by_id = {"701": duplicate_row, "702": new_row, "703": invalid_row}
        return [by_id[item["source_id"]] for item in items if item["source_id"] in by_id]

    monkeypatch.setattr(main_module, "preview_external_expenses", fake_preview)
    monkeypatch.setattr(main_module, "fetch_external_expenses", fake_fetch)

    with TestClient(app) as client:
        login(client)
        old_batch = client.post("/api/batches", json={"name": "external-old", "start_date": "2026-06-01", "end_date": "2026-06-07"}).json()["batch"]
        existing = client.post(
            f"/api/batches/{old_batch['id']}/requests",
            json={"dingding_id": " EXT-DUPLICATE ", "summary": "历史记录", "amount": 1},
        )
        assert existing.status_code == 200
        target_batch = client.post("/api/batches", json={"name": "external-target", "start_date": "2026-07-01", "end_date": "2026-07-15"}).json()["batch"]
        target_id = target_batch["id"]

        preview_filter = {
            "batch_id": target_id,
            "date_from": "2026-07-01",
            "date_to": "2026-07-15",
            "source_types": ["operation", "purchase"],
            "approval_no": "",
            "applicant_ids": [],
            "page": 1,
            "page_size": 50,
        }
        preview = client.post("/api/external-expenses/preview", json=preview_filter)
        assert preview.status_code == 200
        payload = preview.json()
        expected_summary = {"matched": 3, "importable": 1, "duplicates": 1, "warnings": 1, "invalid": 1}
        assert payload["summary"] == expected_summary
        assert payload["pagination"]["total"] == 3
        assert [row["approval_no"] for row in payload["all_rows"]] == ["EXT-DUPLICATE", "EXT-NEW", "EXT-INVALID"]
        duplicate_preview = next(row for row in payload["rows"] if row["approval_no"] == "EXT-DUPLICATE")
        new_preview = next(row for row in payload["rows"] if row["approval_no"] == "EXT-NEW")
        assert duplicate_preview["importable"] is False
        assert duplicate_preview["duplicate"]["batch_name"] == "external-old"
        assert new_preview["importable"] is True

        exact_number_without_dates = client.post(
            "/api/external-expenses/preview",
            json={**preview_filter, "date_from": "", "date_to": "not-a-date", "approval_no": "EXT-NEW"},
        )
        assert exact_number_without_dates.status_code == 200

        expected_filtered_rows = {
            "importable": ["EXT-NEW"],
            "duplicates": ["EXT-DUPLICATE"],
            "warnings": ["EXT-NEW"],
            "invalid": ["EXT-INVALID"],
        }
        for result_filter, expected_approval_nos in expected_filtered_rows.items():
            filtered = client.post(
                "/api/external-expenses/preview",
                json={**preview_filter, "result_filter": result_filter},
            )
            assert filtered.status_code == 200
            filtered_payload = filtered.json()
            assert filtered_payload["summary"] == expected_summary
            assert [row["approval_no"] for row in filtered_payload["rows"]] == expected_approval_nos
            assert [row["approval_no"] for row in filtered_payload["all_rows"]] == ["EXT-DUPLICATE", "EXT-NEW", "EXT-INVALID"]
            assert filtered_payload["pagination"]["total"] == len(expected_approval_nos)

        invalid_result_filter = client.post(
            "/api/external-expenses/preview",
            json={**preview_filter, "result_filter": "unknown"},
        )
        assert invalid_result_filter.status_code == 400

        invalid_range = client.post(
            "/api/external-expenses/preview",
            json={"batch_id": target_id, "date_from": "2026-01-01", "date_to": "2026-03-01", "source_types": ["operation"]},
        )
        assert invalid_range.status_code == 400

        imported = client.post(
            f"/api/batches/{target_id}/imports/external-expenses",
            json={"items": [{"source_type": "operation", "source_id": "702"}]},
        )
        assert imported.status_code == 200
        assert imported.json()["imported_rows"] == 1
        assert imported.json()["warnings"] == 1
        requests = client.get(f"/api/batches/{target_id}/requests").json()["requests"]
        assert len(requests) == 1
        request = requests[0]
        assert request["dingding_id"] == "EXT-NEW"
        assert request["paid_amount"] == 0
        assert request["pending_amount"] == new_row["amount"]
        assert request["finance_review"] == "未付款"
        assert request["raw_extra"]["external_source"]["approval_status"] == "COMPLETED"
        assert request["payment_account"] == "公户"
        assert request["project"] == "导入项目"
        with connect() as conn:
            job = conn.execute("SELECT * FROM import_jobs WHERE id = ?", (imported.json()["job_id"],)).fetchone()
            assert job["kind"] == "external-expenses"
            audit = conn.execute("SELECT 1 FROM audit_logs WHERE action = 'import.external_expenses' AND batch_id = ?", (target_id,)).fetchone()
            assert audit

        repeated = client.post(
            f"/api/batches/{target_id}/imports/external-expenses",
            json={"items": [{"source_type": "operation", "source_id": "702"}]},
        )
        assert repeated.status_code == 200
        assert repeated.json()["imported_rows"] == 0
        assert repeated.json()["duplicate_rows"] == 1
        assert repeated.json()["job_id"] is None

        rolled_back = client.post(f"/api/batches/{target_id}/imports/latest/rollback")
        assert rolled_back.status_code == 200
        assert rolled_back.json()["deleted_requests"] == 1
        assert client.get(f"/api/batches/{target_id}/requests").json()["totals"]["count"] == 0


def test_external_expense_import_rejects_mexico_after_source_refetch(monkeypatch):
    mexico_row = external_expense_test_row(
        "IMPORT-MX-GUARD",
        "9201",
        execution_region="墨西哥México",
        source_sheet="YW MOLDES MX模具",
    )
    monkeypatch.setattr(
        main_module,
        "fetch_external_expenses",
        lambda items: [mexico_row],
    )

    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={
                "name": "Mexico import guard",
                "start_date": "2026-08-25",
                "end_date": "2026-08-25",
            },
        ).json()["batch"]
        response = client.post(
            f"/api/batches/{batch['id']}/imports/external-expenses",
            json={"items": [{"source_type": "operation", "source_id": "9201"}]},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["imported_rows"] == 0
        assert payload["invalid_rows"] == 1
        assert payload["job_id"] is None
        assert payload["errors"] == [
            {
                "source_type": "operation",
                "source_id": "9201",
                "messages": [CHINA_WORKBENCH_REGION_ERROR],
            }
        ]
        assert client.get(
            f"/api/batches/{batch['id']}/requests"
        ).json()["requests"] == []


def test_external_expense_concurrent_dedupe_and_source_failure_are_atomic(monkeypatch):
    concurrent_row = external_expense_test_row("EXT-CONCURRENT", "801")
    fetch_barrier = Barrier(2)

    def concurrent_fetch(items):
        fetch_barrier.wait(timeout=10)
        return [concurrent_row]

    monkeypatch.setattr(main_module, "fetch_external_expenses", concurrent_fetch)

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "external-concurrent", "start_date": "2026-07-15", "end_date": "2026-07-15"}).json()["batch"]

    def submit_import():
        with TestClient(app) as thread_client:
            login(thread_client)
            response = thread_client.post(
                f"/api/batches/{batch['id']}/imports/external-expenses",
                json={"items": [{"source_type": "operation", "source_id": "801"}]},
            )
            return response

    with ThreadPoolExecutor(max_workers=2) as executor:
        responses = list(executor.map(lambda _: submit_import(), range(2)))
    assert sorted(response.status_code for response in responses) == [200, 409]
    successful = next(response for response in responses if response.status_code == 200)
    conflicted = next(response for response in responses if response.status_code == 409)
    assert successful.json()["imported_rows"] == 1
    assert conflicted.json()["detail"]["code"] in {
        "BATCH_OPERATION_IN_PROGRESS",
        "VERSION_CONFLICT",
    }
    with connect() as conn:
        assert conn.execute("SELECT COUNT(*) AS count FROM payment_requests WHERE TRIM(dingding_id) = 'EXT-CONCURRENT'").fetchone()["count"] == 1

    def fail_source(items):
        raise ExternalExpenseError("来源数据库暂时不可用")

    monkeypatch.setattr(main_module, "fetch_external_expenses", fail_source)
    with TestClient(app) as client:
        login(client)
        failed_batch = client.post("/api/batches", json={"name": "external-failure", "start_date": "2026-07-15", "end_date": "2026-07-15"}).json()["batch"]
        failed = client.post(
            f"/api/batches/{failed_batch['id']}/imports/external-expenses",
            json={"items": [{"source_type": "operation", "source_id": "999"}]},
        )
        assert failed.status_code == 502
        assert client.get(f"/api/batches/{failed_batch['id']}/requests").json()["totals"]["count"] == 0
        with connect() as conn:
            assert conn.execute("SELECT COUNT(*) AS count FROM import_jobs WHERE batch_id = ?", (failed_batch["id"],)).fetchone()["count"] == 0


def test_expected_payment_account_sync_transition_matrix(monkeypatch):
    def metadata_item(
        approval_no,
        source_id,
        expected_value=None,
        expected_source=None,
    ):
        return {
            "approval_no": approval_no,
            "source_type": "operation",
            "source_label": "运营支出",
            "source_id": source_id,
            "table": "approval_expense_operation",
            "record_id": source_id,
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "execution_region": "中国China",
            "expected_payment_account": expected_value,
            "expected_payment_account_source": expected_source,
        }

    first_metadata = [
        metadata_item(
            "EXPECTED-SYNC-BLANK",
            "9501",
            "悦为智能公司账户",
            "service_subject_default",
        ),
        metadata_item(
            "EXPECTED-SYNC-MANUAL",
            "9502",
            "钉钉明确账户",
            "dingtalk_explicit",
        ),
        metadata_item(
            "EXPECTED-SYNC-DEFAULT",
            "9503",
            "星铭公司账户",
            "service_subject_default",
        ),
        metadata_item(
            "EXPECTED-SYNC-EXPLICIT",
            "9504",
            "钉钉明确账户 B",
            "dingtalk_explicit",
        ),
        metadata_item(
            "EXPECTED-SYNC-LEGACY",
            "9505",
            "新钉钉账户",
            "dingtalk_explicit",
        ),
        metadata_item("EXPECTED-SYNC-UNKNOWN", "9506"),
        metadata_item(
            "EXPECTED-SYNC-CONFLICT",
            "9507",
            "悦为智能公司账户",
            "service_subject_default",
        ),
        metadata_item(
            "EXPECTED-SYNC-CONFLICT",
            "9508",
            "凌翔公司账户",
            "service_subject_default",
        ),
    ]
    metadata_state = {"rows": first_metadata}
    monkeypatch.setattr(
        main_module,
        "fetch_external_expense_metadata",
        lambda approval_nos: metadata_state["rows"],
    )
    monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", lambda approval_nos: [])
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda approval_nos: [])

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "expected-account-sync"}).json()["batch"]
        requests = {}
        for approval_no in (
            "EXPECTED-SYNC-BLANK",
            "EXPECTED-SYNC-MANUAL",
            "EXPECTED-SYNC-DEFAULT",
            "EXPECTED-SYNC-EXPLICIT",
            "EXPECTED-SYNC-LEGACY",
            "EXPECTED-SYNC-UNKNOWN",
            "EXPECTED-SYNC-CONFLICT",
            "EXPECTED-SYNC-UNMATCHED",
        ):
            payload = {
                "dingding_id": approval_no,
                "source_sheet": "悦为智能",
                "amount": 100,
            }
            if approval_no == "EXPECTED-SYNC-MANUAL":
                payload["expected_payment_account"] = "人工账户"
            created = client.post(
                f"/api/batches/{batch['id']}/requests",
                json=payload,
            ).json()["request"]
            requests[approval_no] = created

        with connect() as conn:
            conn.execute(
                """
                UPDATE payment_requests
                SET expected_payment_account = '凌翔公司账户',
                    expected_payment_account_source = 'service_subject_default'
                WHERE id = ?
                """,
                (requests["EXPECTED-SYNC-DEFAULT"]["id"],),
            )
            conn.execute(
                """
                UPDATE payment_requests
                SET expected_payment_account = '钉钉明确账户 A',
                    expected_payment_account_source = 'dingtalk_explicit'
                WHERE id = ?
                """,
                (requests["EXPECTED-SYNC-EXPLICIT"]["id"],),
            )
            conn.execute(
                """
                UPDATE payment_requests
                SET expected_payment_account = '历史人工账户',
                    expected_payment_account_source = NULL
                WHERE id = ?
                """,
                (requests["EXPECTED-SYNC-LEGACY"]["id"],),
            )

        first_sync = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata"
        )
        assert first_sync.status_code == 200
        rows = {
            row["dingding_id"]: row
            for row in client.get(
                f"/api/batches/{batch['id']}/requests",
                params={"dingtalk_lifecycle": "all"},
            ).json()["requests"]
        }
        assert rows["EXPECTED-SYNC-BLANK"]["expected_payment_account"] == "悦为智能公司账户"
        assert rows["EXPECTED-SYNC-BLANK"]["expected_payment_account_source"] == "service_subject_default"
        assert rows["EXPECTED-SYNC-MANUAL"]["expected_payment_account"] == "人工账户"
        assert rows["EXPECTED-SYNC-MANUAL"]["expected_payment_account_source"] == "manual"
        assert rows["EXPECTED-SYNC-DEFAULT"]["expected_payment_account"] == "星铭公司账户"
        assert rows["EXPECTED-SYNC-DEFAULT"]["expected_payment_account_source"] == "service_subject_default"
        assert rows["EXPECTED-SYNC-EXPLICIT"]["expected_payment_account"] == "钉钉明确账户 B"
        assert rows["EXPECTED-SYNC-EXPLICIT"]["expected_payment_account_source"] == "dingtalk_explicit"
        assert rows["EXPECTED-SYNC-LEGACY"]["expected_payment_account"] == "历史人工账户"
        assert rows["EXPECTED-SYNC-LEGACY"]["expected_payment_account_source"] is None
        for approval_no in (
            "EXPECTED-SYNC-UNKNOWN",
            "EXPECTED-SYNC-CONFLICT",
            "EXPECTED-SYNC-UNMATCHED",
        ):
            assert rows[approval_no]["expected_payment_account"] is None
            assert rows[approval_no]["expected_payment_account_source"] is None

        with connect() as conn:
            audit = conn.execute(
                """
                SELECT old_value_json, new_value_json
                FROM audit_logs
                WHERE entity_id = ?
                  AND action = 'external_expenses.metadata_sync.request_fields'
                ORDER BY id DESC LIMIT 1
                """,
                (requests["EXPECTED-SYNC-BLANK"]["id"],),
            ).fetchone()
            assert json.loads(audit["old_value_json"]) == {
                "expected_payment_account": None,
                "expected_payment_account_source": None,
            }
            assert json.loads(audit["new_value_json"]) == {
                "expected_payment_account": "悦为智能公司账户",
                "expected_payment_account_source": "service_subject_default",
            }
            for approval_no in (
                "EXPECTED-SYNC-MANUAL",
                "EXPECTED-SYNC-LEGACY",
                "EXPECTED-SYNC-UNKNOWN",
                "EXPECTED-SYNC-CONFLICT",
                "EXPECTED-SYNC-UNMATCHED",
            ):
                assert conn.execute(
                    """
                    SELECT 1 FROM audit_logs
                    WHERE entity_id = ?
                      AND action = 'external_expenses.metadata_sync.request_fields'
                    """,
                    (requests[approval_no]["id"],),
                ).fetchone() is None

        default_row = rows["EXPECTED-SYNC-DEFAULT"]
        full_form_update = client.patch(
            f"/api/batches/{batch['id']}/requests/{default_row['id']}",
            json={
                "expected_payment_account": default_row["expected_payment_account"],
                "remark": "只修改备注",
                "expected_version": default_row["version"],
            },
        )
        assert full_form_update.status_code == 200
        assert full_form_update.json()["request"]["expected_payment_account_source"] == "service_subject_default"

        metadata_state["rows"] = [
            *[
                item
                for item in first_metadata
                if item["approval_no"] not in {
                    "EXPECTED-SYNC-DEFAULT",
                    "EXPECTED-SYNC-EXPLICIT",
                    "EXPECTED-SYNC-CONFLICT",
                }
            ],
            metadata_item(
                "EXPECTED-SYNC-DEFAULT",
                "9503",
                "凌翔 8899 主账户",
                "dingtalk_explicit",
            ),
            metadata_item("EXPECTED-SYNC-EXPLICIT", "9504"),
        ]
        second_sync = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata"
        )
        assert second_sync.status_code == 200
        second_rows = {
            row["dingding_id"]: row
            for row in client.get(
                f"/api/batches/{batch['id']}/requests",
                params={"dingtalk_lifecycle": "all"},
            ).json()["requests"]
        }
        assert second_rows["EXPECTED-SYNC-DEFAULT"]["expected_payment_account"] == "凌翔 8899 主账户"
        assert second_rows["EXPECTED-SYNC-DEFAULT"]["expected_payment_account_source"] == "dingtalk_explicit"
        assert second_rows["EXPECTED-SYNC-EXPLICIT"]["expected_payment_account"] == "钉钉明确账户 B"
        assert second_rows["EXPECTED-SYNC-EXPLICIT"]["expected_payment_account_source"] == "dingtalk_explicit"

        manual_row = second_rows["EXPECTED-SYNC-MANUAL"]
        cleared = client.patch(
            f"/api/batches/{batch['id']}/requests/{manual_row['id']}",
            json={
                "expected_payment_account": "",
                "expected_version": manual_row["version"],
            },
        )
        assert cleared.status_code == 200
        metadata_state["rows"] = [
            *[
                item
                for item in metadata_state["rows"]
                if item["approval_no"] != "EXPECTED-SYNC-MANUAL"
            ],
            metadata_item(
                "EXPECTED-SYNC-MANUAL",
                "9502",
                "清空后钉钉账户",
                "dingtalk_explicit",
            ),
        ]
        third_sync = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata"
        )
        assert third_sync.status_code == 200
        third_rows = {
            row["dingding_id"]: row
            for row in client.get(
                f"/api/batches/{batch['id']}/requests",
                params={"dingtalk_lifecycle": "all"},
            ).json()["requests"]
        }
        assert third_rows["EXPECTED-SYNC-MANUAL"]["expected_payment_account"] == "清空后钉钉账户"
        assert third_rows["EXPECTED-SYNC-MANUAL"]["expected_payment_account_source"] == "dingtalk_explicit"

        with connect() as conn:
            assert conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM audit_logs
                WHERE entity_id = ?
                  AND action = 'external_expenses.metadata_sync.request_fields'
                """,
                (requests["EXPECTED-SYNC-BLANK"]["id"],),
            ).fetchone()["count"] == 1


def test_identical_expected_account_sync_does_not_bump_business_versions(monkeypatch):
    metadata = [{
        "approval_no": "EXPECTED-SYNC-IDEMPOTENT",
        "source_type": "operation",
        "source_label": "运营支出",
        "source_id": "9599",
        "table": "approval_expense_operation",
        "record_id": "9599",
        "approval_status": "RUNNING",
        "approval_result": "agree",
        "execution_region": "中国China",
        "expected_payment_account": "悦为智能公司账户",
        "expected_payment_account_source": "service_subject_default",
    }]
    workflows = [{
        "approval_no": "EXPECTED-SYNC-IDEMPOTENT",
        "process_instance_id": "expected-sync-process",
        "status": "RUNNING",
        "result": "agree",
        "title": "幂等同步测试",
        "events": [{
            "event_key": "stable-comment",
            "process_instance_id": "expected-sync-process",
            "activity_id": "finance-node",
            "event_type": "ADD_REMARK",
            "stage_name": "评论",
            "result": None,
            "operator_id": "finance-user",
            "operator_name": "测试财务",
            "event_time": "2026-09-03T08:00:00+08:00",
            "sequence_index": 0,
            "comment": "待后续处理",
            "images": [],
            "attachments": [],
            "trusted_finance": False,
            "current": False,
        }],
    }]
    monkeypatch.setattr(main_module, "fetch_external_expense_metadata", lambda approval_nos: metadata)
    monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", lambda approval_nos: workflows)
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda approval_nos: [])

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "expected-account-idempotent"}).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "dingding_id": "EXPECTED-SYNC-IDEMPOTENT",
                "source_sheet": "悦为智能",
                "amount": 100,
            },
        ).json()["request"]

        first = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert first.status_code == 200
        after_first = client.get(
            f"/api/batches/{batch['id']}/requests",
            params={"dingtalk_lifecycle": "all"},
        ).json()["requests"][0]
        batch_after_first = client.get(f"/api/batches/{batch['id']}").json()["batch"]
        with connect() as conn:
            history_after_first = conn.execute(
                "SELECT COUNT(*) AS count FROM payable_history_versions WHERE source_request_id = ?",
                (request["id"],),
            ).fetchone()["count"]

        second = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert second.status_code == 200
        assert second.json()["updated_requests"] == 0
        after_second = client.get(
            f"/api/batches/{batch['id']}/requests",
            params={"dingtalk_lifecycle": "all"},
        ).json()["requests"][0]
        batch_after_second = client.get(f"/api/batches/{batch['id']}").json()["batch"]
        with connect() as conn:
            history_after_second = conn.execute(
                "SELECT COUNT(*) AS count FROM payable_history_versions WHERE source_request_id = ?",
                (request["id"],),
            ).fetchone()["count"]

        assert after_second["version"] == after_first["version"]
        assert batch_after_second["version"] == batch_after_first["version"]
        assert history_after_second == history_after_first


def test_external_expense_metadata_sync_statuses_conflicts_and_atomic_failure(monkeypatch):
    metadata = [
        {
            "approval_no": "SYNC-MATCH",
            "source_type": "operation",
            "source_label": "运营支出",
            "source_id": "901",
            "table": "approval_expense_operation",
            "record_id": "901",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "applicant_id": "user-901",
            "applicant": "同步申请人",
            "applicant_name_source": "ding_user_snapshot",
            "applicant_department": "产品与采购部",
            "payment_account": "公户",
            "project": "钉钉项目 A",
        },
        {
            "approval_no": "SYNC-CONFLICT",
            "source_type": "operation",
            "source_label": "运营支出",
            "source_id": "902",
            "approval_status": "COMPLETED",
            "payment_account": "私户",
            "project": "冲突项目",
        },
        {
            "approval_no": "SYNC-CONFLICT",
            "source_type": "purchase",
            "source_label": "采购支出",
            "source_id": "903",
            "approval_status": "TERMINATED",
            "payment_account": "公户",
            "project": "冲突项目 2",
        },
        {
            "approval_no": "SYNC-TERMINATED",
            "source_type": "purchase",
            "source_label": "采购支出",
            "source_id": "904",
            "approval_status": "TERMINATED",
        },
    ]
    monkeypatch.setattr(main_module, "fetch_external_expense_metadata", lambda approval_nos: metadata)
    monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", lambda approval_nos: [])
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda approval_nos: [])

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "metadata-sync", "start_date": "2026-07-15", "end_date": "2026-07-15"}).json()["batch"]
        batch_id = batch["id"]
        request_ids = []
        for dingding_id, amount, source_sheet in (
            ("SYNC-MATCH", 100, "用户手动 Sheet"),
            ("SYNC-MATCH", 101, "用户手动 Sheet"),
            ("SYNC-CONFLICT", 200, "冲突 Sheet"),
            ("SYNC-MISSING", 300, "未匹配 Sheet"),
            ("SYNC-TERMINATED", 400, "终止 Sheet"),
        ):
            response = client.post(
                f"/api/batches/{batch_id}/requests",
                json={"dingding_id": dingding_id, "amount": amount, "source_sheet": source_sheet},
            )
            assert response.status_code == 200
            request_ids.append(response.json()["request"]["id"])

        manual_applicant = client.patch(
            f"/api/batches/{batch_id}/requests/{request_ids[0]}",
            json={"applicant": "外部供应商申请人"},
        )
        assert manual_applicant.status_code == 200
        assert manual_applicant.json()["request"]["applicant"] == "外部供应商申请人"

        synced = client.post(f"/api/batches/{batch_id}/external-expenses/sync-metadata")
        assert synced.status_code == 200
        synced_payload = synced.json()
        timings = synced_payload.pop("timings")
        assert set(timings) == {"metadata_query_ms", "workflow_query_ms", "attachment_query_ms", "commit_ms"}
        assert all(value >= 0 for value in timings.values())
        assert synced_payload == {
            "status": "synced",
            "batch_id": batch_id,
            "unique_approval_nos": 4,
            "matched": 2,
            "unmatched": 1,
            "conflicts": 1,
            "updated_requests": 5,
            "workflow_events": 0,
            "payment_candidates": 0,
            "auto_payments": 0,
            "review_required": 0,
            "already_applied": 0,
            "skipped": 0,
            "auto_payment_mode": "preview",
            "attachment_downloaded": 0,
            "attachment_synced": 0,
            "attachment_existing": 0,
            "attachment_failed": 0,
            "attachment_errors": [],
        }
        visible_rows = client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        assert request_ids[4] not in {row["id"] for row in visible_rows}
        inactive_rows = client.get(
            f"/api/batches/{batch_id}/requests",
            params={"dingtalk_lifecycle": "inactive"},
        ).json()["requests"]
        assert [row["id"] for row in inactive_rows] == [request_ids[4]]
        rows = client.get(
            f"/api/batches/{batch_id}/requests",
            params={"dingtalk_lifecycle": "all"},
        ).json()["requests"]
        by_id = {row["id"]: row for row in rows}
        matched_source = by_id[request_ids[0]]["raw_extra"]["external_source"]
        assert matched_source["lookup_status"] == "matched"
        assert matched_source["approval_status"] == "RUNNING"
        assert matched_source["applicant"] == "同步申请人"
        assert matched_source["applicant_department"] == "产品与采购部"
        assert by_id[request_ids[0]]["applicant"] == "外部供应商申请人"
        assert by_id[request_ids[0]]["amount"] == 100
        assert by_id[request_ids[0]]["source_sheet"] == "用户手动 Sheet"
        assert by_id[request_ids[2]]["raw_extra"]["external_source"]["lookup_status"] == "conflict"
        assert by_id[request_ids[3]]["raw_extra"]["external_source"]["lookup_status"] == "unmatched"
        assert by_id[request_ids[2]]["payment_account"] is None
        assert by_id[request_ids[2]]["project"] is None
        assert by_id[request_ids[3]]["payment_account"] is None
        assert by_id[request_ids[3]]["project"] is None
        with connect() as conn:
            for request_id in (request_ids[2], request_ids[3]):
                assert conn.execute(
                    "SELECT 1 FROM audit_logs WHERE entity_id = ? AND action = 'external_expenses.metadata_sync.request_fields'",
                    (request_id,),
                ).fetchone() is None
        assert by_id[request_ids[4]]["raw_extra"]["external_source"]["approval_status"] == "TERMINATED"
        assert by_id[request_ids[4]]["general_manager_approval"] == "无需审批"
        visible_batch = client.get(f"/api/batches/{batch_id}").json()["batch"]
        assert visible_batch["request_count"] == 4
        assert visible_batch["total_amount"] == 701
        listed_batch = next(item for item in client.get("/api/batches").json()["batches"] if item["id"] == batch_id)
        assert listed_batch["request_count"] == 4
        exported = client.get(f"/api/batches/{batch_id}/export.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(io.BytesIO(exported.content))
        all_sheet = workbook["全部"]
        headers = [cell.value for cell in all_sheet[2]]
        dingding_column = headers.index("钉钉申请单号") + 1
        exported_dingding_ids = {
            all_sheet.cell(row, dingding_column).value
            for row in range(3, all_sheet.max_row + 1)
        }
        assert "SYNC-TERMINATED" not in exported_dingding_ids
        with connect() as conn:
            audit_row = conn.execute(
                "SELECT new_value_json FROM audit_logs WHERE batch_id = ? AND action = 'external_expenses.metadata_sync'",
                (batch_id,),
            ).fetchone()
            assert audit_row is not None
            audit_payload = json.loads(audit_row["new_value_json"])
            audit_timings = audit_payload.pop("timings")
            assert set(audit_timings) == {
                "metadata_query_ms",
                "workflow_query_ms",
                "attachment_query_ms",
            }
            assert json.loads(json.dumps(audit_payload)) == {
                "unique_approval_nos": 4,
                "matched": 2,
                "unmatched": 1,
                "conflicts": 1,
                "updated_requests": 5,
                "workflow_events": 0,
                "payment_candidates": 0,
                "auto_payments": 0,
                "review_required": 0,
                "already_applied": 0,
                "skipped": 0,
                "auto_payment_mode": "preview",
                "attachment_downloaded": 0,
                "attachment_synced": 0,
                "attachment_existing": 0,
                "attachment_failed": 0,
                "attachment_errors": [],
            }

        refreshed_metadata = [
            {**item, "approval_status": "RUNNING"}
            if item["approval_no"] == "SYNC-TERMINATED"
            else item
            for item in metadata
        ]
        monkeypatch.setattr(main_module, "fetch_external_expense_metadata", lambda approval_nos: refreshed_metadata)
        resynced = client.post(f"/api/batches/{batch_id}/external-expenses/sync-metadata")
        assert resynced.status_code == 200
        rows = client.get(f"/api/batches/{batch_id}/requests").json()["requests"]
        by_id = {row["id"]: row for row in rows}
        assert by_id[request_ids[4]]["raw_extra"]["external_source"]["approval_status"] == "RUNNING"
        assert by_id[request_ids[4]]["general_manager_approval"] is None

        archived = client.post(f"/api/batches/{batch_id}/archive")
        assert archived.status_code == 200
        rejected = client.post(f"/api/batches/{batch_id}/external-expenses/sync-metadata")
        assert rejected.status_code == 400

        failure_batch = client.post("/api/batches", json={"name": "metadata-sync-failure"}).json()["batch"]
        created = client.post(
            f"/api/batches/{failure_batch['id']}/requests",
            json={"dingding_id": "SYNC-FAIL", "amount": 50, "raw_extra": {"kept": True}},
        ).json()["request"]

        def fail_metadata(approval_nos):
            raise ExternalExpenseError("来源数据库暂时不可用")

        monkeypatch.setattr(main_module, "fetch_external_expense_metadata", fail_metadata)
        failed = client.post(f"/api/batches/{failure_batch['id']}/external-expenses/sync-metadata")
        assert failed.status_code == 502
        unchanged = client.get(f"/api/batches/{failure_batch['id']}/requests").json()["requests"][0]
        assert unchanged["id"] == created["id"]
        assert unchanged["raw_extra"] == {"kept": True}


def test_dingtalk_sync_fills_blank_payee_and_manager_fields_without_overwriting_manual_values(monkeypatch):
    approval_nos = (
        "SYNC-FILL-BLANK",
        "SYNC-KEEP-MANUAL",
        "SYNC-IGNORE-INVALID-DATE",
    )
    source_dates = ("2026-07-24", "2026-07-25", "not-a-date")
    metadata = [
        {
            "approval_no": approval_no,
            "source_type": "purchase",
            "source_label": "采购支出",
            "source_id": str(index),
            "approval_status": "RUNNING" if index == 1 else "TERMINATED",
            "approval_result": "agree",
            "beneficiary": "钉钉收款人",
            "execution_region": "中国China",
            "needed_payment_date": source_dates[index - 1],
            "payment_account": ("公户" if index == 1 else "私户" if index == 2 else "未知账户"),
            "project": ("钉钉项目 A" if index == 1 else "  外部项目  " if index == 2 else ""),
        }
        for index, approval_no in enumerate(approval_nos, start=1)
    ]
    workflows = [
        {
            "approval_no": approval_no,
            "process_instance_id": f"process-{index}",
            "status": "RUNNING",
            "result": "agree",
            "events": [
                {
                    "event_key": f"manager-{index}",
                    "process_instance_id": f"process-{index}",
                    "activity_id": "ceo-node",
                    "event_type": "EXECUTE_TASK_NORMAL",
                    "stage_name": "悦为智能 CEO 审批",
                    "result": "AGREE",
                    "operator_id": "ceo-user",
                    "operator_name": "CEO",
                    "event_time": "2026-08-24T16:20:00+08:00",
                    "sequence_index": 3,
                    "comment": None,
                    "images": [],
                    "attachments": [],
                    "trusted_finance": False,
                    "current": False,
                }
            ],
            "current_tasks": [],
        }
        for index, approval_no in enumerate(approval_nos, start=1)
    ]
    monkeypatch.setattr(main_module, "fetch_external_expense_metadata", lambda values: metadata)
    monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", lambda values: workflows)
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda values: [])

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "sync-fill-fields"}).json()["batch"]
        blank = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": approval_nos[0], "source_sheet": "悦为智能", "amount": 100},
        ).json()["request"]
        manual = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "dingding_id": approval_nos[1],
                "source_sheet": "悦为智能",
                "amount": 200,
                "payee_name": "人工收款人",
                "payee_account": "人工账号",
                "needed_payment_date": "2026-08-20",
                "payment_account": "人工账户",
                "project": "人工项目",
                "general_manager_approval": "存在争议",
                "general_manager_approval_date": "2026-08-20",
            },
        ).json()["request"]
        invalid_source = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "dingding_id": approval_nos[2],
                "source_sheet": "悦为智能",
                "amount": 300,
            },
        ).json()["request"]

        response = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert response.status_code == 200
        rows = client.get(
            f"/api/batches/{batch['id']}/requests",
            params={"dingtalk_lifecycle": "all"},
        ).json()["requests"]
        by_id = {row["id"]: row for row in rows}

        assert by_id[blank["id"]]["needed_payment_date"] == "2026-07-24"
        assert by_id[manual["id"]]["needed_payment_date"] == "2026-08-20"
        assert by_id[invalid_source["id"]]["needed_payment_date"] is None
        assert by_id[blank["id"]]["payment_account"] == "公户"
        assert by_id[blank["id"]]["project"] == "钉钉项目 A"
        assert by_id[manual["id"]]["payment_account"] == "人工账户"
        assert by_id[manual["id"]]["project"] == "人工项目"
        assert by_id[invalid_source["id"]]["payment_account"] is None
        assert by_id[invalid_source["id"]]["project"] is None

        with connect() as conn:
            history = conn.execute(
                """
                SELECT needed_payment_date, event_type
                FROM payable_history_versions
                WHERE source_request_id = ? AND event_type = 'dingtalk.sync'
                ORDER BY id DESC
                LIMIT 1
                """,
                (blank["id"],),
            ).fetchone()
            assert history["needed_payment_date"] == "2026-07-24"
            assert history["event_type"] == "dingtalk.sync"
            field_audit = conn.execute(
                """
                SELECT old_value_json, new_value_json
                FROM audit_logs
                WHERE entity_id = ? AND action = 'external_expenses.metadata_sync.request_fields'
                ORDER BY id DESC LIMIT 1
                """,
                (blank["id"],),
            ).fetchone()
            assert json.loads(field_audit["old_value_json"]) == {"payment_account": None, "project": None}
            assert json.loads(field_audit["new_value_json"]) == {"payment_account": "公户", "project": "钉钉项目 A"}
            for request in (manual, invalid_source):
                assert conn.execute(
                    "SELECT 1 FROM audit_logs WHERE entity_id = ? AND action = 'external_expenses.metadata_sync.request_fields'",
                    (request["id"],),
                ).fetchone() is None

        details = client.get(
            "/api/daily-payables/details",
            params={"date": date.today().isoformat()},
        )
        assert details.status_code == 200
        assert any(
            item["dingding_id"] == "SYNC-FILL-BLANK"
            for item in details.json()["items"]
        )

        second = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert second.status_code == 200
        second_rows = client.get(
            f"/api/batches/{batch['id']}/requests",
            params={"dingtalk_lifecycle": "all"},
        ).json()["requests"]
        second_by_id = {row["id"]: row for row in second_rows}
        assert second_by_id[blank["id"]]["needed_payment_date"] == "2026-07-24"
        assert second_by_id[manual["id"]]["needed_payment_date"] == "2026-08-20"
        assert second_by_id[invalid_source["id"]]["needed_payment_date"] is None
        assert second_by_id[blank["id"]]["payment_account"] == "公户"
        assert second_by_id[blank["id"]]["project"] == "钉钉项目 A"
        assert second_by_id[manual["id"]]["payment_account"] == "人工账户"
        assert second_by_id[manual["id"]]["project"] == "人工项目"
        assert second_by_id[invalid_source["id"]]["payment_account"] is None
        assert second_by_id[invalid_source["id"]]["project"] is None
        with connect() as conn:
            assert conn.execute(
                "SELECT COUNT(*) AS count FROM audit_logs WHERE entity_id = ? AND action = 'external_expenses.metadata_sync.request_fields'",
                (blank["id"],),
            ).fetchone()["count"] == 1

    blank_row = by_id[blank["id"]]
    assert blank_row["payee_name"] == "钉钉收款人"
    assert blank_row["payee_account"] == "钉钉收款人"
    assert blank_row["general_manager_approval"] == "同意付款"
    assert blank_row["general_manager_approval_date"] == "2026-08-24"

    manual_row = by_id[manual["id"]]
    assert manual_row["payee_name"] == "人工收款人"
    assert manual_row["payee_account"] == "人工账号"
    assert manual_row["general_manager_approval"] == "存在争议"
    assert manual_row["general_manager_approval_date"] == "2026-08-20"


@pytest.mark.parametrize(
    "stage_name",
    ["悦为智能CEO审批", "悦为智能 CEO 审批", "总经理审批", "Gerente General", "Dirección General"],
)
def test_general_manager_approval_uses_latest_explicit_manager_node(stage_name):
    result = general_manager_approval_from_workflow_events([
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": stage_name,
            "result": "AGREE",
            "event_time": "2026-08-24T16:20:00+08:00",
            "sequence_index": 3,
        }
    ])
    assert result == ("同意付款", "2026-08-24")


def test_general_manager_approval_uses_latest_decision_and_ignores_department_manager():
    assert general_manager_approval_from_workflow_events([
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "采购经理审批",
            "result": "AGREE",
            "event_time": "2026-08-24T09:00:00+08:00",
            "sequence_index": 1,
        }
    ]) is None

    for stage_name in (
        "副总经理审批",
        "总经理助理审批",
        "Subgerente General",
        "CEO助理审批",
    ):
        assert general_manager_approval_from_workflow_events([
            {
                "event_type": "EXECUTE_TASK_NORMAL",
                "stage_name": stage_name,
                "result": "AGREE",
                "event_time": "2026-08-24T09:00:00+08:00",
                "sequence_index": 1,
            }
        ]) is None

    assert general_manager_approval_from_workflow_events([
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "CEO 审批",
            "result": "AGREE",
            "event_time": "2026-08-23T09:00:00+08:00",
            "sequence_index": 1,
        },
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "CEO 审批",
            "result": "REFUSE",
            "event_time": "2026-08-24T09:00:00+08:00",
            "sequence_index": 2,
        },
    ]) is None

    assert general_manager_approval_from_workflow_events([
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "CEO审批",
            "result": "AGREE",
            "event_time": "2026-08-23T09:00:00+08:00",
            "sequence_index": 1,
        },
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "CEO审批",
            "result": "REFUSE",
            "event_time": None,
            "sequence_index": 2,
        },
    ]) is None

    assert general_manager_approval_from_workflow_events([
        {
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "CEO审批",
            "result": "AGREE",
            "event_time": "not-a-date",
            "sequence_index": 1,
        },
    ]) is None


def test_dingtalk_payment_comment_classifier_is_strict():
    trusted_event = {"comment": "该笔已支付，付款截图如上", "trusted_finance": True}
    assert classify_dingtalk_payment_event(
        trusted_event,
        approval_no="202607270000000001",
        pending_amount=60,
        workflow_status="RUNNING",
        workflow_result="agree",
    )[0] == "eligible"
    classification, _ = classify_dingtalk_payment_event(
        {"comment": "发票金额100元，悦为支付", "trusted_finance": True},
        approval_no="202607270943000076977",
        pending_amount=100,
        workflow_status="RUNNING",
        workflow_result="agree",
    )
    assert classification == "eligible"
    for comment in (
        "悦为支付",
        "发票金额100元，待悦为支付",
        "发票金额100元，请悦为支付",
        "发票金额100元，将由悦为支付",
    ):
        classification, _ = classify_dingtalk_payment_event(
            {"comment": comment, "trusted_finance": True},
            approval_no="202607270943000076977",
            pending_amount=100,
            workflow_status="RUNNING",
            workflow_result="agree",
        )
        assert classification == "review_required"
    classification, _ = classify_dingtalk_payment_event(
        {"comment": "发票金额80元，悦为支付", "trusted_finance": True},
        approval_no="202607270943000076977",
        pending_amount=100,
        workflow_status="RUNNING",
        workflow_result="agree",
    )
    assert classification == "review_required"
    for comment in (
        "客户已付款",
        "已支付 40 元，剩余 20 元未支付",
        "已支付5千元",
        "两张审批单合并付款，已支付",
        "审批已支付该笔，无需再次支付",
        "202607270000000002 已付款",
    ):
        classification, _ = classify_dingtalk_payment_event(
            {"comment": comment, "trusted_finance": True},
            approval_no="202607270000000001",
            pending_amount=60,
            workflow_status="RUNNING",
            workflow_result="agree",
        )
        assert classification == "review_required"
    assert classify_dingtalk_payment_event(
        {"comment": "已支付", "trusted_finance": False},
        approval_no="202607270000000001",
        pending_amount=60,
        workflow_status="RUNNING",
        workflow_result="agree",
    )[0] == "ignored"
    classification, reason = classify_dingtalk_payment_event(
        {"comment": "已支付20000元", "trusted_finance": True},
        approval_no="202605061052000163127",
        pending_amount=20000,
        paid_amount=20000,
        workflow_status="RUNNING",
        workflow_result="agree",
    )
    assert classification == "review_required"
    assert "累计已付" in reason


def test_dingtalk_metadata_sync_adds_only_unsynced_request_attachments(monkeypatch):
    approval_no = "SYNC-ATTACHMENT"
    content = b"%PDF-1.4\nsynced attachment\n"
    metadata = [{
        "approval_no": approval_no,
        "source_type": "operation",
        "source_label": "运营支出",
        "source_id": "1201",
        "approval_status": "RUNNING",
        "approval_result": "agree",
        "applicant_id": "user-1201",
        "applicant": "附件申请人",
        "applicant_department": "财务中心",
    }]
    source_attachments = [{
        "source_type": "operation",
        "source_id": "1201",
        "approval_no": approval_no,
        "attachment_id": "501",
        "row_no": 1,
        "file_id": "file-501",
        "file_name": "付款凭证.pdf",
        "file_type": "pdf",
        "file_size": len(content),
        "created_at": "2026-07-27T10:00:00",
    }]
    workflows = [{
        "approval_no": approval_no,
        "process_instance_id": "process-attachment",
        "status": "RUNNING",
        "result": "agree",
        "title": "附件同步测试",
        "events": [],
    }]
    download_calls = []

    class FakeAttachmentClient:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return None

        def download(self, process_instance_id, file_id):
            download_calls.append((process_instance_id, file_id))
            return content, "application/pdf"

    monkeypatch.setattr(main_module, "fetch_external_expense_metadata", lambda approval_nos: metadata)
    monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", lambda approval_nos: workflows)
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda approval_nos: source_attachments)
    monkeypatch.setattr(main_module, "DingtalkAttachmentClient", FakeAttachmentClient)

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "attachment-sync"}).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": approval_no, "amount": 100},
        ).json()["request"]

        first = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert first.status_code == 200
        assert first.json()["attachment_downloaded"] == 1
        assert first.json()["attachment_synced"] == 1
        assert first.json()["attachment_existing"] == 0
        assert first.json()["attachment_failed"] == 0
        assert download_calls == [("process-attachment", "file-501")]

        attachments = client.get(
            f"/api/batches/{batch['id']}/requests/{request['id']}/attachments"
        ).json()["attachments"]
        assert len(attachments) == 1
        attachment = attachments[0]
        assert attachment["source_system"] == "dingtalk_expense_database"
        assert attachment["source_attachment_id"] == "file-501"
        assert attachment["attachment_type"] == "file"
        assert attachment["original_filename"] == "付款凭证.pdf"
        with main_module.connect() as conn:
            stored = conn.execute(
                "SELECT file_object_id, source_instance_id FROM attachment_links WHERE id = ?",
                (attachment["id"],),
            ).fetchone()
            assert stored["file_object_id"] is not None
            assert stored["source_instance_id"] == "process-attachment"
        downloaded = client.get(attachment["file_url"])
        assert downloaded.status_code == 200
        assert downloaded.content == content
        assert downloaded.headers["content-type"].startswith("application/pdf")

        rejected_delete = client.delete(
            f"/api/batches/{batch['id']}/requests/{request['id']}/attachments/{attachment['id']}"
        )
        assert rejected_delete.status_code == 400

        # The middle table may regenerate its own row id.  The DingTalk fileId is
        # the stable identity and must prevent another download/link.
        source_attachments[0]["attachment_id"] = "999"
        source_attachments.append({**source_attachments[0], "attachment_id": "1000"})
        second = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert second.status_code == 200
        assert second.json()["attachment_downloaded"] == 0
        assert second.json()["attachment_synced"] == 0
        assert second.json()["attachment_existing"] == 1
        assert download_calls == [("process-attachment", "file-501")]
        assert len(
            client.get(
                f"/api/batches/{batch['id']}/requests/{request['id']}/attachments"
            ).json()["attachments"]
        ) == 1


def test_dingtalk_workflow_events_respect_stage_and_finance_approval_order():
    operations = [
        {
            "userId": "finance-user",
            "date": "2026-07-27T01:00:00Z",
            "type": "ADD_REMARK",
            "activityId": None,
            "showName": "评论",
            "remark": "已支付",
        },
        {
            "userId": "finance-user",
            "date": "2026-07-27T02:00:00Z",
            "type": "EXECUTE_TASK_NORMAL",
            "activityId": "finance-node",
            "showName": "财务审批",
            "result": "AGREE",
        },
        {
            "userId": "finance-user",
            "date": "2026-07-27T03:00:00Z",
            "type": "ADD_REMARK",
            "activityId": None,
            "showName": "评论",
            "remark": "付款完成",
        },
    ]
    events = _parse_workflow_events(
        "process-order",
        operations,
        {"finance-node"},
        {"finance-user": "测试财务"},
    )
    assert [event["comment"] for event in events] == ["已支付", None, "付款完成"]
    assert [event["stage_name"] for event in events] == ["财务审批前评论", "财务审批", "财务审批后评论"]
    assert [event["sequence_index"] for event in events] == [0, 1, 2]
    assert events[0]["trusted_finance"] is False
    assert events[1]["trusted_finance"] is True
    assert events[2]["trusted_finance"] is True
    assert [event["current"] for event in events] == [False, True, False]


def test_dingtalk_workflow_sync_creates_idempotent_remaining_payment(monkeypatch):
    approval_no = "202607270000000001"
    monkeypatch.setenv("DINGTALK_AUTO_PAYMENT_MODE", "apply")
    monkeypatch.setattr(
        main_module,
        "fetch_external_expense_metadata",
        lambda approval_nos: [{
            "approval_no": approval_no,
            "source_type": "operation",
            "source_label": "运营支出",
            "source_id": "9901",
            "approval_status": "RUNNING",
            "approval_result": "agree",
            "applicant_id": "finance-user",
            "applicant": "测试财务",
            "applicant_department": "财务中心",
            "payment_account": " 公户 ",
            "project": " 钉钉项目自动付款 ",
        }],
    )
    events = [
        {
            "event_key": "finance-agree",
            "process_instance_id": "process-1",
            "activity_id": "finance-node",
            "event_type": "EXECUTE_TASK_NORMAL",
            "stage_name": "财务审批",
            "result": "AGREE",
            "operator_id": "finance-user",
            "operator_name": "测试财务",
            "event_time": "2026-07-27T17:58:00+08:00",
            "comment": None,
            "images": [],
            "attachments": [],
            "trusted_finance": True,
        },
        {
            "event_key": "paid-comment",
            "process_instance_id": "process-1",
            "activity_id": "finance-node",
            "event_type": "ADD_REMARK",
            "stage_name": "评论",
            "result": None,
            "operator_id": "finance-user",
            "operator_name": "测试财务",
            "event_time": "2026-07-27T18:00:00+08:00",
            "comment": "发票金额60元，悦为支付",
            "images": [{"url": "https://example.invalid/proof.png"}],
            "attachments": [],
            "trusted_finance": True,
        },
        {
            "event_key": "complex-comment",
            "process_instance_id": "process-1",
            "activity_id": "finance-node",
            "event_type": "ADD_REMARK",
            "stage_name": "评论",
            "result": None,
            "operator_id": "finance-user",
            "operator_name": "测试财务",
            "event_time": "2026-07-27T18:01:00+08:00",
            "comment": "另一笔部分付款，剩余金额待付",
            "images": [],
            "attachments": [],
            "trusted_finance": True,
        },
        {
            "event_key": "second-paid-comment",
            "process_instance_id": "process-1",
            "activity_id": "finance-node",
            "event_type": "ADD_REMARK",
            "stage_name": "评论",
            "result": None,
            "operator_id": "finance-user",
            "operator_name": "测试财务",
            "event_time": "2026-07-27T18:02:00+08:00",
            "comment": "付款完成",
            "images": [],
            "attachments": [],
            "trusted_finance": True,
        },
    ]
    monkeypatch.setattr(
        main_module,
        "fetch_dingtalk_workflows",
        lambda approval_nos: [{
            "approval_no": approval_no,
            "process_instance_id": "process-1",
            "status": "RUNNING",
            "result": "agree",
            "title": "测试审批",
            "events": events,
        }],
    )
    monkeypatch.setattr(main_module, "fetch_external_expense_attachments", lambda approval_nos: [])

    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "workflow-auto-payment"}).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": approval_no, "amount": 100},
        ).json()["request"]
        manual_payment = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
            json={"amount": 40, "payment_date": "2026-07-26", "payer": "人工出纳"},
        )
        assert manual_payment.status_code == 200

        synced = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert synced.status_code == 200
        result = synced.json()
        assert result["workflow_events"] == 4
        assert result["payment_candidates"] == 1
        assert result["auto_payments"] == 1
        assert result["review_required"] == 1
        payments = client.get(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments"
        ).json()
        assert payments["summary"]["paid_amount"] == 100
        assert payments["summary"]["pending_amount"] == 0
        assert len(payments["payments"]) == 2
        automatic = next(payment for payment in payments["payments"] if payment["source_type"] == "dingtalk_workflow")
        assert automatic["amount"] == 60
        assert automatic["payment_date"] == "2026-07-27"
        assert automatic["payer"] == "测试财务"
        assert automatic["payment_account"] == "公户"
        synced_request = next(
            item for item in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
            if item["id"] == request["id"]
        )
        assert synced_request["project"] == "钉钉项目自动付款"
        batch_after_first_sync = client.get(f"/api/batches/{batch['id']}").json()["batch"]
        with connect() as conn:
            history_after_first_sync = conn.execute(
                "SELECT COUNT(*) AS count FROM payable_history_versions WHERE source_request_id = ?",
                (request["id"],),
            ).fetchone()["count"]

        workflow = client.get(
            f"/api/batches/{batch['id']}/requests/{request['id']}/dingtalk-workflow"
        )
        assert workflow.status_code == 200
        workflow_payload = workflow.json()
        assert workflow_payload["summary"] == {
            "total": 4,
            "active": 4,
            "applied": 1,
            "review_required": 1,
        }
        paid_event = next(event for event in workflow_payload["events"] if event["event_key"] == "paid-comment")
        assert paid_event["payment_record_id"] == automatic["id"]
        assert paid_event["images"] == [{"url": "https://example.invalid/proof.png"}]

        repeated = client.post(f"/api/batches/{batch['id']}/external-expenses/sync-metadata")
        assert repeated.status_code == 200
        assert repeated.json()["auto_payments"] == 0
        assert repeated.json()["already_applied"] == 2
        assert repeated.json()["updated_requests"] == 0
        repeated_payments = client.get(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments"
        ).json()["payments"]
        assert len(repeated_payments) == 2
        request_after_repeated_sync = next(
            item for item in client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
            if item["id"] == request["id"]
        )
        batch_after_repeated_sync = client.get(f"/api/batches/{batch['id']}").json()["batch"]
        with connect() as conn:
            history_after_repeated_sync = conn.execute(
                "SELECT COUNT(*) AS count FROM payable_history_versions WHERE source_request_id = ?",
                (request["id"],),
            ).fetchone()["count"]
        assert request_after_repeated_sync["version"] == synced_request["version"]
        assert batch_after_repeated_sync["version"] == batch_after_first_sync["version"]
        assert history_after_repeated_sync == history_after_first_sync

        fresh = client.post(
            f"/api/batches/{batch['id']}/external-expenses/sync-metadata",
            params={"only_if_stale_seconds": 300},
        )
        assert fresh.status_code == 200
        assert fresh.json()["status"] == "fresh"

        with connect() as conn:
            auto_audits = conn.execute(
                "SELECT COUNT(*) AS count FROM audit_logs WHERE action = 'payment.auto_create_from_dingtalk'",
            ).fetchone()["count"]
            assert auto_audits == 1


def visible_data_sheet(workbook):
    return next(
        sheet
        for sheet in workbook.worksheets
        if sheet.title not in {"全部", "付款明细", "_系统信息"}
    )


def header_lookup(sheet):
    header_row = 2 if sheet.cell(2, 1).value == "序号" else 1
    return header_row, {
        str(sheet.cell(header_row, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(header_row, column).value
    }


def test_weekly_merge_roundtrip_update_add_payment_and_rollback():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "merge-roundtrip", "start_date": "2026-07-16", "end_date": "2026-07-23"},
        ).json()["batch"]
        created = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "dingding_id": "MERGE-001",
                "applicant": "原申请人",
                "payment_account": "私户",
                "expected_payment_account": "原人工账户",
                "summary": "原摘要",
                "amount": 100,
                "source_sheet": "原部门",
            },
        ).json()["request"]
        exported = client.get(f"/api/batches/{batch['id']}/export.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(io.BytesIO(exported.content))
        assert workbook["_系统信息"].sheet_state == "veryHidden"
        sheet = visible_data_sheet(workbook)
        header_row, headers = header_lookup(sheet)
        assert sheet.column_dimensions[sheet.cell(header_row, headers["请款标识"]).column_letter].hidden
        assert sheet.cell(header_row + 1, headers["请款标识"]).value == created["id"]
        assert sheet.cell(header_row + 1, headers["预计支付账户"]).value == "原人工账户"
        payment_sheet = workbook["付款明细"]
        assert payment_sheet.column_dimensions["A"].hidden
        assert payment_sheet.cell(1, 1).value == "付款标识"

        unchanged_bytes = io.BytesIO()
        workbook.save(unchanged_bytes)
        unchanged_bytes.seek(0)
        unchanged_preview = client.post(
            "/api/import/weekly-excel/merge-preview",
            data={"batch_id": str(batch["id"])},
            files={"file": ("roundtrip.xlsx", unchanged_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert unchanged_preview.status_code == 200
        assert unchanged_preview.json()["summary"]["unchanged"] == 1
        assert unchanged_preview.json()["summary"]["conflict"] == 0

        sheet.title = "新部门"
        header_row, headers = header_lookup(sheet)
        sheet.cell(header_row + 1, headers["摘要"], "更新后的摘要")
        sheet.cell(header_row + 1, headers["预计支付账户"], "合并人工账户")
        sheet.cell(header_row + 1, headers["已支付金额"], 20)
        new_row = header_row + 2
        sheet.cell(new_row, headers["请款标识"], None)
        sheet.cell(new_row, headers["钉钉申请单号"], "MERGE-002")
        sheet.cell(new_row, headers["申请人"], "新增申请人")
        sheet.cell(new_row, headers["付款账户"], "公户")
        sheet.cell(new_row, headers["预计支付账户"], "新增人工账户")
        sheet.cell(new_row, headers["摘要"], "新增摘要")
        sheet.cell(new_row, headers["应付金额"], 50)
        sheet.cell(new_row, headers["已支付金额"], 0)
        sheet.cell(new_row, headers["待付款金额"], 50)
        modified = io.BytesIO()
        workbook.save(modified)
        modified.seek(0)

        preview_response = client.post(
            "/api/import/weekly-excel/merge-preview",
            data={"batch_id": str(batch["id"])},
            files={"file": ("modified.xlsx", modified, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["summary"]["create"] == 1
        assert preview["summary"]["update"] == 1
        assert preview["summary"]["payment"] == 1
        payment_key = next(
            key
            for row in preview["rows"]
            if row["dingding_id"] == "MERGE-001"
            for key in row["payment_date_keys"]
        )
        applied = client.post(
            f"/api/import-jobs/{preview['job_id']}/apply-merge",
            json={
                "resolutions": [],
                "payment_dates": {payment_key: "2026-07-23"},
            },
        )
        assert applied.status_code == 200, applied.text
        rows = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
        assert len(rows) == 2
        by_ding = {row["dingding_id"]: row for row in rows}
        assert by_ding["MERGE-001"]["summary"] == "更新后的摘要"
        assert by_ding["MERGE-001"]["expected_payment_account"] == "合并人工账户"
        assert by_ding["MERGE-001"]["expected_payment_account_source"] == "manual"
        assert by_ding["MERGE-001"]["source_sheet"] == "新部门"
        assert by_ding["MERGE-001"]["paid_amount"] == 20
        assert by_ding["MERGE-002"]["summary"] == "新增摘要"
        assert by_ding["MERGE-002"]["expected_payment_account"] == "新增人工账户"
        assert by_ding["MERGE-002"]["expected_payment_account_source"] == "manual"
        assert client.get(f"/api/batches/{batch['id']}").json()["batch"]["sheet_order"][0] == "新部门"

        rolled_back = client.post(f"/api/batches/{batch['id']}/imports/latest/rollback")
        assert rolled_back.status_code == 200, rolled_back.text
        assert rolled_back.json()["restored_requests"] == 1
        restored_rows = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
        assert len(restored_rows) == 1
        assert restored_rows[0]["summary"] == "原摘要"
        assert restored_rows[0]["expected_payment_account"] == "原人工账户"
        assert restored_rows[0]["expected_payment_account_source"] == "manual"
        assert restored_rows[0]["source_sheet"] == "原部门"
        assert restored_rows[0]["paid_amount"] == 0


def test_weekly_merge_legacy_ambiguity_and_concurrent_change():
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "merge-legacy"}).json()["batch"]
        existing_ids = []
        for amount in (100, 200):
            response = client.post(
                f"/api/batches/{batch['id']}/requests",
                json={
                    "dingding_id": "LEGACY-DUP",
                    "payment_account": "私户",
                    "summary": f"原记录 {amount}",
                    "amount": amount,
                    "source_sheet": "旧版 Sheet",
                },
            )
            existing_ids.append(response.json()["request"]["id"])

        legacy = Workbook()
        sheet = legacy.active
        sheet.title = "旧版 Sheet"
        sheet.append(["钉钉申请单号", "申请人", "付款账户", "摘要", "应付金额", "已支付金额", "待付款金额"])
        sheet.append(["LEGACY-DUP", "旧版申请人", "私户", "旧版修改", 180, 0, 180])
        payload = io.BytesIO()
        legacy.save(payload)
        payload.seek(0)
        preview_response = client.post(
            "/api/import/weekly-excel/merge-preview",
            data={"batch_id": str(batch["id"])},
            files={"file": ("legacy.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        row = preview["rows"][0]
        assert row["action"] == "conflict"
        assert {candidate["id"] for candidate in row["candidates"]} == set(existing_ids)

        applied = client.post(
            f"/api/import-jobs/{preview['job_id']}/apply-merge",
            json={
                "resolutions": [{"row_id": row["row_id"], "action": "update", "request_id": existing_ids[0]}],
                "payment_dates": {},
            },
        )
        assert applied.status_code == 200, applied.text
        selected = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"]
        assert next(item for item in selected if item["id"] == existing_ids[0])["summary"] == "旧版修改"

        exported = client.get(f"/api/batches/{batch['id']}/export.xlsx")
        concurrent_preview = client.post(
            "/api/import/weekly-excel/merge-preview",
            data={"batch_id": str(batch["id"])},
            files={"file": ("concurrent.xlsx", io.BytesIO(exported.content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()
        changed = client.patch(
            f"/api/batches/{batch['id']}/requests/{existing_ids[1]}",
            json={"remark": "预览后修改"},
        )
        assert changed.status_code == 200
        rejected = client.post(
            f"/api/import-jobs/{concurrent_preview['job_id']}/apply-merge",
            json={"resolutions": [], "payment_dates": {}},
        )
        assert rejected.status_code == 409


def test_weekly_merge_payment_detail_is_authoritative_and_missing_main_row_is_preserved():
    with TestClient(app) as client:
        login(client)
        batch = client.post("/api/batches", json={"name": "merge-payment-detail"}).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": "MERGE-PAY", "payment_account": "公户", "amount": 100, "source_sheet": "付款测试"},
        ).json()["request"]
        payment = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
            json={"amount": 30, "payment_date": "2026-07-20", "payer": "测试出纳"},
        ).json()["payment"]
        exported = client.get(f"/api/batches/{batch['id']}/export.xlsx")
        workbook = load_workbook(io.BytesIO(exported.content))
        main_sheet = visible_data_sheet(workbook)
        main_header_row, _ = header_lookup(main_sheet)
        main_sheet.delete_rows(main_header_row + 1, 1)
        payment_sheet = workbook["付款明细"]
        payment_headers = {cell.value: cell.column for cell in payment_sheet[1] if cell.value}
        assert payment_sheet.cell(2, payment_headers["付款标识"]).value == payment["id"]
        payment_sheet.cell(2, payment_headers["本次金额"], 40)
        modified = io.BytesIO()
        workbook.save(modified)
        modified.seek(0)
        preview_response = client.post(
            "/api/import/weekly-excel/merge-preview",
            data={"batch_id": str(batch["id"])},
            files={"file": ("payment-update.xlsx", modified, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["summary"]["conflict"] == 0
        assert preview["summary"]["payment"] == 1
        preserved = next(row for row in preview["rows"] if row["request_id"] == request["id"])
        assert preserved["payment_change"]
        assert any("系统记录将保留" in warning for warning in preserved["warnings"])
        applied = client.post(
            f"/api/import-jobs/{preview['job_id']}/apply-merge",
            json={"resolutions": [], "payment_dates": {}},
        )
        assert applied.status_code == 200, applied.text
        updated = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
        assert updated["id"] == request["id"]
        assert updated["paid_amount"] == 40

        rolled_back = client.post(f"/api/batches/{batch['id']}/imports/latest/rollback")
        assert rolled_back.status_code == 200, rolled_back.text
        restored = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
        assert restored["paid_amount"] == 30
        restored_payment = client.get(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments"
        ).json()["payments"][0]
        assert restored_payment["amount"] == 30


def teardown_module():
    shutil.rmtree(TEST_DIR, ignore_errors=True)


def test_currency_conversion_preview_apply_and_roundtrip(monkeypatch):
    def fake_rates(selected_date, currencies):
        values = {"CNY": 1.0, "USD": 6.8, "MXN": 0.4}
        return {
            currency: {
                "currency": currency,
                "cny_per_unit": values[currency],
                "requested_date": selected_date.isoformat(),
                "actual_date": "2026-08-07" if currency != "CNY" else selected_date.isoformat(),
                "fallback": currency != "CNY" and selected_date.isoformat() != "2026-08-07",
            }
            for currency in currencies
        }

    monkeypatch.setattr(main_module, "fetch_rates", fake_rates)
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "currency-conversion", "start_date": "2026-08-01", "end_date": "2026-08-10"},
        ).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"amount": 680, "currency": "CNY", "source_sheet": "汇率测试"},
        ).json()["request"]
        for amount, day in ((100, "2026-08-05"), (240, "2026-08-06")):
            response = client.post(
                f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
                json={"amount": amount, "payment_date": day},
            )
            assert response.status_code == 200

        preview_response = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/currency-conversion/preview",
            json={"target_currency": "USD", "rate_date": "2026-08-08"},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()["preview"]
        assert preview["after"] == {"amount": 100.0, "paid_amount": 50.0, "pending_amount": 50.0}
        assert preview["actual_rate_date"] == "2026-08-07"
        assert preview["used_previous_rate"] is True

        applied = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/currency-conversion/apply",
            json={"target_currency": "USD", "rate_date": "2026-08-08", "expected_updated_at": request["updated_at"]},
        )
        assert applied.status_code == 200
        converted = applied.json()["request"]
        assert converted["currency"] == "USD"
        assert converted["amount"] == 100
        assert converted["paid_amount"] == 50
        assert converted["pending_amount"] == 50
        payments = client.get(f"/api/batches/{batch['id']}/requests/{request['id']}/payments").json()["payments"]
        assert round(sum(item["amount"] for item in payments), 2) == 50
        assert all(item["fx_rate_cny_per_unit"] == 6.8 for item in payments)

        roundtrip = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/currency-conversion/apply",
            json={"target_currency": "CNY", "rate_date": "2026-08-08", "expected_updated_at": converted["updated_at"]},
        )
        assert roundtrip.status_code == 200
        restored = roundtrip.json()["request"]
        assert restored["currency"] == "CNY"
        assert restored["amount"] == 680
        assert restored["paid_amount"] == 340
        assert restored["pending_amount"] == 340


def test_currency_correction_keeps_numeric_amounts_and_updates_cny_value(monkeypatch):
    def fake_rates(selected_date, currencies):
        values = {"CNY": 1.0, "USD": 6.8, "MXN": 0.4}
        return {
            currency: {
                "currency": currency,
                "cny_per_unit": values[currency],
                "requested_date": selected_date.isoformat(),
                "actual_date": selected_date.isoformat(),
                "fallback": False,
            }
            for currency in currencies
        }

    monkeypatch.setattr(main_module, "fetch_rates", fake_rates)
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "currency-correction", "start_date": "2026-08-01", "end_date": "2026-08-10"},
        ).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"amount": 25000, "currency": "CNY", "source_sheet": "币种更正"},
        ).json()["request"]
        payment = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
            json={"amount": 5000, "payment_date": "2026-08-08"},
        )
        assert payment.status_code == 200

        preview_response = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/currency-conversion/preview",
            json={"target_currency": "USD", "rate_date": "2026-08-08", "mode": "correct"},
        )
        assert preview_response.status_code == 200, preview_response.text
        preview = preview_response.json()["preview"]
        assert preview["mode"] == "correct"
        assert preview["before"] == {"amount": 25000.0, "paid_amount": 5000.0, "pending_amount": 20000.0}
        assert preview["after"] == preview["before"]
        assert preview["before_base_amount_cny"] == 25000.0
        assert preview["base_amount_cny"] == 170000.0

        current = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
        applied = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/currency-conversion/apply",
            json={
                "target_currency": "USD",
                "rate_date": "2026-08-08",
                "mode": "correct",
                "expected_updated_at": current["updated_at"],
            },
        )
        assert applied.status_code == 200, applied.text
        corrected = applied.json()["request"]
        assert corrected["currency"] == "USD"
        assert corrected["amount"] == 25000
        assert corrected["paid_amount"] == 5000
        assert corrected["pending_amount"] == 20000
        assert corrected["base_amount_cny"] == 170000
        payments = client.get(f"/api/batches/{batch['id']}/requests/{request['id']}/payments").json()["payments"]
        assert payments[0]["amount"] == 5000
        assert payments[0]["base_amount_cny"] == 34000
        with connect() as conn:
            audit = conn.execute(
                "SELECT action, new_value_json FROM audit_logs WHERE entity_id = ? AND action = 'request.currency_correct' ORDER BY id DESC LIMIT 1",
                (request["id"],),
            ).fetchone()
            assert audit is not None
            assert json.loads(audit["new_value_json"])["mode"] == "correct"


def test_foreign_amount_correction_uses_confirmed_rate_and_keeps_payments(monkeypatch):
    def fake_rates(selected_date, currencies):
        values = {"CNY": 1.0, "USD": 6.8, "MXN": 0.4}
        return {
            currency: {
                "currency": currency,
                "cny_per_unit": values[currency],
                "requested_date": selected_date.isoformat(),
                "actual_date": selected_date.isoformat(),
                "fallback": False,
            }
            for currency in currencies
        }

    monkeypatch.setattr(main_module, "fetch_rates", fake_rates)
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "foreign-amount-correction", "start_date": "2026-08-01", "end_date": "2026-08-10"},
        ).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"amount": 680, "currency": "CNY", "source_sheet": "汇率测试"},
        ).json()["request"]
        payment = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
            json={"amount": 340, "payment_date": "2026-08-08"},
        )
        assert payment.status_code == 200
        current = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
        converted_response = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/currency-conversion/apply",
            json={"target_currency": "USD", "rate_date": "2026-08-08", "expected_updated_at": current["updated_at"]},
        )
        assert converted_response.status_code == 200, converted_response.text
        converted = converted_response.json()["request"]
        assert converted["amount"] == 100
        assert converted["paid_amount"] == 50

        preview_response = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/amount-correction/preview",
            json={"amount": 120, "rate_date": "2026-08-09", "expected_updated_at": converted["updated_at"]},
        )
        assert preview_response.status_code == 200, preview_response.text
        preview = preview_response.json()["preview"]
        assert preview["before"] == {"amount": 100.0, "paid_amount": 50.0, "pending_amount": 50.0}
        assert preview["after"] == {"amount": 120.0, "paid_amount": 50.0, "pending_amount": 70.0}
        assert preview["base_amount_cny"] == 816.0

        applied_response = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/amount-correction/apply",
            json={"amount": 120, "rate_date": "2026-08-09", "expected_updated_at": converted["updated_at"]},
        )
        assert applied_response.status_code == 200, applied_response.text
        corrected = applied_response.json()["request"]
        assert corrected["currency"] == "USD"
        assert corrected["amount"] == 120
        assert corrected["paid_amount"] == 50
        assert corrected["pending_amount"] == 70
        assert corrected["base_amount_cny"] == 816

        unchanged_patch = client.patch(
            f"/api/batches/{batch['id']}/requests/{request['id']}",
            json={"amount": 120, "summary": "其他字段同时保存"},
        )
        assert unchanged_patch.status_code == 200, unchanged_patch.text
        assert unchanged_patch.json()["request"]["summary"] == "其他字段同时保存"

        too_low = client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/amount-correction/preview",
            json={"amount": 40, "rate_date": "2026-08-09"},
        )
        assert too_low.status_code == 400
        assert "累计已支付" in too_low.json()["detail"]
        with connect() as conn:
            audit = conn.execute(
                "SELECT action FROM audit_logs WHERE entity_id = ? AND action = 'request.amount_correct' ORDER BY id DESC LIMIT 1",
                (request["id"],),
            ).fetchone()
            assert audit is not None


def test_historical_currency_restore_can_be_rolled_back():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "historical-currency", "start_date": "2026-08-01", "end_date": "2026-08-10"},
        ).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": "FX-HISTORY-1", "amount": 680, "currency": "CNY", "source_sheet": "汇率测试"},
        ).json()["request"]
        client.post(
            f"/api/batches/{batch['id']}/requests/{request['id']}/payments",
            json={"amount": 340, "payment_date": "2026-08-07"},
        )
        external_source = {
            "system": "dingtalk_expense_database",
            "source_currency": "USD",
            "source_amount": 100,
            "base_currency_amount": 680,
            "application_date": "2026-08-07",
        }
        with connect() as conn:
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": external_source}), request["id"]),
            )

        preview = client.get(f"/api/batches/{batch['id']}/historical-currency-restore/preview")
        assert preview.status_code == 200
        candidate = next(item for item in preview.json()["rows"] if item["request_id"] == request["id"])
        assert candidate["status"] == "recoverable"

        applied = client.post(
            f"/api/batches/{batch['id']}/historical-currency-restore/apply",
            json={"request_ids": [request["id"]]},
        )
        assert applied.status_code == 200, applied.text
        converted = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
        assert converted["currency"] == "USD"
        assert converted["amount"] == 100
        assert converted["paid_amount"] == 50

        rolled_back = client.post(f"/api/batches/{batch['id']}/imports/latest/rollback")
        assert rolled_back.status_code == 200, rolled_back.text
        restored = client.get(f"/api/batches/{batch['id']}/requests").json()["requests"][0]
        assert restored["currency"] == "CNY"
        assert restored["amount"] == 680
        assert restored["paid_amount"] == 340


def test_execution_region_filter_respects_china_workbench_isolation():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "execution-region-filter", "start_date": "2026-08-11", "end_date": "2026-08-14"},
        ).json()["batch"]
        china = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": "REGION-CN", "amount": 100, "currency": "CNY", "source_sheet": "中国公司"},
        ).json()["request"]
        mexico = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": "REGION-MX", "amount": 400, "currency": "CNY", "source_sheet": "墨西哥公司"},
        ).json()["request"]
        explicit_cny = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={"dingding_id": "REGION-MX-CNY", "amount": 400, "currency": "CNY", "source_sheet": "墨西哥公司"},
        ).json()["request"]
        with connect() as conn:
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {
                    "system": "dingtalk_expense_database",
                    "execution_region": "中国China",
                    "source_currency": "CNY",
                    "source_amount": 100,
                    "base_currency_amount": 100,
                }}), china["id"]),
            )
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {
                    "system": "dingtalk_expense_database",
                    "execution_region": "墨西哥 México",
                    "source_currency": "CNY",
                    "source_amount": 1000,
                    "base_currency_amount": 400,
                    "application_date": "2026-08-12",
                }}), mexico["id"]),
            )
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {
                    "system": "dingtalk_expense_database",
                    "execution_region": "墨西哥 México",
                    "source_currency": "CNY",
                    "source_currency_raw": "CNY",
                    "currency_source": "approval_currency",
                    "source_amount": 400,
                    "base_currency_amount": 400,
                }}), explicit_cny["id"]),
            )
            main_module.persist_request_region(conn, china["id"], actor_id=None)
            main_module.persist_request_region(conn, mexico["id"], actor_id=None)
            main_module.persist_request_region(conn, explicit_cny["id"], actor_id=None)

        mexico_rows = client.get(f"/api/batches/{batch['id']}/requests?execution_region=mexico")
        assert mexico_rows.status_code == 200, mexico_rows.text
        assert mexico_rows.json()["requests"] == []
        china_rows = client.get(f"/api/batches/{batch['id']}/requests?execution_region=china")
        assert china_rows.status_code == 200, china_rows.text
        assert [row["dingding_id"] for row in china_rows.json()["requests"]] == ["REGION-CN"]
        assert client.get(f"/api/batches/{batch['id']}/requests?execution_region=other").status_code == 400

def test_historical_currency_restore_uses_explicit_peso_total_in_summary():
    with TestClient(app) as client:
        login(client)
        batch = client.post(
            "/api/batches",
            json={"name": "summary-peso-restore", "start_date": "2026-08-11", "end_date": "2026-08-14"},
        ).json()["batch"]
        request = client.post(
            f"/api/batches/{batch['id']}/requests",
            json={
                "dingding_id": "SUMMARY-MXN",
                "summary": "本地购买货架，合计76,800比索（不含税）",
                "amount": 29952,
                "currency": "CNY",
                "source_sheet": "UV IMPRESION MX彩印",
            },
        ).json()["request"]
        with connect() as conn:
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {
                    "system": "legacy_excel_import",
                    "execution_region": "墨西哥 México",
                    "source_currency": "CNY",
                    "source_currency_raw": "CNY",
                    "currency_source": "approval_currency",
                    "source_amount": 29952,
                    "base_currency_amount": 29952,
                    "application_date": "2026-08-11",
                }}), request["id"]),
            )

        preview = client.get(f"/api/batches/{batch['id']}/historical-currency-restore/preview")
        assert preview.status_code == 200, preview.text
        candidate = next(row for row in preview.json()["rows"] if row["request_id"] == request["id"])
        assert candidate["status"] == "recoverable"
        assert candidate["source_currency"] == "MXN"
        assert candidate["source_amount"] == 76800
        assert candidate["base_amount_cny"] == 29952
        assert candidate["currency_source"] == "summary_text"


def test_mexico_tracking_api_enforces_mexico_access_scope_and_participation():
    with (
        TestClient(app) as admin_client,
        TestClient(app) as participant_client,
        TestClient(app) as all_client,
        TestClient(app) as none_client,
    ):
        login(admin_client)
        for username, scope, identity in (
            ("mexico-participant", "participant", "Ana"),
            ("mexico-all", "all", None),
            ("mexico-none", "none", None),
        ):
            created_user = admin_client.post(
                "/api/admin/users",
                json={
                    "username": username,
                    "password": "Yuewei123",
                    "display_name": username,
                    "role": "business",
                    "active": True,
                    "sheet_permissions": [],
                    "mexico_access_scope": scope,
                    "mexico_identity_name": identity,
                },
            )
            assert created_user.status_code == 200, created_user.text
        timestamp = now_iso()
        with connect() as conn:
            tracking_ids: dict[str, int] = {}
            for approval_no, applicant_name, company_name in (
                ("MX-API-APPLICANT", "Ana", "Applicant Company"),
                ("MX-API-TASK", "Bruno", "Task Company"),
                ("MX-API-EVENT", "Carla", "Event Company"),
                ("MX-API-DENIED", "Diego", "Hidden Company"),
            ):
                tracking_ids[approval_no] = int(
                    conn.execute(
                        """
                        INSERT INTO mexico_approval_tracking (
                            approval_no, source_type, resolved_region,
                            region_resolution_source, region_review_status,
                            source_sheet, company_name, applicant_name, summary,
                            workflow_status, current_node_name, current_approver_name,
                            current_node_entered_at, created_at, updated_at
                        ) VALUES (?, 'purchase', 'mexico', 'execution_region', 'resolved',
                                  'NO CHINA SHEET', ?, ?, ?, 'RUNNING', '审批',
                                  'Otro', ?, ?, ?)
                        """,
                        (
                            approval_no,
                            company_name,
                            applicant_name,
                            approval_no,
                            timestamp,
                            timestamp,
                            timestamp,
                        ),
                    ).lastrowid
                )
            conn.execute(
                """
                INSERT INTO mexico_approval_current_tasks (
                    approval_no, task_key, node_name, approver_name,
                    entered_at, synced_at, created_at, updated_at
                ) VALUES ('MX-API-TASK', 'task-ana', '财务审批', 'Ana', ?, ?, ?, ?)
                """,
                (timestamp, timestamp, timestamp, timestamp),
            )
            conn.execute(
                """
                INSERT INTO mexico_approval_events (
                    approval_no, event_key, sequence_index, event_type,
                    operator_name, event_time, is_current, created_at, updated_at
                ) VALUES ('MX-API-EVENT', 'event-ana', 1, 'CC', 'Ana', ?, 0, ?, ?)
                """,
                (timestamp, timestamp, timestamp),
            )
            conn.execute(
                """
                INSERT INTO mexico_approval_tracking (
                    approval_no, source_type, resolved_region,
                    region_resolution_source, region_review_status,
                    source_sheet, summary, workflow_status, created_at, updated_at
                ) VALUES (?, 'operation', 'review', 'conflict', 'pending',
                          ?, '地区冲突', 'RUNNING', ?, ?)
                """,
                ("MX-API-REVIEW", "YW MOLDES MX模具", timestamp, timestamp),
            )

        login(participant_client, "mexico-participant", "Yuewei123")
        listed = participant_client.get("/api/mexico-tracking?view=pending")
        assert listed.status_code == 200, listed.text
        assert {item["approval_no"] for item in listed.json()["items"]} == {
            "MX-API-APPLICANT",
            "MX-API-TASK",
            "MX-API-EVENT",
        }
        summary = participant_client.get("/api/mexico-tracking/summary")
        assert summary.status_code == 200, summary.text
        assert summary.json()["summary"]["pending"] == 3
        assert summary.json()["summary"]["review"] == 0
        options = participant_client.get("/api/mexico-tracking/filter-options")
        assert options.status_code == 200, options.text
        assert "Hidden Company" not in options.json()["options"]["companies"]
        assert participant_client.get(
            f"/api/mexico-tracking/{tracking_ids['MX-API-EVENT']}"
        ).status_code == 200
        assert participant_client.get(
            f"/api/mexico-tracking/{tracking_ids['MX-API-DENIED']}"
        ).status_code == 404
        assert participant_client.get("/api/mexico-tracking?view=review").status_code == 403

        login(all_client, "mexico-all", "Yuewei123")
        all_items = all_client.get("/api/mexico-tracking?view=pending")
        assert all_items.status_code == 200, all_items.text
        assert all_items.json()["total"] == 4
        all_summary = all_client.get("/api/mexico-tracking/summary")
        assert all_summary.status_code == 200, all_summary.text
        assert all_summary.json()["summary"]["review"] == 0

        login(none_client, "mexico-none", "Yuewei123")
        for method, path in (
            ("get", "/api/mexico-tracking"),
            ("get", "/api/mexico-tracking/summary"),
            ("get", "/api/mexico-tracking/filter-options"),
            ("get", "/api/mexico-tracking/settings"),
            ("get", f"/api/mexico-tracking/{tracking_ids['MX-API-APPLICANT']}"),
            ("post", "/api/mexico-tracking/sync"),
        ):
            response = getattr(none_client, method)(path)
            assert response.status_code == 403, (path, response.text)

        review = admin_client.get("/api/mexico-tracking?view=review")
        assert review.status_code == 200, review.text
        assert any(item["approval_no"] == "MX-API-REVIEW" for item in review.json()["items"])
        admin_summary = admin_client.get("/api/mexico-tracking/summary")
        assert admin_summary.status_code == 200, admin_summary.text
        assert admin_summary.json()["summary"]["review"] >= 1
        assert participant_client.put(
            "/api/mexico-tracking/settings",
            json={
                "yellow_days": 2,
                "red_days": 5,
                "cache_stale_seconds": 300,
                "china_region_isolation_enabled": False,
            },
        ).status_code == 403


def test_mexico_approver_stats_api_uses_participant_visibility():
    token = uuid.uuid4().hex
    username = f"mexico-stats-{token}"
    identity = f"Stats Ana {token}"
    approver = f"Stats Bruno {token}"
    approval_no = f"MX-STATS-API-{token}"
    with TestClient(app) as admin_client, TestClient(app) as participant_client:
        login(admin_client)
        created = admin_client.post(
            "/api/admin/users",
            json={
                "username": username,
                "password": "Yuewei123",
                "display_name": username,
                "role": "business",
                "active": True,
                "sheet_permissions": [],
                "mexico_access_scope": "participant",
                "mexico_identity_name": identity,
            },
        )
        assert created.status_code == 200, created.text
        timestamp = now_iso()
        with connect() as conn:
            conn.execute(
                """
                INSERT INTO mexico_approval_tracking (
                    approval_no, source_type, resolved_region,
                    region_resolution_source, region_review_status,
                    applicant_name, company_name, workflow_status,
                    current_node_entered_at, created_at, updated_at
                ) VALUES (?, 'purchase', 'mexico', 'execution_region', 'resolved',
                          ?, 'Stats company', 'RUNNING', ?, ?, ?)
                """,
                (approval_no, identity, timestamp, timestamp, timestamp),
            )
            conn.execute(
                """
                INSERT INTO mexico_approval_current_tasks (
                    approval_no, task_key, node_name, approver_name,
                    entered_at, synced_at, created_at, updated_at
                ) VALUES (?, 'stats-task', 'Finance', ?, ?, ?, ?, ?)
                """,
                (approval_no, approver, timestamp, timestamp, timestamp, timestamp),
            )

        login(participant_client, username, "Yuewei123")
        response = participant_client.get("/api/mexico-tracking/approver-stats")

        assert response.status_code == 200, response.text
        assert response.json()["items"] == [
            {
                "approver_name": approver,
                "pending": 1,
                "overdue": 0,
                "severe": 0,
            }
        ]


def test_mexico_tracking_row_attachment_sync_is_idempotent_and_authorized(
    monkeypatch: pytest.MonkeyPatch,
):
    username = f"mexico-row-attachments-{uuid.uuid4().hex}"
    password = "Yuewei123"
    with TestClient(app) as admin_client, TestClient(app) as business_client:
        login(admin_client)
        created_user = admin_client.post(
            "/api/admin/users",
            json={
                "username": username,
                "password": password,
                "display_name": "墨西哥附件用户",
                "role": "business",
                "active": True,
                "sheet_permissions": [],
                "mexico_access_scope": "participant",
                "mexico_identity_name": "Attachment Ana",
            },
        )
        assert created_user.status_code == 200, created_user.text
        timestamp = now_iso()
        allowed_no = f"MX-ROW-ALLOWED-{uuid.uuid4().hex}"
        denied_no = f"MX-ROW-DENIED-{uuid.uuid4().hex}"
        with connect() as conn:
            allowed_id = int(
                conn.execute(
                    """
                        INSERT INTO mexico_approval_tracking (
                            approval_no, source_type, process_instance_id,
                            resolved_region, region_resolution_source,
                            region_review_status, source_sheet, applicant_name, workflow_status,
                            created_at, updated_at
                        ) VALUES (?, 'purchase', ?, 'mexico', 'execution_region',
                                  'resolved', 'NO CHINA SHEET', 'Attachment Ana', 'RUNNING', ?, ?)
                    """,
                    (allowed_no, f"process-{allowed_no}", timestamp, timestamp),
                ).lastrowid
            )
            denied_id = int(
                conn.execute(
                    """
                        INSERT INTO mexico_approval_tracking (
                            approval_no, source_type, process_instance_id,
                            resolved_region, region_resolution_source,
                            region_review_status, source_sheet, applicant_name, workflow_status,
                            created_at, updated_at
                        ) VALUES (?, 'purchase', ?, 'mexico', 'execution_region',
                                  'resolved', 'NO CHINA SHEET', 'Somebody Else', 'RUNNING', ?, ?)
                    """,
                    (denied_no, f"process-{denied_no}", timestamp, timestamp),
                ).lastrowid
            )

        def fake_workflows(approval_nos):
            approval_no = next(iter(approval_nos))
            return [
                {
                    "approval_no": approval_no,
                    "process_instance_id": f"process-{approval_no}",
                    "events": [
                        {
                            "event_key": "event-1",
                            "attachments": [
                                {"fileId": "row-file-1", "fileName": "invoice.pdf"}
                            ],
                            "images": [],
                        }
                    ],
                }
            ]

        submitted: list[str] = []
        monkeypatch.setattr(main_module, "fetch_dingtalk_workflows", fake_workflows)
        monkeypatch.setattr(
            main_module,
            "fetch_external_expense_attachments",
            lambda approval_nos: [],
        )
        monkeypatch.setattr(
            main_module,
            "_submit_mexico_attachment_task",
            lambda run_id: submitted.append(run_id),
            raising=False,
        )

        login(business_client, username, password)
        first = business_client.post(
            f"/api/mexico-tracking/{allowed_id}/attachments/sync"
        )
        assert first.status_code == 202, first.text
        assert first.json()["attachment_status"]["queued"] == 1
        assert first.json()["run"]["kind"] == "mexico-attachments"

        second = business_client.post(
            f"/api/mexico-tracking/{allowed_id}/attachments/sync"
        )
        assert second.status_code == 202, second.text
        assert second.json()["run"]["id"] == first.json()["run"]["id"]
        assert business_client.post(
            f"/api/mexico-tracking/{denied_id}/attachments/sync"
        ).status_code == 404

        with connect() as conn:
            assert conn.execute(
                "SELECT COUNT(*) FROM mexico_approval_attachments WHERE approval_no = ?",
                (allowed_no,),
            ).fetchone()[0] == 1
            assert conn.execute(
                "SELECT COUNT(*) FROM mexico_sync_runs "
                "WHERE kind = 'mexico-attachments' AND status IN ('queued', 'running')"
            ).fetchone()[0] == 1
        assert submitted == [first.json()["run"]["id"], first.json()["run"]["id"]]
        with connect() as conn:
            conn.execute(
                "UPDATE mexico_approval_attachments SET status = 'failed' "
                "WHERE approval_no = ?",
                (allowed_no,),
            )
            conn.execute(
                "UPDATE mexico_sync_runs SET status = 'completed', phase = 'complete', "
                "completed_at = ?, lease_until = NULL WHERE id = ?",
                (now_iso(), first.json()["run"]["id"]),
            )


def test_china_region_isolation_filters_workbench_totals_sheets_and_export():
    with TestClient(app) as client:
        login(client)
        settings_payload = {
            "yellow_days": 2,
            "red_days": 5,
            "cache_stale_seconds": 300,
            "china_region_isolation_enabled": False,
        }
        assert client.put("/api/mexico-tracking/settings", json=settings_payload).status_code == 200
        try:
            batch = client.post(
                "/api/batches",
                json={
                    "name": f"中国口径隔离-{uuid.uuid4().hex}",
                    "start_date": "2026-08-24",
                    "end_date": "2026-08-31",
                },
            ).json()["batch"]
            rows = [
                ("凌翔产品&开发", "CN-ISOLATION", 100),
                ("YW MOLDES MX模具", "MX-ISOLATION", 200),
                (f"地区待核对-{uuid.uuid4().hex}", "REVIEW-ISOLATION", 300),
            ]
            for sheet_name, approval_no, amount in rows:
                created = client.post(
                    f"/api/batches/{batch['id']}/requests",
                    json={
                        "source_sheet": sheet_name,
                        "dingding_id": approval_no,
                        "summary": approval_no,
                        "amount": amount,
                        "currency": "CNY",
                    },
                )
                assert created.status_code == 200, created.text

            listed = client.get(f"/api/batches/{batch['id']}/requests")
            assert listed.status_code == 200, listed.text
            assert [row["dingding_id"] for row in listed.json()["requests"]] == ["CN-ISOLATION"]
            assert listed.json()["totals"]["count"] == 1
            assert listed.json()["totals"]["amount"] == 100

            detail = client.get(f"/api/batches/{batch['id']}")
            assert detail.status_code == 200, detail.text
            assert detail.json()["batch"]["request_count"] == 1
            assert detail.json()["batch"]["total_amount"] == 100
            assert "凌翔产品&开发" in detail.json()["batch"]["sheet_order"]
            assert "YW MOLDES MX模具" not in detail.json()["batch"]["sheet_order"]
            assert not any(name.startswith("地区待核对-") for name in detail.json()["batch"]["sheet_order"])

            batches = client.get("/api/batches").json()["batches"]
            public_batch = next(item for item in batches if item["id"] == batch["id"])
            assert public_batch["request_count"] == 1
            assert public_batch["total_amount"] == 100

            exported = client.get(f"/api/batches/{batch['id']}/export.xlsx")
            assert exported.status_code == 200, exported.text
            workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
            workbook_text = "\n".join(
                str(cell.value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            assert "CN-ISOLATION" in workbook_text
            assert "MX-ISOLATION" not in workbook_text
            assert "REVIEW-ISOLATION" not in workbook_text
        finally:
            settings_payload["china_region_isolation_enabled"] = False
            client.put("/api/mexico-tracking/settings", json=settings_payload)
