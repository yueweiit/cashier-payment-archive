from __future__ import annotations

import asyncio
import gc
import io
import json
import sys
import tempfile
import uuid
from calendar import monthrange
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import unquote

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app import daily_payables_export
import backend.app.main as main_module
from backend.app.db import connect, get_daily_payables_history_start_date, now_iso
from backend.app.main import app
from backend.app.daily_payables import daily_snapshot, daily_trend
from backend.app.payable_history import payment_effective_at, record_request_state
from backend.app.security import hash_password


def login(client: TestClient) -> None:
    response = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    assert response.status_code == 200


def login_business_for_sheets(client: TestClient, *sheet_names: str) -> str:
    username = f"daily-business-{uuid.uuid4().hex}"
    with connect() as conn:
        cursor = conn.execute(
            """
            INSERT INTO users (username, password_hash, role, display_name, active, created_at)
            VALUES (?, ?, 'business', ?, 1, ?)
            """,
            (username, hash_password("daily123"), username, now_iso()),
        )
        user_id = int(cursor.lastrowid)
        for sheet_name in sheet_names:
            conn.execute(
                """
                INSERT INTO user_sheet_permissions (user_id, sheet_name, created_by, created_at)
                VALUES (?, ?, 1, ?)
                """,
                (user_id, sheet_name, now_iso()),
            )
    client.post("/api/auth/logout")
    response = client.post(
        "/api/auth/login",
        json={"username": username, "password": "daily123"},
    )
    assert response.status_code == 200
    return username


def create_request(
    client: TestClient,
    *,
    amount: float = 20000,
    source_sheet: str = "每日应付测试公司",
    needed_payment_date: str | None = "2026-08-21",
    currency: str = "CNY",
    execution_region: str | None = "China",
    dingding_id: str | None = None,
    summary: str = "历史状态测试",
) -> dict:
    batch = client.post(
        "/api/batches",
        json={"name": f"每日应付-{uuid.uuid4().hex}", "start_date": "2026-08-21", "end_date": "2026-08-28"},
    ).json()["batch"]
    payload = {
        "source_sheet": source_sheet,
        "summary": summary,
        "applicant": "测试申请人",
        "amount": amount,
        "currency": currency,
        "needed_payment_date": needed_payment_date,
    }
    if dingding_id is not None:
        payload["dingding_id"] = dingding_id
    if execution_region is not None:
        payload["raw_extra"] = {
            "external_source": {"execution_region": execution_region}
        }
    return client.post(
        f"/api/batches/{batch['id']}/requests",
        json=payload,
    ).json()["request"]


def shift_calendar_months(value: date, months: int) -> date:
    month_index = value.year * 12 + value.month - 1 + months
    year, month_zero = divmod(month_index, 12)
    month = month_zero + 1
    return value.replace(year=year, month=month, day=min(value.day, monthrange(year, month)[1]))


def test_record_request_state_is_idempotent_and_keeps_full_state():
    with TestClient(app) as client:
        login(client)
        request = create_request(client)
        event_key = f"test:{uuid.uuid4().hex}"
        with connect() as conn:
            record_request_state(
                conn,
                request["id"],
                event_type="request.updated",
                event_key=event_key,
                effective_at="2026-08-21T12:00:00.000000",
                actor_id=1,
            )
            record_request_state(
                conn,
                request["id"],
                event_type="request.updated",
                event_key=event_key,
                effective_at="2026-08-21T12:00:00.000000",
                actor_id=1,
            )
            rows = conn.execute(
                "SELECT * FROM payable_history_versions WHERE event_key = ?",
                (event_key,),
            ).fetchall()
            current = conn.execute(
                "SELECT logical_request_id FROM payment_requests WHERE id = ?",
                (request["id"],),
            ).fetchone()
            assert len(rows) == 1
            row = rows[0]
            assert row["logical_request_id"] == current["logical_request_id"]
            assert row["source_batch_id"] == request["batch_id"]
            assert row["dingding_id"] == request.get("dingding_id")
            assert row["amount"] == 20000
            assert row["paid_amount"] == 0
            assert row["pending_amount"] == 20000
            assert row["needed_payment_date"] == "2026-08-21"
            assert row["source_sheet"] == "每日应付测试公司"
            assert row["included"] == 1


def test_record_request_state_tracks_terminated_refused_and_reopened():
    with TestClient(app) as client:
        login(client)
        request = create_request(client)
        with connect() as conn:
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {"approval_status": "TERMINATED"}}), request["id"]),
            )
            record_request_state(
                conn,
                request["id"],
                event_type="dingtalk.terminated",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at="2026-08-21T13:00:00.000000",
                actor_id=1,
            )
            terminated = conn.execute(
                "SELECT * FROM payable_history_versions WHERE source_request_id = ? ORDER BY id DESC LIMIT 1",
                (request["id"],),
            ).fetchone()
            assert terminated["approval_status"] == "TERMINATED"
            assert terminated["included"] == 0

            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {"approval_status": "RUNNING", "approval_result": "agree"}}), request["id"]),
            )
            record_request_state(
                conn,
                request["id"],
                event_type="dingtalk.reopened",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at="2026-08-21T14:00:00.000000",
                actor_id=1,
            )
            reopened = conn.execute(
                "SELECT * FROM payable_history_versions WHERE source_request_id = ? ORDER BY id DESC LIMIT 1",
                (request["id"],),
            ).fetchone()
            assert reopened["approval_status"] == "RUNNING"
            assert reopened["approval_result"] == "agree"
            assert reopened["included"] == 1


def test_history_write_rolls_back_with_business_transaction():
    with TestClient(app) as client:
        login(client)
        request = create_request(client)
        event_key = f"rollback:{uuid.uuid4().hex}"
        with connect() as conn:
            conn.execute("BEGIN IMMEDIATE")
            record_request_state(
                conn,
                request["id"],
                event_type="request.updated",
                event_key=event_key,
                effective_at=now_iso(),
                actor_id=1,
            )
            conn.rollback()
            assert not conn.execute(
                "SELECT 1 FROM payable_history_versions WHERE event_key = ?",
                (event_key,),
            ).fetchone()


def test_payment_effective_at_uses_end_of_backdated_payment_day():
    assert payment_effective_at(
        "2026-08-20",
        recorded_at="2026-08-21T09:00:00.000000",
    ) == "2026-08-20T23:59:59.999999"
    assert payment_effective_at(
        "2026-08-21",
        recorded_at="2026-08-21T09:00:00.000000",
    ) == "2026-08-21T09:00:00.000000"


def test_request_and_payment_endpoints_append_history_versions():
    with TestClient(app) as client:
        login(client)
        request = create_request(client)
        request_id = request["id"]
        with connect() as conn:
            created = conn.execute(
                "SELECT * FROM payable_history_versions WHERE source_request_id = ? ORDER BY id DESC LIMIT 1",
                (request_id,),
            ).fetchone()
            assert created is not None
            assert created["event_type"] == "request.create"
            logical_request_id = created["logical_request_id"]

        updated_response = client.patch(
            f"/api/batches/{request['batch_id']}/requests/{request_id}",
            json={"summary": "更新后的摘要", "expected_version": request["version"]},
        )
        assert updated_response.status_code == 200, updated_response.text
        updated = updated_response.json()["request"]

        payment_response = client.post(
            f"/api/batches/{request['batch_id']}/requests/{request_id}/payments",
            json={
                "amount": 10000,
                "payment_date": "2026-08-21",
                "payer": "测试付款人",
                "expected_request_version": updated["version"],
            },
        )
        assert payment_response.status_code == 200, payment_response.text
        after_payment = payment_response.json()["request"]

        with connect() as conn:
            rows = conn.execute(
                "SELECT * FROM payable_history_versions WHERE logical_request_id = ? ORDER BY id",
                (logical_request_id,),
            ).fetchall()
            assert [row["event_type"] for row in rows[-3:]] == [
                "request.create",
                "request.update",
                "payment.create",
            ]
            assert rows[-1]["paid_amount"] == 10000
            assert rows[-1]["pending_amount"] == 10000

        delete_response = client.delete(
            f"/api/batches/{request['batch_id']}/requests/{request_id}",
            params={"expected_version": after_payment["version"]},
        )
        assert delete_response.status_code == 200, delete_response.text
        with connect() as conn:
            tombstone = conn.execute(
                "SELECT * FROM payable_history_versions WHERE logical_request_id = ? ORDER BY id DESC LIMIT 1",
                (logical_request_id,),
            ).fetchone()
            assert tombstone["event_type"] == "request.delete"
            assert tombstone["deleted"] == 1
            assert tombstone["included"] == 0


def test_rollover_inherits_logical_request_id_without_duplicate_identity():
    with TestClient(app) as client:
        login(client)
        request = create_request(client)
        source_batch = next(
            batch
            for batch in client.get("/api/batches").json()["batches"]
            if batch["id"] == request["batch_id"]
        )
        response = client.post(
            f"/api/batches/{source_batch['id']}/rollover",
            json={
                "name": f"结转-{uuid.uuid4().hex}",
                "start_date": "2026-08-29",
                "end_date": "2026-09-04",
                "copy_mode": "all",
                "expected_batch_version": source_batch["version"],
            },
        )
        assert response.status_code == 200, response.text
        target_batch = response.json()["batch"]
        target_requests = client.get(f"/api/batches/{target_batch['id']}/requests").json()["requests"]
        copied = next(item for item in target_requests if item["copied_from_request_id"] == request["id"])
        assert copied["logical_request_id"] == request["id"]


def test_daily_payables_deduplicates_reimports_by_dingtalk_business_id():
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        due_date = (selected - timedelta(days=1)).isoformat()
        dingding_id = f"daily-dedup-{uuid.uuid4().hex}"
        source_sheet = f"每日应付去重-{uuid.uuid4().hex}"
        older = create_request(
            client,
            amount=290,
            source_sheet=source_sheet,
            needed_payment_date=due_date,
            dingding_id=dingding_id,
            summary="跨周重复导入",
        )
        newer = create_request(
            client,
            amount=290,
            source_sheet=source_sheet,
            needed_payment_date=due_date,
            dingding_id=dingding_id,
            summary="跨周重复导入",
        )

        with connect() as conn:
            snapshot = daily_snapshot(
                conn,
                selected,
                allowed_sheets={source_sheet},
                include_details=True,
                china_only=True,
            )
            trend = daily_trend(
                conn,
                selected,
                selected,
                allowed_sheets={source_sheet},
                china_only=True,
            )

        matches = [
            item for item in snapshot["items"]
            if item["dingding_id"] == dingding_id
        ]
        assert len(matches) == 1
        assert matches[0]["logical_request_id"] == newer["logical_request_id"]
        assert matches[0]["source_request_id"] == newer["id"]
        assert snapshot["totals_cny"]["end_pending"] == 290
        assert snapshot["counts"]["end_pending"] == 1
        assert trend["points"][0]["totals_cny"]["end_pending"] == 290
        assert trend["points"][0]["counts"]["end_pending"] == 1


def test_daily_payables_export_generates_summary_and_detail_workbook():
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        selected_iso = selected.isoformat()
        source_sheet = f"每日应付导出-{uuid.uuid4().hex}"
        request = create_request(
            client,
            amount=290,
            source_sheet=source_sheet,
            needed_payment_date=selected_iso,
            dingding_id="=1+1",
            summary='=HYPERLINK("https://example.invalid", "中转站290")\x01',
        )
        login_business_for_sheets(client, source_sheet)
        temp_directory = Path(tempfile.gettempdir())
        temp_before = set(temp_directory.glob("daily-payables-*.xlsx"))

        response = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": selected_iso, "end": selected_iso},
        )

        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert int(response.headers["content-length"]) == len(response.content)
        assert set(temp_directory.glob("daily-payables-*.xlsx")) == temp_before
        assert f"每日应付_{selected.strftime('%Y%m%d')}-{selected.strftime('%Y%m%d')}.xlsx" in unquote(
            response.headers["content-disposition"]
        )
        workbook = load_workbook(io.BytesIO(response.content), data_only=False)
        assert workbook.sheetnames == ["每日汇总", "逐日明细"]

        summary = workbook["每日汇总"]
        summary_headers = [cell.value for cell in summary[1]]
        assert summary_headers[:8] == [
            "统计日期",
            "当日到期数量",
            "日终未付数量",
            "逾期数量",
            "当天新增到期（折合人民币）",
            "当日支付（折合人民币）",
            "日终待付（折合人民币）",
            "逾期待付（折合人民币）",
        ]
        assert summary.cell(2, 1).value.strftime("%Y-%m-%d") == selected_iso
        assert [summary.cell(2, column).value for column in range(2, 9)] == [1, 1, 0, 290, 0, 290, 0]
        cny_pending_column = summary_headers.index("CNY 日终待付") + 1
        assert summary.cell(2, cny_pending_column).value == 290

        detail = workbook["逐日明细"]
        detail_headers = [cell.value for cell in detail[1]]
        assert detail_headers == [
            "统计日期",
            "状态",
            "请款标识",
            "钉钉申请单号",
            "应付款公司（来源 Sheet）",
            "申请人",
            "摘要",
            "需求付款日期",
            "应付金额",
            "累计已付",
            "当日支付",
            "日终待付",
            "币种",
            "折合人民币应付金额",
            "折合人民币累计已付",
            "折合人民币当日支付",
            "折合人民币日终待付",
            "审批状态",
            "审批结果",
        ]
        detail_row = [cell.value for cell in detail[2]]
        assert detail_row[0].strftime("%Y-%m-%d") == selected_iso
        assert detail_row[1] == "当日到期"
        assert detail_row[2] == request["logical_request_id"]
        assert detail_row[4] == source_sheet
        assert detail_row[3] == "'=1+1"
        assert detail.cell(2, 4).data_type == "s"
        assert detail_row[6] == "'=HYPERLINK(\"https://example.invalid\", \"中转站290\")"
        assert detail.cell(2, 7).data_type == "s"
        assert detail_row[8:13] == [290, 0, 0, 290, "CNY"]


def test_daily_payables_export_validates_historical_six_month_range():
    with TestClient(app) as client:
        login(client)
        today = date.today()
        start = shift_calendar_months(today, -6)
        valid_end = today - timedelta(days=1)
        with connect() as conn:
            conn.execute(
                """
                INSERT INTO app_settings (key, value, updated_at)
                VALUES ('daily_payables_history_start_date', ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    updated_at = excluded.updated_at
                """,
                (start.isoformat(), now_iso()),
            )

        valid = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": start.isoformat(), "end": valid_end.isoformat()},
        )
        assert valid.status_code == 200, valid.text

        too_large = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": start.isoformat(), "end": today.isoformat()},
        )
        assert too_large.status_code == 422
        assert too_large.json()["detail"]["code"] == "EXPORT_RANGE_TOO_LARGE"

        future = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": today.isoformat(), "end": (today + timedelta(days=1)).isoformat()},
        )
        assert future.status_code == 422
        assert future.json()["detail"]["code"] == "FUTURE_EXPORT_DATE"

        inverted = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": today.isoformat(), "end": valid_end.isoformat()},
        )
        assert inverted.status_code == 422
        assert inverted.json()["detail"]["code"] == "INVALID_EXPORT_RANGE"

        before_history = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": (start - timedelta(days=1)).isoformat(), "end": start.isoformat()},
        )
        assert before_history.status_code == 422
        assert before_history.json()["detail"]["code"] == "HISTORY_NOT_AVAILABLE"


def test_daily_payables_export_replays_deduplicated_payment_history_and_formats_excel():
    with TestClient(app) as client:
        login(client)
        second_day = date.today()
        first_day = second_day - timedelta(days=1)
        source_sheet = f"每日应付区间导出-{uuid.uuid4().hex}"
        dingding_id = f"daily-range-{uuid.uuid4().hex}"
        with connect() as conn:
            conn.execute(
                """
                UPDATE app_settings SET value = ?, updated_at = ?
                WHERE key = 'daily_payables_history_start_date'
                """,
                (first_day.isoformat(), now_iso()),
            )
        older = create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=first_day.isoformat(),
            dingding_id=dingding_id,
            summary="区间去重测试",
        )
        newer = create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=first_day.isoformat(),
            dingding_id=dingding_id,
            summary="区间去重测试",
        )

        with connect() as conn:
            conn.execute(
                "UPDATE payable_history_versions SET effective_at = ? WHERE logical_request_id = ?",
                (f"{first_day.isoformat()}T08:00:00.000000", older["logical_request_id"]),
            )
            conn.execute(
                "UPDATE payable_history_versions SET effective_at = ? WHERE logical_request_id = ?",
                (f"{first_day.isoformat()}T09:00:00.000000", newer["logical_request_id"]),
            )
            conn.execute(
                "UPDATE payment_requests SET paid_amount = 40, pending_amount = 60 WHERE id = ?",
                (newer["id"],),
            )
            record_request_state(
                conn,
                newer["id"],
                event_type="payment.create",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at=f"{second_day.isoformat()}T10:00:00.000000",
                actor_id=1,
            )
        login_business_for_sheets(client, source_sheet)

        response = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": first_day.isoformat(), "end": second_day.isoformat()},
        )
        assert response.status_code == 200, response.text
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        summary = workbook["每日汇总"]
        detail = workbook["逐日明细"]

        assert summary.max_row == 3
        assert summary.cell(2, 7).value == 100
        assert summary.cell(3, 6).value == 40
        assert summary.cell(3, 7).value == 60
        assert summary.cell(3, 8).value == 60
        assert summary.cell(2, 5).number_format == "#,##0.00"
        assert summary.freeze_panes == "A2"
        assert summary.auto_filter.ref == f"A1:T{summary.max_row}"

        assert detail.max_row == 3
        first_detail = [cell.value for cell in detail[2]]
        second_detail = [cell.value for cell in detail[3]]
        assert first_detail[1] == "当日到期"
        assert first_detail[2] == newer["logical_request_id"]
        assert second_detail[1] == "逾期待付"
        assert second_detail[2] == newer["logical_request_id"]
        assert second_detail[9:12] == [40, 40, 60]
        assert second_detail[14:17] == [40, 40, 60]
        assert detail.cell(2, 9).number_format == "#,##0.00"
        assert detail.freeze_panes == "A2"
        assert detail.auto_filter.ref == f"A1:S{detail.max_row}"


def test_daily_payables_export_includes_prior_due_items_paid_in_full_that_day():
    with TestClient(app) as client:
        login(client)
        paid_day = date.today()
        due_day = paid_day - timedelta(days=1)
        source_sheet = f"每日应付当日付清-{uuid.uuid4().hex}"
        with connect() as conn:
            conn.execute(
                "UPDATE app_settings SET value = ?, updated_at = ? WHERE key = 'daily_payables_history_start_date'",
                (due_day.isoformat(), now_iso()),
            )
        request = create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=due_day.isoformat(),
            summary="前日到期当日付清",
        )
        with connect() as conn:
            conn.execute(
                "UPDATE payable_history_versions SET effective_at = ? WHERE logical_request_id = ?",
                (f"{due_day.isoformat()}T09:00:00.000000", request["logical_request_id"]),
            )
            conn.execute(
                "UPDATE payment_requests SET paid_amount = 100, pending_amount = 0 WHERE id = ?",
                (request["id"],),
            )
            record_request_state(
                conn,
                request["id"],
                event_type="payment.create",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at=f"{paid_day.isoformat()}T10:00:00.000000",
                actor_id=1,
            )
        login_business_for_sheets(client, source_sheet)

        response = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": due_day.isoformat(), "end": paid_day.isoformat()},
        )
        assert response.status_code == 200, response.text
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        summary = workbook["每日汇总"]
        detail = workbook["逐日明细"]

        assert summary.cell(3, 6).value == 100
        paid_rows = [
            row
            for row in detail.iter_rows(min_row=2, values_only=True)
            if row[0].strftime("%Y-%m-%d") == paid_day.isoformat()
        ]
        assert len(paid_rows) == 1
        assert paid_rows[0][1] == "当日付清"
        assert paid_rows[0][9:12] == (100, 100, 0)


def test_daily_payables_export_respects_sheet_permissions_and_china_region():
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        selected_iso = selected.isoformat()
        visible_sheet = f"每日导出授权-{uuid.uuid4().hex}"
        hidden_sheet = f"每日导出无权-{uuid.uuid4().hex}"
        mexico_sheet = f"每日导出墨西哥-{uuid.uuid4().hex}"
        visible = create_request(
            client,
            amount=100,
            source_sheet=visible_sheet,
            needed_payment_date=selected_iso,
            summary="可导出记录",
        )
        create_request(
            client,
            amount=200,
            source_sheet=hidden_sheet,
            needed_payment_date=selected_iso,
            summary="无权记录",
        )
        create_request(
            client,
            amount=300,
            source_sheet=mexico_sheet,
            needed_payment_date=selected_iso,
            execution_region="Mexico",
            summary="墨西哥记录",
        )
        username = f"daily-export-{uuid.uuid4().hex}"
        with connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO users (username, password_hash, role, display_name, active, created_at)
                VALUES (?, ?, 'business', ?, 1, ?)
                """,
                (username, hash_password("daily123"), username, now_iso()),
            )
            user_id = int(cursor.lastrowid)
            for sheet_name in (visible_sheet, mexico_sheet):
                conn.execute(
                    """
                    INSERT INTO user_sheet_permissions (user_id, sheet_name, created_by, created_at)
                    VALUES (?, ?, 1, ?)
                    """,
                    (user_id, sheet_name, now_iso()),
                )

        client.post("/api/auth/logout")
        login_response = client.post(
            "/api/auth/login",
            json={"username": username, "password": "daily123"},
        )
        assert login_response.status_code == 200
        response = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": selected_iso, "end": selected_iso},
        )
        assert response.status_code == 200, response.text
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        summary = workbook["每日汇总"]
        detail = workbook["逐日明细"]
        assert summary.cell(2, 7).value == 100
        assert detail.max_row == 2
        assert detail.cell(2, 3).value == visible["logical_request_id"]
        assert detail.cell(2, 7).value == "可导出记录"


def test_daily_payables_export_rejects_excel_detail_overflow(monkeypatch):
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        selected_iso = selected.isoformat()
        source_sheet = f"每日导出上限-{uuid.uuid4().hex}"
        create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=selected_iso,
        )
        create_request(
            client,
            amount=200,
            source_sheet=source_sheet,
            needed_payment_date=selected_iso,
        )
        login_business_for_sheets(client, source_sheet)
        monkeypatch.setattr(daily_payables_export, "EXCEL_MAX_ROWS", 2)
        unraisable: list[BaseException] = []
        monkeypatch.setattr(sys, "unraisablehook", lambda args: unraisable.append(args.exc_value))

        response = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": selected_iso, "end": selected_iso},
        )

        assert response.status_code == 422
        assert response.json()["detail"]["code"] == "EXPORT_TOO_LARGE"
        gc.collect()
        assert unraisable == []


def test_daily_payables_workbook_is_materialized_to_a_bounded_temporary_file():
    snapshot = {
        "date": date.today().isoformat(),
        "totals_cny": {"due_today": 0, "paid_today": 0, "end_pending": 0, "overdue_pending": 0},
        "currency_totals": [],
        "counts": {"due_today": 0, "end_pending": 0, "overdue_pending": 0},
        "items": [],
    }

    path = daily_payables_export.export_daily_payables_workbook([snapshot])
    try:
        assert isinstance(path, Path)
        assert path.exists()
        assert path.suffix == ".xlsx"
        assert daily_payables_export.MAX_EXPORT_DETAIL_ROWS < daily_payables_export.EXCEL_MAX_ROWS
        assert load_workbook(path, read_only=True).sheetnames == ["每日汇总", "逐日明细"]
    finally:
        if isinstance(path, Path):
            path.unlink(missing_ok=True)


def test_daily_payables_export_rejects_concurrent_over_capacity(monkeypatch):
    class BusyExportSlots:
        def acquire(self, *, blocking: bool) -> bool:
            assert blocking is False
            return False

        def release(self) -> None:
            raise AssertionError("A slot that was not acquired must not be released")

    monkeypatch.setattr(main_module, "_DAILY_PAYABLES_EXPORT_SLOTS", BusyExportSlots())
    with TestClient(app) as client:
        login(client)
        selected = date.today().isoformat()
        response = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": selected, "end": selected},
        )

    assert response.status_code == 429
    assert response.json()["detail"]["code"] == "EXPORT_BUSY"


def test_daily_payables_export_cleans_files_and_slots_after_invalid_range():
    temp_directory = Path(tempfile.gettempdir())
    before = set(temp_directory.glob("daily-payables-*.xlsx"))
    selected = date.today().isoformat()
    with TestClient(app) as client:
        login(client)
        for _ in range(daily_payables_export.MAX_CONCURRENT_EXPORTS):
            response = client.get(
                "/api/daily-payables/export.xlsx",
                params={"start": selected, "end": selected},
                headers={"Range": "bytes=invalid"},
            )
            assert response.status_code == 400

        normal = client.get(
            "/api/daily-payables/export.xlsx",
            params={"start": selected, "end": selected},
        )

    assert normal.status_code == 200, normal.text
    assert set(temp_directory.glob("daily-payables-*.xlsx")) == before


def test_daily_payables_export_cleans_files_and_slots_when_response_send_fails(monkeypatch):
    class TrackingExportSlots:
        def __init__(self) -> None:
            self.acquired = False
            self.release_count = 0

        def acquire(self, *, blocking: bool) -> bool:
            assert blocking is False
            self.acquired = True
            return True

        def release(self) -> None:
            self.release_count += 1

    slots = TrackingExportSlots()
    monkeypatch.setattr(main_module, "_DAILY_PAYABLES_EXPORT_SLOTS", slots)
    selected = date.today()
    response = main_module.export_daily_payables(
        start=selected,
        end=selected,
        user={"id": 1, "role": "admin"},
    )
    output_path = Path(response.path)
    assert output_path.exists()

    async def receive() -> dict:
        return {"type": "http.disconnect"}

    async def failing_send(_: dict) -> None:
        raise RuntimeError("simulated client disconnect")

    scope = {
        "type": "http",
        "method": "GET",
        "path": "/api/daily-payables/export.xlsx",
        "headers": [],
        "extensions": {},
    }
    try:
        asyncio.run(response(scope, receive, failing_send))
    except RuntimeError as exc:
        assert str(exc) == "simulated client disconnect"
    else:
        raise AssertionError("The simulated send failure must propagate")

    response.cleanup()
    assert not output_path.exists()
    assert slots.acquired is True
    assert slots.release_count == 1


def test_daily_payables_workbook_reports_invalid_historical_dates_and_cleans_temp_files():
    snapshot = {
        "date": date.today().isoformat(),
        "totals_cny": {"due_today": 1, "paid_today": 0, "end_pending": 1, "overdue_pending": 0},
        "currency_totals": [],
        "counts": {"due_today": 1, "end_pending": 1, "overdue_pending": 0},
        "items": [{"needed_payment_date": "invalid-date"}],
    }
    temp_directory = Path(tempfile.gettempdir())
    before = set(temp_directory.glob("daily-payables-*.xlsx"))

    try:
        daily_payables_export.export_daily_payables_workbook([snapshot])
    except Exception as exc:
        assert getattr(exc, "code", None) == "INVALID_EXPORT_DATA"
    else:
        raise AssertionError("Invalid historical dates must not produce a workbook")

    assert set(temp_directory.glob("daily-payables-*.xlsx")) == before


def test_daily_trend_does_not_apply_future_dingtalk_identity_to_past():
    with TestClient(app) as client:
        login(client)
        first_day = date.today()
        second_day = first_day + timedelta(days=1)
        due_date = (first_day - timedelta(days=1)).isoformat()
        source_sheet = f"每日应付未来身份-{uuid.uuid4().hex}"
        older = create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=due_date,
        )
        newer = create_request(
            client,
            amount=200,
            source_sheet=source_sheet,
            needed_payment_date=due_date,
        )
        dingding_id = f"daily-future-{uuid.uuid4().hex}"

        with connect() as conn:
            for sequence, request in enumerate((older, newer), start=10):
                conn.execute(
                    "UPDATE payment_requests SET dingding_id = ? WHERE id = ?",
                    (dingding_id, request["id"]),
                )
                record_request_state(
                    conn,
                    request["id"],
                    event_type="dingtalk.linked",
                    event_key=f"test:{uuid.uuid4().hex}",
                    effective_at=f"{second_day.isoformat()}T{sequence}:00:00.000000",
                    actor_id=1,
                )
            trend = daily_trend(
                conn,
                first_day,
                second_day,
                allowed_sheets={source_sheet},
                china_only=True,
            )

        assert trend["points"][0]["totals_cny"]["end_pending"] == 300
        assert trend["points"][0]["counts"]["end_pending"] == 2
        assert trend["points"][1]["totals_cny"]["end_pending"] == 200
        assert trend["points"][1]["counts"]["end_pending"] == 1


def test_daily_payables_treats_cleared_dingtalk_ids_as_independent():
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        due_date = (selected - timedelta(days=1)).isoformat()
        source_sheet = f"每日应付清空身份-{uuid.uuid4().hex}"
        dingding_id = f"daily-cleared-{uuid.uuid4().hex}"
        requests = [
            create_request(
                client,
                amount=amount,
                source_sheet=source_sheet,
                needed_payment_date=due_date,
                dingding_id=dingding_id,
            )
            for amount in (100, 200)
        ]

        with connect() as conn:
            for sequence, request in enumerate(requests, start=20):
                conn.execute(
                    "UPDATE payment_requests SET dingding_id = NULL WHERE id = ?",
                    (request["id"],),
                )
                record_request_state(
                    conn,
                    request["id"],
                    event_type="request.update",
                    event_key=f"test:{uuid.uuid4().hex}",
                    effective_at=f"{selected.isoformat()}T{sequence}:00:00.000000",
                    actor_id=1,
                )
            snapshot = daily_snapshot(
                conn,
                selected,
                allowed_sheets={source_sheet},
                include_details=True,
                china_only=True,
            )

        assert snapshot["totals_cny"]["end_pending"] == 300
        assert snapshot["counts"]["end_pending"] == 2
        assert {item["logical_request_id"] for item in snapshot["items"]} == {
            request["logical_request_id"] for request in requests
        }


def test_daily_payables_uses_only_winning_logical_payment_delta():
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        due_date = (selected - timedelta(days=1)).isoformat()
        source_sheet = f"每日应付付款去重-{uuid.uuid4().hex}"
        dingding_id = f"daily-payment-{uuid.uuid4().hex}"
        older = create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=due_date,
            dingding_id=dingding_id,
        )
        newer = create_request(
            client,
            amount=100,
            source_sheet=source_sheet,
            needed_payment_date=due_date,
            dingding_id=dingding_id,
        )

        with connect() as conn:
            for request, reset_hour, payment_hour in (
                (older, 20, 21),
                (newer, 22, 23),
            ):
                conn.execute(
                    "UPDATE payment_requests SET paid_amount = 0, pending_amount = 100 WHERE id = ?",
                    (request["id"],),
                )
                record_request_state(
                    conn,
                    request["id"],
                    event_type="request.update",
                    event_key=f"test:{uuid.uuid4().hex}",
                    effective_at=f"{selected.isoformat()}T{reset_hour}:00:00.000000",
                    actor_id=1,
                )
                conn.execute(
                    "UPDATE payment_requests SET paid_amount = 50, pending_amount = 50 WHERE id = ?",
                    (request["id"],),
                )
                record_request_state(
                    conn,
                    request["id"],
                    event_type="payment.create",
                    event_key=f"test:{uuid.uuid4().hex}",
                    effective_at=f"{selected.isoformat()}T{payment_hour}:00:00.000000",
                    actor_id=1,
                )
            snapshot = daily_snapshot(
                conn,
                selected,
                allowed_sheets={source_sheet},
                include_details=True,
                china_only=True,
            )

        assert snapshot["counts"]["end_pending"] == 1
        assert snapshot["totals_cny"]["end_pending"] == 50
        assert snapshot["totals_cny"]["paid_today"] == 50
        assert len(snapshot["items"]) == 1
        assert snapshot["items"][0]["logical_request_id"] == newer["logical_request_id"]
        assert snapshot["items"][0]["paid_today"] == 50


def test_daily_payables_summary_details_trend_and_sheet_permissions():
    with TestClient(app) as client:
        login(client)
        with connect() as conn:
            conn.execute(
                """
                INSERT INTO app_settings (key, value, updated_at)
                VALUES ('daily_payables_history_start_date', '2026-08-21', ?)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    updated_at = excluded.updated_at
                """,
                (now_iso(),),
            )
        sheet_name = f"每日应付授权-{uuid.uuid4().hex}"
        hidden_sheet = f"每日应付无权-{uuid.uuid4().hex}"
        request = create_request(client, source_sheet=sheet_name)
        hidden_request = create_request(client, amount=99999, source_sheet=hidden_sheet)
        no_due_request = create_request(
            client,
            amount=88888,
            source_sheet=sheet_name,
            needed_payment_date=None,
        )
        usd_request = create_request(
            client,
            amount=100,
            source_sheet=sheet_name,
        )
        reopened_request = create_request(
            client,
            amount=300,
            source_sheet=sheet_name,
        )
        with connect() as conn:
            conn.execute(
                """
                UPDATE payment_requests
                SET currency = 'USD', base_amount_cny = 680, fx_rate_cny_per_unit = 6.8,
                    fx_rate_date = '2026-08-21', fx_rate_actual_date = '2026-08-21'
                WHERE id = ?
                """,
                (usd_request["id"],),
            )
            record_request_state(
                conn,
                usd_request["id"],
                event_type="test.currency",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at="2026-08-21T09:00:00.000000",
                actor_id=1,
            )
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {"approval_status": "TERMINATED"}}), reopened_request["id"]),
            )
            record_request_state(
                conn,
                reopened_request["id"],
                event_type="dingtalk.terminated",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at="2026-08-21T20:00:00.000000",
                actor_id=1,
            )
            conn.execute(
                "UPDATE payment_requests SET raw_extra_json = ? WHERE id = ?",
                (json.dumps({"external_source": {"approval_status": "RUNNING"}}), reopened_request["id"]),
            )
            record_request_state(
                conn,
                reopened_request["id"],
                event_type="dingtalk.reopened",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at="2026-08-22T10:00:00.000000",
                actor_id=1,
            )

        payment_response = client.post(
            f"/api/batches/{request['batch_id']}/requests/{request['id']}/payments",
            json={
                "amount": 10000,
                "payment_date": "2026-08-21",
                "payer": "测试付款人",
                "expected_request_version": request["version"],
            },
        )
        assert payment_response.status_code == 200, payment_response.text

        source_batch = next(
            batch
            for batch in client.get("/api/batches").json()["batches"]
            if batch["id"] == request["batch_id"]
        )
        rollover = client.post(
            f"/api/batches/{request['batch_id']}/rollover",
            json={
                "name": f"每日应付结转-{uuid.uuid4().hex}",
                "start_date": "2026-08-22",
                "end_date": "2026-08-28",
                "copy_mode": "all",
                "expected_batch_version": source_batch["version"],
            },
        )
        assert rollover.status_code == 200, rollover.text

        with connect() as conn:
            conn.execute(
                "UPDATE payment_requests SET paid_amount = 20000, pending_amount = 0 WHERE id = ?",
                (request["id"],),
            )
            record_request_state(
                conn,
                request["id"],
                event_type="payment.create",
                event_key=f"test:{uuid.uuid4().hex}",
                effective_at="2026-08-22T11:00:00.000000",
                actor_id=1,
            )
            conn.execute(
                "UPDATE payment_requests SET paid_amount = 10000, pending_amount = 10000 WHERE id = ?",
                (request["id"],),
            )

        username = f"daily-{uuid.uuid4().hex}"
        with connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO users (username, password_hash, role, display_name, active, created_at)
                VALUES (?, ?, 'business', ?, 1, ?)
                """,
                (username, hash_password("daily123"), username, now_iso()),
            )
            user_id = int(cursor.lastrowid)
            conn.execute(
                """
                INSERT INTO user_sheet_permissions (user_id, sheet_name, created_by, created_at)
                VALUES (?, ?, 1, ?)
                """,
                (user_id, sheet_name, now_iso()),
            )
            history_start = get_daily_payables_history_start_date(conn)

        client.post("/api/auth/logout")
        response = client.post(
            "/api/auth/login",
            json={"username": username, "password": "daily123"},
        )
        assert response.status_code == 200

        summary_response = client.get("/api/daily-payables/summary", params={"date": "2026-08-21"})
        assert summary_response.status_code == 200, summary_response.text
        summary = summary_response.json()
        assert summary["history_start_date"] == history_start
        assert summary["totals_cny"]["due_today"] == 20680
        assert summary["totals_cny"]["paid_today"] == 10000
        assert summary["totals_cny"]["end_pending"] == 10680
        currencies = {item["currency"]: item for item in summary["currency_totals"]}
        assert currencies["CNY"]["end_pending"] == 10000
        assert currencies["USD"]["end_pending"] == 100

        details_response = client.get("/api/daily-payables/details", params={"date": "2026-08-21"})
        assert details_response.status_code == 200, details_response.text
        details = details_response.json()["items"]
        logical_ids = [item["logical_request_id"] for item in details]
        assert logical_ids.count(request["logical_request_id"]) == 1
        assert hidden_request["logical_request_id"] not in logical_ids
        assert no_due_request["logical_request_id"] not in logical_ids
        assert reopened_request["logical_request_id"] not in logical_ids
        assert next(item for item in details if item["logical_request_id"] == request["logical_request_id"])[
            "pending_amount"
        ] == 10000

        trend_response = client.get(
            "/api/daily-payables/trend",
            params={"start": "2026-08-21", "end": "2026-08-22"},
        )
        assert trend_response.status_code == 200, trend_response.text
        trend = trend_response.json()["points"]
        assert trend[0]["date"] == "2026-08-21"
        assert trend[0]["totals_cny"]["end_pending"] == 10680
        assert trend[1]["totals_cny"]["paid_today"] == 10000
        assert trend[1]["totals_cny"]["overdue_pending"] == 980

        next_day_details = client.get(
            "/api/daily-payables/details", params={"date": "2026-08-22"}
        ).json()["items"]
        assert reopened_request["logical_request_id"] in {
            item["logical_request_id"] for item in next_day_details
        }


def test_daily_payables_rejects_dates_before_baseline_and_long_trends():
    with TestClient(app) as client:
        login(client)
        with connect() as conn:
            history_start = get_daily_payables_history_start_date(conn)
        earlier = (date.fromisoformat(history_start) - timedelta(days=1)).isoformat()
        response = client.get("/api/daily-payables/summary", params={"date": earlier})
        assert response.status_code == 422
        assert response.json()["detail"]["code"] == "HISTORY_NOT_AVAILABLE"

        response = client.get(
            "/api/daily-payables/trend",
            params={"start": history_start, "end": (date.fromisoformat(history_start) + timedelta(days=93)).isoformat()},
        )
        assert response.status_code == 422
        assert response.json()["detail"]["code"] == "TREND_RANGE_TOO_LARGE"


def test_daily_payables_can_exclude_mexico_and_region_review_history():
    with TestClient(app) as client:
        login(client)
        selected = date.today()
        selected_iso = selected.isoformat()
        china_sheet = f"中国地区-{uuid.uuid4().hex}"
        mexico_sheet = f"墨西哥地区-{uuid.uuid4().hex}"
        review_sheet = f"地区待核对-{uuid.uuid4().hex}"
        china = create_request(
            client,
            amount=100,
            source_sheet=china_sheet,
            needed_payment_date=selected_iso,
        )
        mexico = create_request(
            client,
            amount=200,
            source_sheet=mexico_sheet,
            needed_payment_date=selected_iso,
            execution_region="Mexico",
        )
        review = create_request(
            client,
            amount=300,
            source_sheet=review_sheet,
            needed_payment_date=selected_iso,
            execution_region=None,
        )

        with connect() as conn:
            all_regions = daily_snapshot(
                conn,
                selected,
                include_details=True,
                china_only=False,
                allowed_sheets={china_sheet, mexico_sheet, review_sheet},
            )
            china_only = daily_snapshot(
                conn,
                selected,
                include_details=True,
                china_only=True,
                allowed_sheets={china_sheet, mexico_sheet, review_sheet},
            )

        assert all_regions["totals_cny"]["due_today"] == 600
        assert china_only["totals_cny"]["due_today"] == 100
        assert [item["logical_request_id"] for item in china_only["items"]] == [
            china["logical_request_id"]
        ]
        assert mexico["logical_request_id"] not in {
            item["logical_request_id"] for item in china_only["items"]
        }
        assert review["logical_request_id"] not in {
            item["logical_request_id"] for item in china_only["items"]
        }
