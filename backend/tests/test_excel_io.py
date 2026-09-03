import sys
import io
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook

from backend.app.excel_io import export_workbook, normalize_request_business_fields, parse_weekly_excel, safe_sheet_title


SAMPLE = Path("/Users/smk/Downloads/20260626~20260707请款明细.xlsx")


def test_weekly_excel_imports_known_workbook():
    if not SAMPLE.exists():
        return
    rows, meta = parse_weekly_excel(SAMPLE)
    assert len(rows) == 161
    sheet_counts = {item["sheet"]: item.get("imported", 0) for item in meta["sheets"]}
    assert sheet_counts["汇总7.7可支付"] == 86
    assert sheet_counts["赣瑞模具（7月前）"] == 7
    assert sheet_counts["赣瑞模具（7月后）"] == 1
    assert sheet_counts["志威模具（7月前）"] == 10
    assert sheet_counts["志威模具（7月后）"] == 1
    assert sheet_counts["YUEWEI自研产品+OEM项目款"] == 43
    assert sheet_counts["物流"] == 5
    assert sheet_counts["员工报销"] == 1
    assert sheet_counts["员工成长金+待定付款"] == 7
    first = rows[0]
    assert first["dingding_id"]
    assert first["amount"] is not None
    assert first["content_hash"]
    assert {row.get("finance_review") for row in rows} <= {"未付款", "部分付款", "已付款"}
    assert {row.get("payment_status") for row in rows} <= {None}
    assert {row.get("general_manager_approval") for row in rows if row.get("general_manager_approval")} <= {"同意付款", "延缓批付", "存在争议"}
    paid_remark_row = next(row for row in rows if row["source_sheet"] == "YUEWEI自研产品+OEM项目款" and row["source_row"] == 24)
    assert paid_remark_row["finance_review"] == "未付款"
    assert "已支付10万" in paid_remark_row["remark"]
    assert "Tiffany垫付" in paid_remark_row["remark"]
    assert paid_remark_row["actual_payment_date"] is None
    finance_paid_row = next(row for row in rows if row["source_sheet"] == "志威模具（7月后）" and row["source_row"] == 3)
    assert finance_paid_row["finance_review"] == "已付款"
    assert "已经支付14500元" in finance_paid_row["remark"]
    payment_status_note_row = next(row for row in rows if row["source_sheet"] == "物流" and row["source_row"] == 2)
    assert payment_status_note_row["finance_review"] == "未付款"
    assert "原付款情况：同意支付" in payment_status_note_row["remark"]
    manager_note_row = next(row for row in rows if row["source_sheet"] == "员工报销" and row["source_row"] == 2)
    assert manager_note_row["general_manager_approval"] is None
    assert "我要审核一下" in manager_note_row["general_manager_opinion"]
    assert manager_note_row["remark"] is None
    manager_plan_row = next(row for row in rows if row["source_sheet"] == "志威模具（7月前）" and row["source_row"] == 5)
    assert manager_plan_row["general_manager_approval"] is None
    assert manager_plan_row["general_manager_opinion"] == "对完账已确认付款计划"
    assert "对完账已确认付款计划" not in manager_plan_row["remark"]
    dispute_row = next(row for row in rows if row["source_sheet"] == "员工成长金+待定付款" and row["source_row"] == 9)
    assert dispute_row["general_manager_approval"] == "存在争议"
    assert "CMA船司" in dispute_row["general_manager_opinion"]
    assert "CMA船司" not in dispute_row["remark"]
    advance_note_row = next(row for row in rows if row["source_sheet"] == "员工成长金+待定付款" and row["source_row"] == 8)
    assert advance_note_row["finance_review"] == "未付款"
    assert advance_note_row["general_manager_approval"] is None
    assert "Tiffany垫付" in advance_note_row["general_manager_opinion"]
    assert "Tiffany垫付" not in advance_note_row["remark"]
    assert not any(row.get("finance_review") == "2026/5/8付" for row in rows)


def test_partial_finance_review_is_not_normalized_as_paid():
    row = {
        "finance_review": "部分付款",
        "actual_payment_date": "2026-07-10",
        "payment_status": "",
        "remark": "",
    }
    normalize_request_business_fields(row)
    assert row["finance_review"] == "部分付款"
    assert row["payment_status"] is None

    status_row = {
        "finance_review": "",
        "actual_payment_date": "",
        "payment_status": "部分付款",
        "remark": "",
    }
    normalize_request_business_fields(status_row)
    assert status_row["finance_review"] == "部分付款"
    assert status_row["payment_status"] is None


def test_partial_payment_amounts_are_normalized_and_exported():
    row = {
        "amount": 100,
        "paid_amount": 35.5,
        "pending_amount": 999,
        "finance_review": "未付款",
        "payment_status": "",
        "remark": "",
    }
    normalize_request_business_fields(row)
    assert row["paid_amount"] == 35.5
    assert row["pending_amount"] == 64.5
    assert row["finance_review"] == "部分付款"

    content = export_workbook(
        {"name": "部分付款测试", "end_date": "2026-07-07"},
        [{"id": 1, "source_sheet": "手工录入", **row}],
        {},
    )
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    worksheet = workbook["全部"]
    headers = [cell.value for cell in worksheet[2]]
    assert "申请人" in headers
    amount_columns = [headers.index(header) + 1 for header in ("应付金额", "已支付金额", "待付款金额")]
    assert [worksheet.cell(3, column).value for column in amount_columns] == [100, 35.5, 64.5]
    assert all(str(worksheet.cell(4, column).value).startswith("=SUM(") for column in amount_columns)


def test_weekly_excel_extracts_embedded_images():
    if not SAMPLE.exists():
        return
    rows, meta = parse_weekly_excel(SAMPLE)
    embedded_images = [image for row in rows for image in row.get("_embedded_images", [])]
    assert meta["images"] == {"found": 9, "attached": 9, "skipped": 0}
    assert len(embedded_images) == 9
    assert all(image["data"] and image["label"].startswith("Excel图片 ") for image in embedded_images)
    image_rows = {(row["source_sheet"], row["source_row"]) for row in rows if row.get("_embedded_images")}
    assert ("汇总7.7可支付", 3) in image_rows
    assert ("志威模具（7月前）", 7) in image_rows
    assert ("员工成长金+待定付款", 10) in image_rows


def test_export_workbook_roundtrip_bytes():
    rows, _ = parse_weekly_excel(SAMPLE) if SAMPLE.exists() else ([], {})
    content = export_workbook(
        {"name": "测试批次", "end_date": "2026-07-07"},
        [{"id": idx + 1, **row} for idx, row in enumerate(rows[:5])],
        {},
    )
    assert content.startswith(b"PK")
    assert len(content) > 1000
    workbook = load_workbook(io.BytesIO(content))
    headers = [cell.value for worksheet in workbook.worksheets for row in worksheet.iter_rows(max_row=2) for cell in row]
    assert "付款情况" not in headers
    assert "已支付金额" in headers
    assert "待付款金额" in headers


def test_expected_payment_account_excel_roundtrip(tmp_path):
    content = export_workbook(
        {"id": 9, "name": "预计支付账户导出", "end_date": "2026-09-02"},
        [{
            "id": 1,
            "dingding_id": "EXPECTED-EXCEL-1",
            "source_sheet": "悦为智能",
            "payment_account": "公户",
            "expected_payment_account": "悦为智能 6221 主账户",
            "amount": 100,
            "summary": "预计支付账户 Excel 往返",
        }],
        {},
    )
    exported = tmp_path / "expected-payment-account.xlsx"
    exported.write_bytes(content)

    workbook = load_workbook(exported, data_only=False)
    for sheet_name, header_row in (("全部", 2), ("悦为智能", 1)):
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[header_row]]
        payment_account_index = headers.index("付款账户")
        assert headers[payment_account_index + 1] == "预计支付账户"
        expected_column = headers.index("预计支付账户") + 1
        assert worksheet.cell(header_row + 1, expected_column).value == "悦为智能 6221 主账户"

    rows, _meta = parse_weekly_excel(exported)
    assert len(rows) == 1
    assert rows[0]["expected_payment_account"] == "悦为智能 6221 主账户"
    assert "expected_payment_account" in rows[0]["_present_fields"]


def test_department_sheet_titles_are_safe_and_unique():
    used: set[str] = set()
    first = safe_sheet_title("产品/采购:中国区*2026?very-long-department-name", used)
    second = safe_sheet_title("产品\\采购:中国区*2026?very-long-department-name", used)
    assert len(first) <= 31
    assert len(second) <= 31
    assert not any(character in first + second for character in "[]:*?/\\")
    assert first.casefold() != second.casefold()

    content = export_workbook(
        {"name": "部门 Sheet 测试", "end_date": "2026-07-15"},
        [
            {"id": 1, "source_sheet": "付款明细", "summary": "部门与标准付款明细页同名"},
            {"id": 2, "source_sheet": "A/B", "summary": "非法字符一"},
            {"id": 3, "source_sheet": "A\\B", "summary": "非法字符二"},
        ],
        {},
    )
    workbook = load_workbook(io.BytesIO(content))
    assert len(workbook.sheetnames) == len({name.casefold() for name in workbook.sheetnames})
    assert all(len(name) <= 31 for name in workbook.sheetnames)


def test_export_workbook_uses_saved_sheet_order():
    content = export_workbook(
        {
            "name": "Sheet 顺序测试",
            "end_date": "2026-07-21",
            "sheet_order": ["供应商", "财务中心", "采购中心"],
        },
        [
            {"id": 1, "source_sheet": "采购中心", "amount": 100, "summary": "采购"},
            {"id": 2, "source_sheet": "供应商", "amount": 200, "summary": "供应商"},
            {"id": 3, "source_sheet": "财务中心", "amount": 300, "summary": "财务"},
        ],
        {},
    )
    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames[:4] == ["全部", "供应商", "财务中心", "采购中心"]
    assert workbook.sheetnames[-2:] == ["付款明细", "_系统信息"]
    assert workbook["_系统信息"].sheet_state == "veryHidden"


def test_all_sheet_is_first_and_skipped_during_roundtrip_parse(tmp_path):
    records = [
        {"id": 1, "dingding_id": "DD-1", "source_sheet": "采购中心", "amount": 100, "summary": "采购"},
        {"id": 2, "dingding_id": "DD-2", "source_sheet": "财务中心", "amount": 200, "summary": "财务"},
    ]
    content = export_workbook(
        {
            "id": 7,
            "name": "全部 Sheet 测试",
            "end_date": "2026-07-21",
            "sheet_order": ["财务中心", "采购中心"],
        },
        records,
        {},
    )
    exported = tmp_path / "all-sheet-roundtrip.xlsx"
    exported.write_bytes(content)

    workbook = load_workbook(exported, data_only=False)
    assert workbook.sheetnames[:3] == ["全部", "财务中心", "采购中心"]
    assert workbook["全部"]["A1"].value.startswith("全部记录汇总视图")

    rows, meta = parse_weekly_excel(exported)
    assert len(rows) == len(records)
    assert {row["_request_id"] for row in rows} == {1, 2}
    assert meta["workbook_sheet_order"] == ["财务中心", "采购中心"]
    assert any(item.get("sheet") == "全部" and item.get("skipped") == "view_sheet" for item in meta["sheets"])
